"""
API 测试
验证 Excel 导入导出和自动填充功能
"""
import pytest
import io
import openpyxl
from openpyxl import Workbook


class TestExcelExport:
    """Excel 导出测试"""

    def test_export_excel_success(self, client, sample_data):
        """测试 Excel 导出成功"""
        from models import db, Locomotive

        with client.application.app_context():
            # 添加一个机车以便有数据导出
            loco = Locomotive(
                series_id=1,
                power_type_id=1,
                model_id=1,
                brand_id=1,
                scale='HO',
                locomotive_number='0001'
            )
            db.session.add(loco)
            db.session.commit()

        response = client.get('/api/export/excel')
        assert response.status_code == 200
        assert 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in response.content_type

    def test_export_excel_empty(self, client):
        """测试空数据库导出"""
        response = client.get('/api/export/excel')
        # 空数据库应该返回错误
        assert response.status_code == 400

    def test_export_locomotive_head_no_depot(self, client, sample_data):
        """测试先头车导出不包含动车段列"""
        from models import db, LocomotiveHead

        with client.application.app_context():
            head = LocomotiveHead(
                model_id=1,
                brand_id=1,
                scale='HO',
                head_light=True,
                special_color='红色'
            )
            db.session.add(head)
            db.session.commit()

        response = client.get('/api/export/excel')
        assert response.status_code == 200

        # 解析 Excel 检查列头
        wb = openpyxl.load_workbook(io.BytesIO(response.data))
        if '先头车' in wb.sheetnames:
            sheet = wb['先头车']
            headers = [cell.value for cell in sheet[1]]
            # 确保没有动车段
            assert '动车段' not in headers
            # 确保有涂装（原特涂）
            assert '涂装' in headers


class TestExcelImport:
    """Excel 导入测试"""

    def test_import_locomotive_head_no_depot(self, client, sample_data):
        """测试先头车导入（无动车段）"""
        # 创建测试 Excel
        wb = Workbook()
        ws = wb.create_sheet('先头车')
        ws.append(['车型', '品牌', '涂装', '比例', '头车灯', '室内灯', '价格', '总价', '货号', '购买日期', '购买商家'])
        ws.append(['CRH380A', '测试品牌', '红色', 'HO', '是', 'LED', '100', '100', '001', '', ''])

        # 移除默认 sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # 保存到内存
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = client.post(
            '/api/import/excel',
            data={'file': (excel_file, 'test.xlsx')},
            content_type='multipart/form-data'
        )

        assert response.status_code == 200
        data = response.get_json()
        assert data.get('success') == True


class TestAutoFillAPI:
    """自动填充 API 测试"""

    def test_auto_fill_locomotive(self, client, sample_data):
        """测试机车自动填充"""
        response = client.get('/api/auto-fill/locomotive/1')
        assert response.status_code == 200
        data = response.get_json()
        assert 'series_id' in data
        assert 'power_type_id' in data

    def test_auto_fill_trainset(self, client, sample_data):
        """测试动车组自动填充"""
        response = client.get('/api/auto-fill/trainset/1')
        assert response.status_code == 200
        data = response.get_json()
        assert 'series_id' in data
        assert 'power_type_id' in data


class TestStatisticsAPI:
    """统计 API 测试"""

    def test_statistics_endpoint(self, client):
        """测试统计 API"""
        response = client.get('/api/statistics')
        assert response.status_code == 200
        data = response.get_json()
        assert 'type_stats' in data
        assert 'scale_stats' in data
        assert 'brand_stats' in data
        assert 'merchant_stats' in data


class TestLightModelAPI:
    """灯型号兼容查询 API 测试"""

    def test_get_all_light_models(self, client, sample_data):
        """测试获取所有灯型号"""
        response = client.get('/api/light-models/all')
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert 'groups' in data
        # 应至少有一个灯品牌分组
        assert len(data['groups']) >= 1
        # 每个分组应包含 light_brand_name 和 models
        for group in data['groups']:
            assert 'light_brand_name' in group
            assert 'models' in group
            assert 'light_brand_id' in group

    def test_compatible_light_models_carriage(self, client, sample_data):
        """测试查询车厢兼容灯型号"""
        with client.application.app_context():
            from models import CarriageModel, Brand
            cm = CarriageModel.query.first()
            br = Brand.query.first()
            cm_id = cm.id if cm else 1
            br_id = br.id if br else 1

        response = client.get(f'/api/light-models/compatible?model_type=carriage&model_id={cm_id}&brand_id={br_id}')
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert 'groups' in data

    def test_compatible_light_models_trainset(self, client, sample_data):
        """测试查询动车组兼容灯型号"""
        with client.application.app_context():
            from models import TrainsetModel, Brand
            tm = TrainsetModel.query.first()
            br = Brand.query.first()
            tm_id = tm.id if tm else 1
            br_id = br.id if br else 1

        response = client.get(f'/api/light-models/compatible?model_type=trainset&model_id={tm_id}&brand_id={br_id}')
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert 'groups' in data

    def test_compatible_invalid_model_type(self, client, sample_data):
        """测试无效 model_type 返回 400"""
        response = client.get('/api/light-models/compatible?model_type=invalid&model_id=1&brand_id=1')
        assert response.status_code == 400
        data = response.get_json()
        assert data['success'] is False

    def test_compatible_missing_params(self, client, sample_data):
        """测试缺少必填参数返回 400"""
        response = client.get('/api/light-models/compatible?model_type=carriage')
        assert response.status_code == 400

    def test_compatible_with_scale_filter(self, client, sample_data):
        """测试 scale 过滤功能"""
        with client.application.app_context():
            from models import CarriageModel, Brand
            cm = CarriageModel.query.first()
            br = Brand.query.first()
            cm_id = cm.id if cm else 1
            br_id = br.id if br else 1

        # 查询 HO 比例
        response = client.get(
            f'/api/light-models/compatible?model_type=carriage&model_id={cm_id}&brand_id={br_id}&scale=HO')
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        # 所有返回的灯型号 scale 应为 HO
        for group in data['groups']:
            for model in group['models']:
                assert model['scale'] == 'HO'

    def test_compatible_wrong_scale_excluded(self, client, sample_data):
        """测试 N 比例灯不出现在 HO 查询中"""
        with client.application.app_context():
            from models import CarriageModel, Brand
            cm = CarriageModel.query.first()
            br = Brand.query.first()
            cm_id = cm.id if cm else 1
            br_id = br.id if br else 1

        response = client.get(
            f'/api/light-models/compatible?model_type=carriage&model_id={cm_id}&brand_id={br_id}&scale=HO')
        data = response.get_json()
        for group in data['groups']:
            for model in group['models']:
                assert model['scale'] != 'N'

    def test_all_light_models_include_scale(self, client, sample_data):
        """测试所有灯型号响应包含 scale 字段"""
        response = client.get('/api/light-models/all')
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        for group in data['groups']:
            for model in group['models']:
                assert 'scale' in model

    def test_compatible_brand_only_mode(self, client, sample_data):
        """测试 brand_only 模式：只按品牌查询，不需要 model_id"""
        with client.application.app_context():
            from models import Brand
            brand = Brand.query.first()
            br_id = brand.id if brand else 1

        # brand_only 模式只需要 brand_id
        response = client.get(f'/api/light-models/compatible?model_type=brand_only&brand_id={br_id}')
        assert response.status_code == 200
        data = response.get_json()
        assert data['success'] is True
        assert 'groups' in data

    def test_compatible_brand_only_no_model_id(self, client, sample_data):
        """测试 brand_only 模式不传 model_id 不报错"""
        with client.application.app_context():
            from models import Brand
            brand = Brand.query.first()
            br_id = brand.id if brand else 1

        response = client.get(f'/api/light-models/compatible?model_type=brand_only&brand_id={br_id}')
        assert response.status_code == 200

    def test_compatible_brand_only_missing_brand(self, client, sample_data):
        """测试 brand_only 模式缺少 brand_id 返回 400"""
        response = client.get('/api/light-models/compatible?model_type=brand_only')
        assert response.status_code == 400

    def test_compatible_brand_only_ignores_model_id(self, client, sample_data):
        """测试 brand_only 模式下传入 model_id 不影响结果（只看品牌级规则）"""
        with client.application.app_context():
            from models import Brand
            brand = Brand.query.first()
            br_id = brand.id if brand else 1

        # 带 model_id 和不带 model_id 结果应相同
        r1 = client.get(f'/api/light-models/compatible?model_type=brand_only&brand_id={br_id}')
        r2 = client.get(f'/api/light-models/compatible?model_type=brand_only&brand_id={br_id}&model_id=999')
        d1 = r1.get_json()
        d2 = r2.get_json()
        assert d1['success'] and d2['success']
        assert len(d1['groups']) == len(d2['groups'])
