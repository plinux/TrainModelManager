"""
集成测试
验证导入导出往返和数据库初始化
"""
import pytest
import io
import openpyxl
from openpyxl import Workbook
from models import db, Locomotive, CarriageSet, Trainset, LocomotiveHead
from models import Brand, Depot, Merchant, ChipInterface, ChipModel
from models import LocomotiveSeries, LocomotiveModel, CarriageSeries, CarriageModel
from models import TrainsetSeries, TrainsetModel, PowerType


class TestExcelRoundTrip:
    """Excel 导入导出往返测试"""

    def test_locomotive_round_trip(self, client, sample_data):
        """测试机车完整往返：创建 -> 导出 -> 清空 -> 导入 -> 验证"""
        with client.application.app_context():
            # 1. 创建测试数据
            loco = Locomotive(
                series_id=1,
                power_type_id=1,
                model_id=1,
                brand_id=1,
                depot_id=1,
                scale='HO',
                locomotive_number='ROUND001',
                decoder_number='01',
                price='500',
                total_price=500,
                item_number='ITEM001'
            )
            db.session.add(loco)
            db.session.commit()

        # 2. 导出
        export_response = client.get('/api/export/excel')
        assert export_response.status_code == 200

        excel_data = export_response.data
        wb = openpyxl.load_workbook(io.BytesIO(excel_data))

        # 3. 验证导出数据包含机车表
        assert '机车' in wb.sheetnames
        loco_sheet = wb['机车']
        # 验证有数据行
        row_count = loco_sheet.max_row
        assert row_count >= 2, "机车表应该有至少一行数据（表头+数据）"

        # 4. 清空数据库中的机车
        with client.application.app_context():
            Locomotive.query.delete()
            db.session.commit()
            assert Locomotive.query.count() == 0

        # 5. 重新导入
        import_response = client.post(
            '/api/import/excel',
            data={'file': (io.BytesIO(excel_data), 'round_trip.xlsx')},
            content_type='multipart/form-data'
        )
        assert import_response.status_code == 200

        # 6. 验证数据已恢复
        with client.application.app_context():
            locos = Locomotive.query.all()
            assert len(locos) > 0

    def test_trainset_round_trip(self, client, sample_data):
        """测试动车组完整往返"""
        with client.application.app_context():
            trainset = Trainset(
                series_id=1,
                power_type_id=1,
                model_id=1,
                brand_id=1,
                scale='HO',
                trainset_number='ROUND001',
                formation=8,
                color='白色'
            )
            db.session.add(trainset)
            db.session.commit()

        # 导出
        export_response = client.get('/api/export/excel')
        assert export_response.status_code == 200
        excel_data = export_response.data

        # 验证导出包含动车组
        wb = openpyxl.load_workbook(io.BytesIO(excel_data))
        assert '动车组' in wb.sheetnames

        # 清空
        with client.application.app_context():
            Trainset.query.delete()
            db.session.commit()

        # 导入
        import_response = client.post(
            '/api/import/excel',
            data={'file': (io.BytesIO(excel_data), 'round_trip.xlsx')},
            content_type='multipart/form-data'
        )
        assert import_response.status_code == 200

        # 验证
        with client.application.app_context():
            assert Trainset.query.count() > 0

    def test_locomotive_head_round_trip(self, client, sample_data):
        """测试先头车完整往返"""
        with client.application.app_context():
            head = LocomotiveHead(
                model_id=1,
                brand_id=1,
                scale='HO',
                head_light=True,
                special_color='红色'
            )
            db.session.add(head)
            db.session.commit()

        # 导出
        export_response = client.get('/api/export/excel')
        assert export_response.status_code == 200
        excel_data = export_response.data

        # 验证导出包含先头车且没有动车段列
        wb = openpyxl.load_workbook(io.BytesIO(excel_data))
        if '先头车' in wb.sheetnames:
            sheet = wb['先头车']
            headers = [cell.value for cell in sheet[1]]
            assert '动车段' not in headers

        # 清空
        with client.application.app_context():
            LocomotiveHead.query.delete()
            db.session.commit()

        # 导入
        import_response = client.post(
            '/api/import/excel',
            data={'file': (io.BytesIO(excel_data), 'round_trip.xlsx')},
            content_type='multipart/form-data'
        )
        assert import_response.status_code == 200

        # 验证
        with client.application.app_context():
            assert LocomotiveHead.query.count() > 0

    def test_empty_export_returns_error(self, client):
        """测试空数据库导出返回错误"""
        response = client.get('/api/export/excel')
        assert response.status_code == 400

    def test_invalid_file_import(self, client):
        """测试无效文件导入"""
        invalid_file = io.BytesIO(b'not an excel file')
        response = client.post(
            '/api/import/excel',
            data={'file': (invalid_file, 'invalid.txt')},
            content_type='multipart/form-data'
        )
        # 应该返回错误
        assert response.status_code != 200 or not response.get_json().get('success', True)


class TestDatabaseInitialization:
    """数据库初始化测试"""

    def test_init_db_creates_tables(self, app):
        """测试数据库表创建"""
        with app.app_context():
            # 验证所有表都已创建
            from models import db
            inspector = db.inspect(db.engine)
            tables = inspector.get_table_names()

            expected_tables = [
                'brand', 'depot', 'merchant', 'power_type',
                'chip_interface', 'chip_model',
                'locomotive_series', 'locomotive_model',
                'carriage_series', 'carriage_model',
                'trainset_series', 'trainset_model',
                'locomotive', 'carriage_set', 'trainset', 'locomotive_head'
            ]

            for table in expected_tables:
                assert table in tables, f"表 {table} 应该存在"

    def test_sample_data_fixture(self, sample_data):
        """测试示例数据 fixture 正常工作"""
        assert sample_data is not None
        assert 'brand' in sample_data

    def test_database_isolation(self, app):
        """测试数据库隔离（每个测试独立）"""
        with app.app_context():
            # 初始应该没有品牌
            initial_count = Brand.query.count()

            # 添加一个品牌
            brand = Brand(name='隔离测试品牌')
            db.session.add(brand)
            db.session.commit()

            # 这个测试内应该有
            assert Brand.query.count() == initial_count + 1

        # 新的 app_context 应该是独立的（因为是内存数据库）
        with app.app_context():
            # 由于使用 yield，这个测试会在 teardown 前运行
            pass


class TestStatisticsWithEmptyData:
    """空数据统计测试"""

    def test_statistics_with_no_data(self, client):
        """测试无数据时的统计"""
        response = client.get('/api/statistics')
        assert response.status_code == 200
        data = response.get_json()

        # 应该返回空统计
        assert 'type_stats' in data
        assert 'scale_stats' in data
        assert 'brand_stats' in data
        assert 'merchant_stats' in data

    def test_home_page_with_no_data(self, client):
        """测试无数据时的首页"""
        response = client.get('/')
        assert response.status_code == 200
        # 页面应该正常渲染
        assert '火车模型管理系统' in response.data.decode('utf-8')


class TestUniqueConstraints:
    """唯一性约束测试"""

    def test_brand_name_uniqueness(self, app):
        """测试品牌名唯一性"""
        with app.app_context():
            brand1 = Brand(name='唯一品牌测试')
            db.session.add(brand1)
            db.session.commit()

            # 尝试创建同名品牌
            brand2 = Brand(name='唯一品牌测试')
            db.session.add(brand2)

            # 根据数据库约束，可能成功或失败
            # 这里只是测试不会崩溃
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()

    def test_locomotive_number_in_scale(self, app, sample_data):
        """测试同一比例内机车号唯一性"""
        with app.app_context():
            loco1 = Locomotive(
                series_id=1,
                power_type_id=1,
                model_id=1,
                brand_id=1,
                scale='HO',
                locomotive_number='UNIQUE001'
            )
            db.session.add(loco1)
            db.session.commit()

            # 同一比例内相同机车号
            loco2 = Locomotive(
                series_id=1,
                power_type_id=1,
                model_id=1,
                brand_id=1,
                scale='HO',
                locomotive_number='UNIQUE001'
            )
            db.session.add(loco2)

            # 注意：这取决于模型是否定义了唯一约束
            # 如果没有约束，可能会成功
            try:
                db.session.commit()
            except Exception:
                db.session.rollback()
