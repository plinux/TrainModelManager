"""
API 扩展测试
覆盖 routes/api.py 的主要未覆盖路径：
- 自动填充车厢 API
- Excel 导入（各种工作表类型和模式）
- Excel 导出（models/system/all 三种模式）
- 导入冲突检测
- 导入模板 CRUD
- 自定义导入（parse/preview/execute）
"""
import pytest
import io
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


def _make_excel(sheets):
  """创建测试用 Excel 文件，sheets = {sheet_name: [headers, [rows...]]}"""
  wb = Workbook()
  first = True
  for sheet_name, (headers, rows) in sheets.items():
    if first:
      ws = wb.active
      ws.title = sheet_name
      first = False
    else:
      ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    for row in rows:
      ws.append(row)
  buf = io.BytesIO()
  wb.save(buf)
  buf.seek(0)
  return buf


class TestAutoFillCarriage:
  """车厢车型自动填充测试"""

  def test_auto_fill_carriage_success(self, client, sample_data):
    """测试车厢车型自动填充成功"""
    response = client.get('/api/auto-fill/carriage/1')
    assert response.status_code == 200
    data = response.get_json()
    assert 'series_id' in data
    assert 'type' in data

  def test_auto_fill_carriage_not_found(self, client, sample_data):
    """测试车厢车型自动填充不存在返回 404"""
    response = client.get('/api/auto-fill/carriage/99999')
    assert response.status_code == 404


class TestExcelImportModes:
  """Excel 导入各种模式测试"""

  def test_import_no_file(self, client):
    """测试未选择文件"""
    response = client.post('/api/import/excel', content_type='multipart/form-data')
    data = response.get_json()
    assert data['success'] is False
    assert '未选择文件' in data['error']

  def test_import_empty_filename(self, client):
    """测试空文件名"""
    buf = io.BytesIO(b'')
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, '')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_import_wrong_format(self, client):
    """测试非 Excel 文件格式"""
    buf = io.BytesIO(b'not excel')
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.txt')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False
    assert 'Excel' in data['error']

  def test_import_empty_workbook(self, client):
    """测试空工作簿（无可导入数据的工作表）"""
    buf = _make_excel({
      '未识别的工作表': (['名称'], [['测试']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    # 未知工作表被跳过，summary 为空
    assert data.get('summary', {}) == {}

  def test_import_preview_mode(self, client, sample_data):
    """测试预览模式（只检查冲突不导入）"""
    buf = _make_excel({
      '品牌': (['名称', '搜索地址'], [['测试品牌', 'http://example.com']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'preview'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['preview'] is True
    assert 'conflicts' in data
    assert 'has_conflicts' in data

  def test_import_brand_sheet(self, client, sample_data):
    """测试导入品牌工作表"""
    buf = _make_excel({
      '品牌': (['名称', '缩写', '搜索地址'], [['新品牌', 'XP', 'http://new.com']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert '品牌' in data.get('summary', {})

    from models import db, Brand
    with client.application.app_context():
      brand = Brand.query.filter_by(name='新品牌').first()
      assert brand is not None
      assert brand.search_url == 'http://new.com'

  def test_import_brand_overwrite(self, client, sample_data):
    """测试覆盖已有品牌"""
    buf = _make_excel({
      '品牌': (['名称', '搜索地址'], [['测试品牌', 'http://updated.com']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'overwrite'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, Brand
    with client.application.app_context():
      brand = Brand.query.filter_by(name='测试品牌').first()
      assert brand is not None
      assert brand.search_url == 'http://updated.com'

  def test_import_depot_sheet(self, client, sample_data):
    """测试导入机务段工作表"""
    buf = _make_excel({
      '机务段': (['名称'], [['新机务段']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, Depot
    with client.application.app_context():
      depot = Depot.query.filter_by(name='新机务段').first()
      assert depot is not None

  def test_import_depot_skip_duplicate(self, client, sample_data):
    """测试导入机务段跳过重复"""
    from models import db, Depot
    with client.application.app_context():
      existing = Depot.query.filter_by(name='测试机务段').first()
      assert existing is not None

    buf = _make_excel({
      '机务段': (['名称'], [['测试机务段']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_merchant_sheet(self, client, sample_data):
    """测试导入商家工作表"""
    buf = _make_excel({
      '商家': (['名称'], [['新商家']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, Merchant
    with client.application.app_context():
      m = Merchant.query.filter_by(name='新商家').first()
      assert m is not None

  def test_import_power_type_sheet(self, client, sample_data):
    """测试导入动力类型工作表"""
    buf = _make_excel({
      '动力类型': (['名称'], [['蒸汽']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, PowerType
    with client.application.app_context():
      pt = PowerType.query.filter_by(name='蒸汽').first()
      assert pt is not None

  def test_import_chip_interface_sheet(self, client, sample_data):
    """测试导入芯片接口工作表"""
    buf = _make_excel({
      '芯片接口': (['名称'], [['PluX22']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, ChipInterface
    with client.application.app_context():
      ci = ChipInterface.query.filter_by(name='PluX22').first()
      assert ci is not None

  def test_import_chip_model_sheet(self, client, sample_data):
    """测试导入芯片型号工作表"""
    buf = _make_excel({
      '芯片型号': (['名称'], [['TCS WOW']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, ChipModel
    with client.application.app_context():
      cm = ChipModel.query.filter_by(name='TCS WOW').first()
      assert cm is not None

  def test_import_locomotive_series_sheet(self, client, sample_data):
    """测试导入机车系列工作表"""
    buf = _make_excel({
      '机车系列': (['名称'], [['DF系列']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, LocomotiveSeries
    with client.application.app_context():
      s = LocomotiveSeries.query.filter_by(name='DF系列').first()
      assert s is not None

  def test_import_carriage_series_sheet(self, client, sample_data):
    """测试导入车厢系列工作表"""
    buf = _make_excel({
      '车厢系列': (['名称'], [['RW系列']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, CarriageSeries
    with client.application.app_context():
      s = CarriageSeries.query.filter_by(name='RW系列').first()
      assert s is not None

  def test_import_trainset_series_sheet(self, client, sample_data):
    """测试导入动车组系列工作表"""
    buf = _make_excel({
      '动车组系列': (['名称'], [['CR系列']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, TrainsetSeries
    with client.application.app_context():
      s = TrainsetSeries.query.filter_by(name='CR系列').first()
      assert s is not None

  def test_import_locomotive_model_sheet(self, client, sample_data):
    """测试导入机车车型工作表"""
    buf = _make_excel({
      '机车车型': (['名称', '系列', '动力类型'], [['DF4B', 'SS系列', '电力']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_carriage_model_sheet(self, client, sample_data):
    """测试导入车厢车型工作表"""
    buf = _make_excel({
      '车厢车型': (['名称', '系列', '类型'], [['YZ25T', 'YZ系列', '客车']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_trainset_model_sheet(self, client, sample_data):
    """测试导入动车组车型工作表"""
    buf = _make_excel({
      '动车组车型': (['名称', '系列', '动力类型'], [['CRH2A', 'CRH系列', '电力']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_locomotive_sheet_skip_conflict(self, client, sample_data):
    """测试导入机车数据跳过冲突"""
    from models import db, Locomotive
    with client.application.app_context():
      loco = Locomotive(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', locomotive_number='1234'
      )
      db.session.add(loco)
      db.session.commit()

    buf = _make_excel({
      '机车': (['品牌', '比例', '系列', '动力', '车型', '机车号'],
              [['测试品牌', 'HO', 'SS系列', '电力', 'SS4', '1234']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_locomotive_sheet_overwrite(self, client, sample_data):
    """测试导入机车数据覆盖冲突"""
    from models import db, Locomotive
    with client.application.app_context():
      loco = Locomotive(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', locomotive_number='5555', item_number='OLD_ITEM'
      )
      db.session.add(loco)
      db.session.commit()

    buf = _make_excel({
      '机车': (['品牌', '比例', '系列', '动力', '车型', '机车号', '货号'],
              [['测试品牌', 'HO', 'SS系列', '电力', 'SS4', '5555', 'NEW_ITEM']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'overwrite'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    with client.application.app_context():
      loco = Locomotive.query.filter_by(locomotive_number='5555').first()
      assert loco.item_number == 'NEW_ITEM'

  def test_import_locomotive_missing_brand(self, client, sample_data):
    """测试导入机车数据时品牌不存在被跳过"""
    buf = _make_excel({
      '机车': (['品牌', '比例', '机车号'],
              [['不存在的品牌', 'HO', '9999']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['summary']['机车模型'] == 0

  def test_import_trainset_sheet_skip_conflict(self, client, sample_data):
    """测试导入动车组数据跳过冲突"""
    from models import db, Trainset
    with client.application.app_context():
      ts = Trainset(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', trainset_number='3001'
      )
      db.session.add(ts)
      db.session.commit()

    buf = _make_excel({
      '动车组': (['品牌', '比例', '动车号'],
               [['测试品牌', 'HO', '3001']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_trainset_sheet_overwrite(self, client, sample_data):
    """测试导入动车组数据覆盖冲突"""
    from models import db, Trainset
    with client.application.app_context():
      ts = Trainset(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', trainset_number='3001', item_number='OLD'
      )
      db.session.add(ts)
      db.session.commit()

    buf = _make_excel({
      '动车组': (['品牌', '比例', '动车号', '货号'],
               [['测试品牌', 'HO', '3001', 'NEW']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'overwrite'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    with client.application.app_context():
      ts = Trainset.query.filter_by(trainset_number='3001').first()
      assert ts.item_number == 'NEW'

  def test_import_carriage_sheet(self, client, sample_data):
    """测试导入车厢数据"""
    buf = _make_excel({
      '车厢': (['品牌', '比例', '系列', '车辆段', '车次', '车型', '车辆号'],
              [['测试品牌', 'HO', 'YZ系列', '测试机务段', 'G100', 'YZ22', 'YZ-01']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['summary']['车厢模型'] >= 1

  def test_import_unknown_sheet(self, client, sample_data):
    """测试导入未知工作表名称（被忽略）"""
    buf = _make_excel({
      '未知类型': (['名称'], [['测试']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_system_empty_name_skipped(self, client, sample_data):
    """测试导入系统数据时空名称行被跳过"""
    buf = _make_excel({
      '品牌': (['名称', '缩写', '搜索地址'], [['', '', ''], ['有效品牌', 'YX', 'http://x.com']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True


class TestExcelExportModes:
  """Excel 导出三种模式测试"""

  def test_export_models_mode_with_locomotive(self, client, sample_data):
    """测试 models 模式导出包含机车"""
    from models import db, Locomotive
    with client.application.app_context():
      loco = Locomotive(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', locomotive_number='0001'
      )
      db.session.add(loco)
      db.session.commit()

    response = client.get('/api/export/excel?mode=models')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert '机车' in wb.sheetnames

  def test_export_system_mode(self, client, sample_data):
    """测试 system 模式导出系统信息"""
    response = client.get('/api/export/excel?mode=system')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert '品牌' in wb.sheetnames
    assert '机务段' in wb.sheetnames
    assert '商家' in wb.sheetnames

  def test_export_all_mode(self, client, sample_data):
    """测试 all 模式导出模型和系统信息"""
    from models import db, Locomotive
    with client.application.app_context():
      loco = Locomotive(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', locomotive_number='0001'
      )
      db.session.add(loco)
      db.session.commit()

    response = client.get('/api/export/excel?mode=all')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert '机车' in wb.sheetnames
    assert '品牌' in wb.sheetnames

  def test_export_models_no_data(self, client):
    """测试 models 模式无数据返回错误"""
    response = client.get('/api/export/excel?mode=models')
    assert response.status_code == 400

  def test_export_system_no_data(self, client):
    """测试 system 模式无数据返回错误"""
    response = client.get('/api/export/excel?mode=system')
    assert response.status_code == 400

  def test_export_all_no_data(self, client):
    """测试 all 模式无数据返回错误"""
    response = client.get('/api/export/excel?mode=all')
    assert response.status_code == 400

  def test_export_with_carriage_set(self, client, sample_data):
    """测试导出包含车厢套装"""
    from models import db, CarriageSet, CarriageItem
    with client.application.app_context():
      cs = CarriageSet(
        brand_id=1, series_id=1, scale='HO',
        train_number='G100', item_number='EXP_CRG'
      )
      db.session.add(cs)
      db.session.commit()
      item = CarriageItem(
        set_id=cs.id, model_id=1, car_number='YZ-01',
        color='green', light_model_id=None
      )
      db.session.add(item)
      db.session.commit()

    response = client.get('/api/export/excel?mode=models')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert '车厢' in wb.sheetnames

  def test_export_with_trainset(self, client, sample_data):
    """测试导出包含动车组"""
    from models import db, Trainset
    with client.application.app_context():
      ts = Trainset(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', trainset_number='CRH001', head_light=True,
        light_model_id=None
      )
      db.session.add(ts)
      db.session.commit()

    response = client.get('/api/export/excel?mode=models')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert '动车组' in wb.sheetnames

  def test_export_system_with_series_and_models(self, client, sample_data):
    """测试系统导出包含系列和车型"""
    response = client.get('/api/export/excel?mode=system')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert '机车系列' in wb.sheetnames
    assert '机车车型' in wb.sheetnames
    assert '车厢系列' in wb.sheetnames
    assert '车厢车型' in wb.sheetnames
    assert '动车组系列' in wb.sheetnames
    assert '动车组车型' in wb.sheetnames

  def test_export_headers_bold(self, client, sample_data):
    """测试导出标题行加粗"""
    response = client.get('/api/export/excel?mode=system')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    sheet = wb['品牌']
    assert sheet.cell(row=1, column=1).font.bold is True


class TestImportConflictDetection:
  """导入冲突检测测试"""

  def test_preview_detects_brand_conflict(self, client, sample_data):
    """测试预览模式检测品牌冲突"""
    buf = _make_excel({
      '品牌': (['名称'], [['测试品牌']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'preview'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['has_conflicts'] is True
    assert any(c['type'] == '品牌' for c in data['conflicts'])

  def test_preview_detects_locomotive_number_conflict(self, client, sample_data):
    """测试预览模式检测机车号冲突"""
    from models import db, Locomotive
    with client.application.app_context():
      loco = Locomotive(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', locomotive_number='8888'
      )
      db.session.add(loco)
      db.session.commit()

    buf = _make_excel({
      '机车': (['品牌', '比例', '机车号'],
              [['测试品牌', 'HO', '8888']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'preview'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['has_conflicts'] is True
    assert any('8888' in c.get('value', '') for c in data['conflicts'])

  def test_preview_detects_trainset_number_conflict(self, client, sample_data):
    """测试预览模式检测动车号冲突"""
    from models import db, Trainset
    with client.application.app_context():
      ts = Trainset(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', trainset_number='7001'
      )
      db.session.add(ts)
      db.session.commit()

    buf = _make_excel({
      '动车组': (['品牌', '比例', '动车号'],
               [['测试品牌', 'HO', '7001']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'preview'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['has_conflicts'] is True
    assert any('7001' in c.get('value', '') for c in data['conflicts'])

  def test_preview_detects_decoder_number_conflict(self, client, sample_data):
    """测试预览模式检测编号冲突"""
    from models import db, Locomotive
    with client.application.app_context():
      loco = Locomotive(
        series_id=1, power_type_id=1, model_id=1, brand_id=1,
        scale='HO', decoder_number='99'
      )
      db.session.add(loco)
      db.session.commit()

    buf = _make_excel({
      '机车': (['品牌', '比例', '编号'],
              [['测试品牌', 'HO', '99']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'preview'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['has_conflicts'] is True


class TestImportTemplateCRUD:
  """导入模板 CRUD 测试"""

  def test_list_templates_empty(self, client):
    """测试获取空模板列表"""
    response = client.get('/api/import-templates')
    assert response.status_code == 200
    data = response.get_json()
    assert data['success'] is True
    assert isinstance(data['templates'], list)

  def test_create_template(self, client):
    """测试创建导入模板"""
    response = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '测试模板', 'config': {'test': True}}),
      content_type='application/json'
    )
    assert response.status_code == 201
    data = response.get_json()
    assert data['success'] is True
    assert data['template']['name'] == '测试模板'

  def test_create_template_no_name(self, client):
    """测试创建模板缺少名称"""
    response = client.post(
      '/api/import-templates',
      data=json.dumps({'config': {}}),
      content_type='application/json'
    )
    assert response.status_code == 400
    data = response.get_json()
    assert data['success'] is False

  def test_create_template_no_config(self, client):
    """测试创建模板缺少配置"""
    response = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '测试'}),
      content_type='application/json'
    )
    assert response.status_code == 400

  def test_create_template_invalid_config(self, client):
    """测试创建模板配置非对象"""
    response = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '测试', 'config': 'not_dict'}),
      content_type='application/json'
    )
    assert response.status_code == 400

  def test_create_template_empty_body(self, client):
    """测试创建模板空请求体（get_json 失败返回 500）"""
    response = client.post(
      '/api/import-templates',
      content_type='application/json'
    )
    # 空请求体导致 get_json() 失败，被泛 except 捕获返回 500
    assert response.status_code == 500

  def test_get_template(self, client):
    """测试获取单个模板"""
    # 先创建
    resp = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '获取测试', 'config': {'key': 'val'}}),
      content_type='application/json'
    )
    template_id = resp.get_json()['template']['id']

    response = client.get(f'/api/import-templates/{template_id}')
    assert response.status_code == 200
    data = response.get_json()
    assert data['success'] is True
    assert data['template']['name'] == '获取测试'

  def test_get_template_not_found(self, client):
    """测试获取不存在的模板"""
    response = client.get('/api/import-templates/99999')
    assert response.status_code == 404

  def test_update_template(self, client):
    """测试更新模板"""
    resp = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '更新前', 'config': {}}),
      content_type='application/json'
    )
    template_id = resp.get_json()['template']['id']

    response = client.put(
      f'/api/import-templates/{template_id}',
      data=json.dumps({'name': '更新后', 'config': {'updated': True}}),
      content_type='application/json'
    )
    assert response.status_code == 200
    data = response.get_json()
    assert data['template']['name'] == '更新后'

  def test_update_template_not_found(self, client):
    """测试更新不存在的模板"""
    response = client.put(
      '/api/import-templates/99999',
      data=json.dumps({'name': '测试'}),
      content_type='application/json'
    )
    assert response.status_code == 404

  def test_update_template_empty_body(self, client):
    """测试更新模板空请求体（get_json 失败返回 500）"""
    resp = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '空体测试', 'config': {}}),
      content_type='application/json'
    )
    template_id = resp.get_json()['template']['id']

    response = client.put(
      f'/api/import-templates/{template_id}',
      content_type='application/json'
    )
    # 空请求体导致 get_json() 失败，被泛 except 捕获返回 500
    assert response.status_code == 500

  def test_update_template_invalid_config(self, client):
    """测试更新模板配置非对象"""
    resp = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '配置测试', 'config': {}}),
      content_type='application/json'
    )
    template_id = resp.get_json()['template']['id']

    response = client.put(
      f'/api/import-templates/{template_id}',
      data=json.dumps({'config': [1, 2]}),
      content_type='application/json'
    )
    assert response.status_code == 400

  def test_delete_template(self, client):
    """测试删除模板"""
    resp = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '删除测试', 'config': {}}),
      content_type='application/json'
    )
    template_id = resp.get_json()['template']['id']

    response = client.delete(f'/api/import-templates/{template_id}')
    assert response.status_code == 200
    data = response.get_json()
    assert data['success'] is True

    # 确认已删除
    response = client.get(f'/api/import-templates/{template_id}')
    assert response.status_code == 404

  def test_delete_template_not_found(self, client):
    """测试删除不存在的模板"""
    response = client.delete('/api/import-templates/99999')
    assert response.status_code == 404

  def test_copy_template(self, client):
    """测试复制模板"""
    resp = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '原件', 'config': {'copied': True}}),
      content_type='application/json'
    )
    template_id = resp.get_json()['template']['id']

    response = client.post(
      f'/api/import-templates/{template_id}/copy',
      data=json.dumps({'name': '副本'}),
      content_type='application/json'
    )
    assert response.status_code == 200
    data = response.get_json()
    assert data['success'] is True
    assert data['template']['name'] == '副本'
    assert data['template']['config'] == {'copied': True}
    assert data['template']['id'] != template_id

  def test_copy_template_not_found(self, client):
    """测试复制不存在的模板"""
    response = client.post(
      '/api/import-templates/99999/copy',
      data=json.dumps({'name': '副本'}),
      content_type='application/json'
    )
    assert response.status_code == 404

  def test_copy_template_no_name(self, client):
    """测试复制模板缺少名称"""
    resp = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '无名称复制', 'config': {}}),
      content_type='application/json'
    )
    template_id = resp.get_json()['template']['id']

    response = client.post(
      f'/api/import-templates/{template_id}/copy',
      data=json.dumps({}),
      content_type='application/json'
    )
    assert response.status_code == 400

  def test_copy_template_duplicate_name(self, client):
    """测试复制模板名称已存在"""
    client.post(
      '/api/import-templates',
      data=json.dumps({'name': '唯一名', 'config': {}}),
      content_type='application/json'
    )
    resp = client.post(
      '/api/import-templates',
      data=json.dumps({'name': '要复制的', 'config': {}}),
      content_type='application/json'
    )
    template_id = resp.get_json()['template']['id']

    response = client.post(
      f'/api/import-templates/{template_id}/copy',
      data=json.dumps({'name': '唯一名'}),
      content_type='application/json'
    )
    assert response.status_code == 400


class TestCustomImportAPI:
  """自定义导入 API 测试"""

  def test_get_tables(self, client):
    """测试获取可用表配置"""
    response = client.get('/api/custom-import/tables')
    assert response.status_code == 200
    data = response.get_json()
    assert data['success'] is True
    assert len(data['tables']) > 0

  def test_parse_excel(self, client):
    """测试解析 Excel 文件"""
    buf = _make_excel({
      'Sheet1': (['品牌名称', '搜索地址'], [['测试', 'http://x.com']])
    })
    response = client.post(
      '/api/custom-import/parse',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['filename'] == 'test.xlsx'
    assert len(data['sheets']) == 1
    assert data['sheets'][0]['name'] == 'Sheet1'
    assert data['sheets'][0]['row_count'] == 1

  def test_parse_excel_no_file(self, client):
    """测试解析无文件"""
    response = client.post(
      '/api/custom-import/parse',
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_parse_excel_wrong_format(self, client):
    """测试解析非 Excel 文件"""
    buf = io.BytesIO(b'not excel')
    response = client.post(
      '/api/custom-import/parse',
      data={'file': (buf, 'test.txt')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_preview_custom_import(self, client, sample_data):
    """测试自定义导入预览"""
    buf = _make_excel({
      '品牌': (['名称', '搜索地址'], [['预览品牌', 'http://x.com']])
    })
    config = json.dumps({
      'sheet_mappings': [{'sheet_name': '品牌', 'table_name': 'brand'}],
      'column_mappings': {
        'brand': {
          'columns': [
            {'source': '名称', 'target': 'name', 'required': True},
            {'source': '搜索地址', 'target': 'search_url', 'required': False}
          ],
          'conflict_mode': 'skip'
        }
      }
    })
    response = client.post(
      '/api/custom-import/preview',
      data={'file': (buf, 'test.xlsx'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert 'previews' in data
    assert len(data['previews']) == 1
    assert data['previews'][0]['table_name'] == 'brand'

  def test_preview_no_file(self, client):
    """测试预览无文件"""
    config = json.dumps({
      'sheet_mappings': [],
      'column_mappings': {}
    })
    response = client.post(
      '/api/custom-import/preview',
      data={'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_preview_no_config(self, client, sample_data):
    """测试预览缺少配置"""
    buf = _make_excel({'Sheet1': (['A'], [['B']])})
    response = client.post(
      '/api/custom-import/preview',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_preview_invalid_config_json(self, client, sample_data):
    """测试预览配置 JSON 格式错误"""
    buf = _make_excel({'Sheet1': (['A'], [['B']])})
    response = client.post(
      '/api/custom-import/preview',
      data={'file': (buf, 'test.xlsx'), 'config': 'not json'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_preview_missing_required_field(self, client, sample_data):
    """测试预览检测缺失必填字段"""
    buf = _make_excel({
      '品牌': (['搜索地址'], [['http://x.com']])
    })
    config = json.dumps({
      'sheet_mappings': [{'sheet_name': '品牌', 'table_name': 'brand'}],
      'column_mappings': {
        'brand': {
          'columns': [
            {'source': '搜索地址', 'target': 'search_url', 'required': False}
          ],
          'conflict_mode': 'skip'
        }
      }
    })
    response = client.post(
      '/api/custom-import/preview',
      data={'file': (buf, 'test.xlsx'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['can_proceed'] is False
    assert 'name' in data['previews'][0].get('missing_required', [])

  def test_preview_brand_conflict(self, client, sample_data):
    """测试预览检测品牌冲突"""
    buf = _make_excel({
      '品牌': (['名称'], [['测试品牌']])
    })
    config = json.dumps({
      'sheet_mappings': [{'sheet_name': '品牌', 'table_name': 'brand'}],
      'column_mappings': {
        'brand': {
          'columns': [
            {'source': '名称', 'target': 'name', 'required': True}
          ],
          'conflict_mode': 'overwrite'
        }
      }
    })
    response = client.post(
      '/api/custom-import/preview',
      data={'file': (buf, 'test.xlsx'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['has_conflicts'] is True

  def test_execute_custom_import_brand(self, client, sample_data):
    """测试执行自定义导入品牌数据"""
    buf = _make_excel({
      '品牌列表': (['品牌名称', '官网'], [['导入品牌', 'http://import.com']])
    })
    config = json.dumps({
      'sheet_mappings': [{'sheet_name': '品牌列表', 'table_name': 'brand'}],
      'column_mappings': {
        'brand': {
          'columns': [
            {'source': '品牌名称', 'target': 'name', 'required': True},
            {'source': '官网', 'target': 'search_url', 'required': False}
          ],
          'conflict_mode': 'skip'
        }
      }
    })
    response = client.post(
      '/api/custom-import/execute',
      data={'file': (buf, 'test.xlsx'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['summary'].get('brand', 0) >= 1

    from models import db, Brand
    with client.application.app_context():
      brand = Brand.query.filter_by(name='导入品牌').first()
      assert brand is not None

  def test_execute_custom_import_depot(self, client, sample_data):
    """测试执行自定义导入机务段"""
    buf = _make_excel({
      '机务段列表': (['名称'], [['导入机务段']])
    })
    config = json.dumps({
      'sheet_mappings': [{'sheet_name': '机务段列表', 'table_name': 'depot'}],
      'column_mappings': {
        'depot': {
          'columns': [
            {'source': '名称', 'target': 'name', 'required': True}
          ],
          'conflict_mode': 'skip'
        }
      }
    })
    response = client.post(
      '/api/custom-import/execute',
      data={'file': (buf, 'test.xlsx'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_execute_custom_import_no_file(self, client):
    """测试执行自定义导入无文件"""
    config = json.dumps({'sheet_mappings': [], 'column_mappings': {}})
    response = client.post(
      '/api/custom-import/execute',
      data={'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_execute_custom_import_no_config(self, client):
    """测试执行自定义导入无配置"""
    buf = _make_excel({'Sheet1': (['A'], [['B']])})
    response = client.post(
      '/api/custom-import/execute',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_execute_custom_import_invalid_config(self, client):
    """测试执行自定义导入无效配置"""
    buf = _make_excel({'Sheet1': (['A'], [['B']])})
    response = client.post(
      '/api/custom-import/execute',
      data={'file': (buf, 'test.xlsx'), 'config': 'not json'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False

  def test_execute_custom_import_brand_overwrite(self, client, sample_data):
    """测试执行自定义导入覆盖品牌"""
    buf = _make_excel({
      '品牌列表': (['品牌名称', '官网'], [['测试品牌', 'http://overwritten.com']])
    })
    config = json.dumps({
      'sheet_mappings': [{'sheet_name': '品牌列表', 'table_name': 'brand'}],
      'column_mappings': {
        'brand': {
          'columns': [
            {'source': '品牌名称', 'target': 'name', 'required': True},
            {'source': '官网', 'target': 'search_url', 'required': False}
          ],
          'conflict_mode': 'overwrite'
        }
      }
    })
    response = client.post(
      '/api/custom-import/execute',
      data={'file': (buf, 'test.xlsx'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

    from models import db, Brand
    with client.application.app_context():
      brand = Brand.query.filter_by(name='测试品牌').first()
      assert brand.search_url == 'http://overwritten.com'

  def test_execute_custom_import_series(self, client, sample_data):
    """测试执行自定义导入系列数据"""
    buf = _make_excel({
      '系列': (['名称'], [['新系列XY']])
    })
    config = json.dumps({
      'sheet_mappings': [{'sheet_name': '系列', 'table_name': 'locomotive_series'}],
      'column_mappings': {
        'locomotive_series': {
          'columns': [
            {'source': '名称', 'target': 'name', 'required': True}
          ],
          'conflict_mode': 'skip'
        }
      }
    })
    response = client.post(
      '/api/custom-import/execute',
      data={'file': (buf, 'test.xlsx'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_execute_custom_import_model(self, client, sample_data):
    """测试执行自定义导入车型数据"""
    buf = _make_excel({
      '车型': (['名称', '系列', '动力类型'], [['新车型XX', 'SS系列', '电力']])
    })
    config = json.dumps({
      'sheet_mappings': [{'sheet_name': '车型', 'table_name': 'locomotive_model'}],
      'column_mappings': {
        'locomotive_model': {
          'columns': [
            {'source': '名称', 'target': 'name', 'required': True},
            {'source': '系列', 'target': 'series_id', 'required': False, 'ref': 'locomotive_series'},
            {'source': '动力类型', 'target': 'power_type_id', 'required': False, 'ref': 'power_type'}
          ],
          'conflict_mode': 'skip'
        }
      }
    })
    response = client.post(
      '/api/custom-import/execute',
      data={'file': (buf, 'test.xlsx'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_execute_custom_import_wrong_format(self, client):
    """测试执行自定义导入非 Excel 文件"""
    buf = io.BytesIO(b'not excel')
    config = json.dumps({'sheet_mappings': [], 'column_mappings': {}})
    response = client.post(
      '/api/custom-import/execute',
      data={'file': (buf, 'test.txt'), 'config': config},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is False


class TestAdditionalCoverage:
  """补充覆盖率的额外测试"""

  def test_import_brand_skip_duplicate(self, client, sample_data):
    """测试 skip 模式下品牌重复被跳过"""
    buf = _make_excel({
      '品牌': (['名称', '缩写', '搜索地址'], [['测试品牌', 'CSP', 'http://skip.com']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    # skip 模式下重复品牌不计入 count
    assert data['summary'].get('品牌', 0) == 0

  def test_import_depot_overwrite_no_change(self, client, sample_data):
    """测试 overwrite 模式下机务段重复（只有 name 字段无变化）"""
    buf = _make_excel({
      '机务段': (['名称'], [['测试机务段']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'overwrite'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_merchant_skip_duplicate(self, client, sample_data):
    """测试 skip 模式下商家重复被跳过"""
    buf = _make_excel({
      '商家': (['名称'], [['测试商家']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['summary'].get('商家', 0) == 0

  def test_import_power_type_skip_duplicate(self, client, sample_data):
    """测试 skip 模式下动力类型重复被跳过"""
    buf = _make_excel({
      '动力类型': (['名称'], [['电力']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_chip_interface_skip_duplicate(self, client, sample_data):
    """测试 skip 模式下芯片接口重复被跳过"""
    buf = _make_excel({
      '芯片接口': (['名称'], [['Next18']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_chip_model_skip_duplicate(self, client, sample_data):
    """测试 skip 模式下芯片型号重复被跳过"""
    buf = _make_excel({
      '芯片型号': (['名称'], [['ESU LokSound']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_locomotive_series_skip_duplicate(self, client, sample_data):
    """测试 skip 模式下机车系列重复被跳过"""
    buf = _make_excel({
      '机车系列': (['名称'], [['SS系列']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_carriage_series_skip_duplicate(self, client, sample_data):
    """测试 skip 模式下车厢系列重复被跳过"""
    buf = _make_excel({
      '车厢系列': (['名称'], [['YZ系列']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_trainset_series_skip_duplicate(self, client, sample_data):
    """测试 skip 模式下动车组系列重复被跳过"""
    buf = _make_excel({
      '动车组系列': (['名称'], [['CRH系列']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'skip'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_merchant_overwrite_no_change(self, client, sample_data):
    """测试 overwrite 模式下商家重复（只有 name 字段无变化）"""
    buf = _make_excel({
      '商家': (['名称'], [['测试商家']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'overwrite'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_power_type_overwrite_no_change(self, client, sample_data):
    """测试 overwrite 模式下动力类型重复"""
    buf = _make_excel({
      '动力类型': (['名称'], [['电力']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'overwrite'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_chip_interface_overwrite_no_change(self, client, sample_data):
    """测试 overwrite 模式下芯片接口重复"""
    buf = _make_excel({
      '芯片接口': (['名称'], [['Next18']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'overwrite'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_chip_model_overwrite_no_change(self, client, sample_data):
    """测试 overwrite 模式下芯片型号重复"""
    buf = _make_excel({
      '芯片型号': (['名称'], [['ESU LokSound']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'overwrite'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_import_locomotive_missing_scale(self, client, sample_data):
    """测试导入机车数据时缺少比例字段被跳过"""
    buf = _make_excel({
      '机车': (['品牌', '机车号'],
              [['测试品牌', '9998']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['summary']['机车模型'] == 0

  def test_import_trainset_missing_brand(self, client, sample_data):
    """测试导入动车组数据时品牌不存在被跳过"""
    buf = _make_excel({
      '动车组': (['品牌', '比例', '动车号'],
               [['不存在的品牌', 'HO', '9999']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['summary']['动车组模型'] == 0

  def test_import_trainset_missing_scale(self, client, sample_data):
    """测试导入动车组数据时缺少比例字段被跳过"""
    buf = _make_excel({
      '动车组': (['品牌', '动车号'],
               [['测试品牌', '9998']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['summary']['动车组模型'] == 0

  def test_import_empty_name_rows_all_system_tables(self, client, sample_data):
    """测试所有系统表空名称行被跳过"""
    buf = _make_excel({
      '商家': (['名称'], [['']]),
      '动力类型': (['名称'], [['']]),
      '芯片接口': (['名称'], [['']]),
      '芯片型号': (['名称'], [['']]),
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx')},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True

  def test_export_system_mode_has_data(self, client, sample_data):
    """测试 system 模式导出所有系统表都有数据"""
    from models import db, ChipInterface, ChipModel
    with client.application.app_context():
      ci = ChipInterface.query.first()
      cm = ChipModel.query.first()
      assert ci is not None
      assert cm is not None

    response = client.get('/api/export/excel?mode=system')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert '芯片接口' in wb.sheetnames
    assert '芯片型号' in wb.sheetnames

  def test_export_with_locomotive_head(self, client, sample_data):
    """测试导出包含先头车"""
    from models import db, LocomotiveHead
    with client.application.app_context():
      head = LocomotiveHead(
        model_id=1, brand_id=1, scale='HO',
        item_number='EXP_LH', head_light=True, light_model_id=None
      )
      db.session.add(head)
      db.session.commit()

    response = client.get('/api/export/excel?mode=models')
    assert response.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(response.data))
    assert '先头车' in wb.sheetnames


class TestCheckImportConflicts:
  """冲突检测逻辑的间接测试（通过 preview 模式）"""

  def test_system_conflict_multiple_types(self, client, sample_data):
    """测试多个系统表同时冲突"""
    buf = _make_excel({
      '品牌': (['名称'], [['测试品牌']]),
      '商家': (['名称'], [['测试商家']]),
      '机务段': (['名称'], [['测试机务段']])
    })
    response = client.post(
      '/api/import/excel',
      data={'file': (buf, 'test.xlsx'), 'mode': 'preview'},
      content_type='multipart/form-data'
    )
    data = response.get_json()
    assert data['success'] is True
    assert data['has_conflicts'] is True
    conflict_types = {c['type'] for c in data['conflicts']}
    assert '品牌' in conflict_types
    assert '商家' in conflict_types
    assert '机务段' in conflict_types
