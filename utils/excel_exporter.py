"""
Excel 导出工具模块

从 routes/excel_io.py 拆分而出，封装数据导出核心逻辑。
接收 mode 参数（models / system / all），生成 Excel 工作簿并返回 BytesIO 流。
"""
from io import BytesIO
from datetime import datetime
import random
import logging

import openpyxl
from openpyxl.styles import Font

from models import (
  LocomotiveModel, CarriageModel, TrainsetModel,
  Locomotive, CarriageSet, Trainset, LocomotiveHead,
  Brand, Depot, Merchant, ChipInterface, ChipModel,
  LocomotiveSeries, CarriageSeries, TrainsetSeries, PowerType,
)

logger = logging.getLogger(__name__)


def export_to_excel_core(mode='models'):
  """导出数据到 Excel 工作簿。

  Args:
    mode: 导出模式，可选 'models'（模型数据）、'system'（系统信息）、'all'（全部）

  Returns:
    BytesIO: 包含 .xlsx 文件的字节流（已 seek(0)），调用方可直接用于 send_file。

  Raises:
    ValueError: 当对应模式没有可导出数据时抛出，message 为中文错误描述。
  """
  # 检查是否有数据
  has_model_data = (
    Locomotive.query.count() > 0 or
    CarriageSet.query.count() > 0 or
    Trainset.query.count() > 0 or
    LocomotiveHead.query.count() > 0
  )
  has_system_data = (
    Brand.query.count() > 0 or
    Depot.query.count() > 0 or
    Merchant.query.count() > 0 or
    PowerType.query.count() > 0 or
    ChipInterface.query.count() > 0 or
    ChipModel.query.count() > 0
  )

  if mode == 'models' and not has_model_data:
    raise ValueError('当前没有可导出的模型数据')
  if mode == 'system' and not has_system_data:
    raise ValueError('当前没有可导出的系统信息')
  if mode == 'all' and not has_model_data and not has_system_data:
    raise ValueError('当前没有可导出的数据')

  workbook = openpyxl.Workbook()

  if 'Sheet' in workbook.sheetnames:
    workbook.remove(workbook['Sheet'])

  # 导出模型数据（models 或 all 模式）
  if mode in ('models', 'all'):
    _export_models(workbook)

  # 导出系统信息（system 或 all 模式）
  if mode in ('system', 'all'):
    _export_system(workbook)

  # 加粗所有工作表的第一行（标题行）
  bold_font = Font(bold=True)
  for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for cell in sheet[1]:
      cell.font = bold_font

  output = BytesIO()
  workbook.save(output)
  output.seek(0)

  logger.info(f"Excel export completed successfully, mode={mode}")
  return output


def build_export_filename(mode):
  """根据导出模式构建下载文件名。

  Args:
    mode: 导出模式（models / system / all）

  Returns:
    str: 形如 TMM_Models_YYYYMMDD_HHMMSS_RAND.xlsx 的文件名
  """
  mode_names = {'models': 'Models', 'system': 'System', 'all': 'All'}
  return f'TMM_{mode_names.get(mode, "Export")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{random.randint(1000, 9999)}.xlsx'


def _export_models(workbook):
  """导出四种模型类型数据到工作簿。"""
  # 导出机车模型
  if Locomotive.query.count() > 0:
    sheet = workbook.create_sheet('机车')
    headers = ['系列', '动力', '车型', '品牌', '机务段', '挂牌', '颜色', '比例', '机车号', '编号',
         '芯片接口', '芯片型号', '价格', '总价', '货号', '购买日期', '购买商家']
    sheet.append(headers)

    for loco in Locomotive.query.all():
      sheet.append([
        loco.series.name if loco.series else '',
        loco.power_type.name if loco.power_type else '',
        loco.model.name if loco.model else '',
        loco.brand.name if loco.brand else '',
        loco.depot.name if loco.depot else '',
        loco.plaque or '',
        loco.color or '',
        loco.scale or '',
        loco.locomotive_number or '',
        loco.decoder_number or '',
        loco.chip_interface.name if loco.chip_interface else '',
        loco.chip_model.name if loco.chip_model else '',
        loco.price or '',
        loco.total_price or '',
        loco.item_number or '',
        loco.purchase_date.strftime('%Y-%m-%d') if loco.purchase_date else '',
        loco.merchant.name if loco.merchant else ''
      ])

  # 导出车厢模型（使用合并单元格标识套装）
  if CarriageSet.query.count() > 0:
    sheet = workbook.create_sheet('车厢')
    headers = ['品牌', '系列', '车辆段', '车次', '挂牌', '货号', '比例', '车型', '车辆号', '颜色', '灯光', '总价', '购买日期', '购买商家']
    sheet.append(headers)

    current_row = 2  # 从第2行开始（第1行是表头）
    for carriage_set in CarriageSet.query.all():
      items = carriage_set.items
      if not items:
        # 无车厢项的套装，单独一行
        sheet.append([
          carriage_set.brand.name if carriage_set.brand else '',
          carriage_set.series.name if carriage_set.series else '',
          carriage_set.depot.name if carriage_set.depot else '',
          carriage_set.train_number or '',
          carriage_set.plaque or '',
          carriage_set.item_number or '',
          carriage_set.scale or '',
          '', '', '', '',
          carriage_set.total_price or '',
          carriage_set.purchase_date.strftime('%Y-%m-%d') if carriage_set.purchase_date else '',
          carriage_set.merchant.name if carriage_set.merchant else ''
        ])
        current_row += 1
      else:
        # 有车厢项的套装，合并公共信息列
        start_row = current_row
        for item in items:
          sheet.append([
            carriage_set.brand.name if carriage_set.brand else '',
            carriage_set.series.name if carriage_set.series else '',
            carriage_set.depot.name if carriage_set.depot else '',
            carriage_set.train_number or '',
            carriage_set.plaque or '',
            carriage_set.item_number or '',
            carriage_set.scale or '',
            item.model.name if item.model else '',
            item.car_number or '',
            item.color or '',
            item.light_model.name if item.light_model else '',
            carriage_set.total_price or '',
            carriage_set.purchase_date.strftime('%Y-%m-%d') if carriage_set.purchase_date else '',
            carriage_set.merchant.name if carriage_set.merchant else ''
          ])
          current_row += 1

        # 合并公共信息列（前7列：A-G，即品牌到比例）
        # 合并总价、购买日期、购买商家列（L-N，即第12-14列）
        if len(items) > 1:
          end_row = current_row - 1
          # 合并前7列（品牌、系列、车辆段、车次、挂牌、货号、比例）
          for col in range(1, 8):  # A-G 列
            sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
          # 合并后3列（总价、购买日期、购买商家）
          for col in range(12, 15):  # L-N 列
            sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

  # 导出动车组模型
  if Trainset.query.count() > 0:
    sheet = workbook.create_sheet('动车组')
    headers = ['系列', '动力', '车型', '品牌', '动车段', '挂牌', '颜色', '比例', '编组', '动车号', '编号',
         '头车灯', '室内灯', '芯片接口', '芯片型号', '价格', '总价', '货号', '购买日期', '购买商家']
    sheet.append(headers)

    for ts in Trainset.query.all():
      sheet.append([
        ts.series.name if ts.series else '',
        ts.power_type.name if ts.power_type else '',
        ts.model.name if ts.model else '',
        ts.brand.name if ts.brand else '',
        ts.depot.name if ts.depot else '',
        ts.plaque or '',
        ts.color or '',
        ts.scale or '',
        ts.formation or '',
        ts.trainset_number or '',
        ts.decoder_number or '',
        '是' if ts.head_light else '否',
        ts.light_model.name if ts.light_model else '',
        ts.chip_interface.name if ts.chip_interface else '',
        ts.chip_model.name if ts.chip_model else '',
        ts.price or '',
        ts.total_price or '',
        ts.item_number or '',
        ts.purchase_date.strftime('%Y-%m-%d') if ts.purchase_date else '',
        ts.merchant.name if ts.merchant else ''
      ])

  # 导出先头车模型
  if LocomotiveHead.query.count() > 0:
    sheet = workbook.create_sheet('先头车')
    headers = ['车型', '品牌', '涂装', '比例', '头车灯', '室内灯', '价格', '总价', '货号', '购买日期', '购买商家']
    sheet.append(headers)

    for head in LocomotiveHead.query.all():
      sheet.append([
        head.model.name if head.model else '',
        head.brand.name if head.brand else '',
        head.special_color or '',
        head.scale or '',
        '是' if head.head_light else '否',
        head.light_model.name if head.light_model else '',
        head.price or '',
        head.total_price or '',
        head.item_number or '',
        head.purchase_date.strftime('%Y-%m-%d') if head.purchase_date else '',
        head.merchant.name if head.merchant else ''
      ])


def _export_system(workbook):
  """导出系统信息（品牌/段/商家/动力/芯片/系列/车型）到工作簿。"""
  # 导出品牌
  if Brand.query.count() > 0:
    sheet = workbook.create_sheet('品牌')
    sheet.append(['名称', '搜索地址'])
    for brand in Brand.query.all():
      sheet.append([brand.name, brand.search_url or ''])

  # 导出机务段
  if Depot.query.count() > 0:
    sheet = workbook.create_sheet('机务段')
    sheet.append(['名称'])
    for depot in Depot.query.all():
      sheet.append([depot.name])

  # 导出商家
  if Merchant.query.count() > 0:
    sheet = workbook.create_sheet('商家')
    sheet.append(['名称'])
    for merchant in Merchant.query.all():
      sheet.append([merchant.name])

  # 导出动力类型
  if PowerType.query.count() > 0:
    sheet = workbook.create_sheet('动力类型')
    sheet.append(['名称'])
    for pt in PowerType.query.all():
      sheet.append([pt.name])

  # 导出芯片接口
  if ChipInterface.query.count() > 0:
    sheet = workbook.create_sheet('芯片接口')
    sheet.append(['名称'])
    for ci in ChipInterface.query.all():
      sheet.append([ci.name])

  # 导出芯片型号
  if ChipModel.query.count() > 0:
    sheet = workbook.create_sheet('芯片型号')
    sheet.append(['名称', '接口'])
    for cm in ChipModel.query.all():
      interface_names = ', '.join([i.name for i in cm.interfaces])
      sheet.append([cm.name, interface_names])

  # 导出机车系列
  if LocomotiveSeries.query.count() > 0:
    sheet = workbook.create_sheet('机车系列')
    sheet.append(['名称'])
    for series in LocomotiveSeries.query.all():
      sheet.append([series.name])

  # 导出车厢系列
  if CarriageSeries.query.count() > 0:
    sheet = workbook.create_sheet('车厢系列')
    sheet.append(['名称'])
    for series in CarriageSeries.query.all():
      sheet.append([series.name])

  # 导出动车组系列
  if TrainsetSeries.query.count() > 0:
    sheet = workbook.create_sheet('动车组系列')
    sheet.append(['名称'])
    for series in TrainsetSeries.query.all():
      sheet.append([series.name])

  # 导出机车车型
  if LocomotiveModel.query.count() > 0:
    sheet = workbook.create_sheet('机车车型')
    sheet.append(['名称', '系列', '动力类型'])
    for model in LocomotiveModel.query.all():
      sheet.append([
        model.name,
        model.series.name if model.series else '',
        model.power_type.name if model.power_type else ''
      ])

  # 导出车厢车型
  if CarriageModel.query.count() > 0:
    sheet = workbook.create_sheet('车厢车型')
    sheet.append(['名称', '系列', '类型'])
    for model in CarriageModel.query.all():
      sheet.append([
        model.name,
        model.series.name if model.series else '',
        model.type or ''
      ])

  # 导出动车组车型
  if TrainsetModel.query.count() > 0:
    sheet = workbook.create_sheet('动车组车型')
    sheet.append(['名称', '系列', '动力类型'])
    for model in TrainsetModel.query.all():
      sheet.append([
        model.name,
        model.series.name if model.series else '',
        model.power_type.name if model.power_type else ''
      ])
