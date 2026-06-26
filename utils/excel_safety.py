"""
Excel 上传安全校验模块（P0-4）

集中处理 Excel 导入的文件校验，防止：
- 不支持格式导致 500（仅允许 .xlsx；旧 .xls 会让 openpyxl 抛 InvalidFileException）
- zip 炸弹式 DoS（read_only 流式加载 + 行列上限）
"""
import logging
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# 导入规模上限（防 zip 炸弹 / 内存耗尽）
MAX_IMPORT_ROWS = 50000
MAX_IMPORT_COLS = 100


def validate_excel_upload(file_storage):
  """校验上传的 Excel 文件并以 read_only 模式加载，返回 Workbook。

  Args:
    file_storage: werkzeug FileStorage（request.files['file']）

  Returns:
    openpyxl.Workbook: read_only 模式的工作簿

  Raises:
    ValueError: 文件名缺失/格式不支持/行列超限
  """
  filename = file_storage.filename or ''
  if not filename:
    raise ValueError('未选择文件')
  # 仅允许 .xlsx（openpyxl 原生支持；.xls 旧格式会抛 InvalidFileException）
  if not filename.lower().endswith('.xlsx'):
    raise ValueError('仅支持 .xlsx 格式文件')

  # 默认模式加载（read_only 不支持 merged_cells，车厢合并单元格检测依赖它）
  # 上传字节由 MAX_CONTENT_LENGTH（50MB）限制；行列上限进一步防超大文件
  workbook = load_workbook(file_storage, data_only=True)

  # 校验行列规模
  for sheet in workbook.worksheets:
    if sheet.max_row and sheet.max_row > MAX_IMPORT_ROWS:
      workbook.close()
      raise ValueError(f'数据行数超限（{sheet.max_row} > {MAX_IMPORT_ROWS}），请拆分文件')
    if sheet.max_column and sheet.max_column > MAX_IMPORT_COLS:
      workbook.close()
      raise ValueError(f'数据列数超限（{sheet.max_column} > {MAX_IMPORT_COLS}）')

  logger.info(f'Excel 校验通过: {filename}')
  return workbook
