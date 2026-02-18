"""
API 路由 Blueprint
包含自动填充 API 和 Excel 导入导出
"""
from flask import Blueprint, request, jsonify, send_file
from models import db, LocomotiveModel, CarriageModel, TrainsetModel
from models import Locomotive, CarriageSet, Trainset, LocomotiveHead
from models import Brand, Depot, Merchant, ChipInterface, ChipModel
from models import LocomotiveSeries, CarriageSeries, TrainsetSeries, PowerType
from models import CarriageItem, ImportTemplate
from utils.helpers import parse_purchase_date, safe_int, safe_float, parse_boolean
from utils.price_calculator import calculate_price
from utils.system_tables import get_table_display_info, SYSTEM_TABLES
from io import BytesIO
from datetime import datetime
import openpyxl
import logging
import random
import json

logger = logging.getLogger(__name__)
api_bp = Blueprint('api', __name__, url_prefix='')


# 自动填充 API
@api_bp.route('/api/auto-fill/locomotive/<int:model_id>')
def auto_fill_locomotive(model_id):
  """机车车型自动填充"""
  model = LocomotiveModel.query.get_or_404(model_id)
  return jsonify({
    'series_id': model.series_id,
    'power_type_id': model.power_type_id
  })


@api_bp.route('/api/auto-fill/carriage/<int:model_id>')
def auto_fill_carriage(model_id):
  """车厢车型自动填充"""
  model = CarriageModel.query.get_or_404(model_id)
  return jsonify({
    'series_id': model.series_id,
    'type': model.type
  })


@api_bp.route('/api/auto-fill/trainset/<int:model_id>')
def auto_fill_trainset(model_id):
  """动车组车型自动填充"""
  model = TrainsetModel.query.get_or_404(model_id)
  return jsonify({
    'series_id': model.series_id,
    'power_type_id': model.power_type_id
  })


# Excel 导入导出
def find_id_by_name(model, name, custom_query=None):
  """根据名称查找模型的ID（大小写不敏感）"""
  if not name:
    return None

  if custom_query:
    result = custom_query(model.query)
    return result.id if result else None

  # 大小写不敏感匹配
  obj = model.query.filter(db.func.lower(model.name) == db.func.lower(name)).first()
  return obj.id if obj else None


def validate_required(value, field_name):
  """验证必填字段"""
  if not value:
    raise ValueError(f"必填字段 '{field_name}' 不能为空")
  return value


@api_bp.route('/api/import/excel', methods=['POST'])
def import_from_excel():
  """从 Excel 文件导入数据 - 支持 preview、skip、overwrite 模式"""
  try:
    if 'file' not in request.files:
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
      return jsonify({'success': False, 'error': '文件格式错误，请上传Excel文件'}), 400

    # 获取模式参数：preview（只检查）、skip（跳过冲突）、overwrite（覆盖冲突）
    # 默认为 skip（向后兼容），前端显式传 preview 进行预检查
    mode = request.form.get('mode', 'skip')

    workbook = openpyxl.load_workbook(file)
    all_data = {}  # 存储 all sheet 数据

    # 读取所有 sheet 数据
    for sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
      data = []
      headers = None
      for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_idx == 1:
          headers = [cell for cell in row if cell is not None]
        else:
          if any(cell is not None for cell in row):
            data.append(dict(zip(headers, row)))
      if data:
        all_data[sheet_name] = data

    if not all_data:
      return jsonify({'success': True, 'summary': {}, 'message': '没有找到可导入的数据'})

    # 预检查冲突
    conflicts = check_import_conflicts(all_data)

    if mode == 'preview':
      # 只返回预检查结果
      return jsonify({
        'success': True,
        'preview': True,
        'conflicts': conflicts,
        'has_conflicts': len(conflicts) > 0,
        'sheets': list(all_data.keys())
      })

    # 执行导入
    summary = {}
    errors = []

    # 模型数据 sheet 名称映射
    model_sheets = {
      '机车': ('机车模型', import_locomotive_data_with_mode),
      '车厢': ('车厢模型', import_carriage_data_with_mode),
      '动车组': ('动车组模型', import_trainset_data_with_mode),
      '先头车': ('先头车模型', import_locomotive_head_data_with_mode)
    }

    # 系统信息 sheet 名称映射
    system_sheets = {
      '品牌': ('品牌', import_brand_data_with_mode),
      '机务段': ('机务段', import_depot_data_with_mode),
      '车辆段': ('机务段', import_depot_data_with_mode),
      '商家': ('商家', import_merchant_data_with_mode),
      '动力类型': ('动力类型', import_power_type_data_with_mode),
      '芯片接口': ('芯片接口', import_chip_interface_data_with_mode),
      '芯片型号': ('芯片型号', import_chip_model_data_with_mode),
      '机车系列': ('机车系列', import_locomotive_series_data_with_mode),
      '机关系列': ('机车系列', import_locomotive_series_data_with_mode),
      '车厢系列': ('车厢系列', import_carriage_series_data_with_mode),
      '动车组系列': ('动车组系列', import_trainset_series_data_with_mode),
      '机车车型': ('机车车型', import_locomotive_model_data_with_mode),
      '车厢车型': ('车厢车型', import_carriage_model_data_with_mode),
      '动车组车型': ('动车组车型', import_trainset_model_data_with_mode)
    }

    for sheet_name, data in all_data.items():
      try:
        if sheet_name in model_sheets:
          display_name, import_func = model_sheets[sheet_name]
          count = import_func(data, mode)
          summary[display_name] = count
        elif sheet_name in system_sheets:
          display_name, import_func = system_sheets[sheet_name]
          count = import_func(data, mode)
          summary[display_name] = count
        else:
          logger.warning(f"Unknown sheet name: {sheet_name}")
      except Exception as e:
        db.session.rollback()
        errors.append(f"{sheet_name}: {str(e)}")
        logger.error(f"Error importing sheet {sheet_name}: {str(e)}", exc_info=True)

    if errors:
      return jsonify({'success': False, 'error': '部分导入失败: ' + '; '.join(errors)}), 400

    logger.info(f"Excel import completed: {summary}, mode={mode}")
    return jsonify({'success': True, 'summary': summary})

  except Exception as e:
    db.session.rollback()
    logger.error(f"Excel import failed: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': str(e)}), 500


def check_import_conflicts(all_data):
  """检查导入数据中的冲突"""
  conflicts = []

  # 检查系统信息名称冲突
  system_name_checks = [
    ('品牌', '品牌', Brand, 'name'),
    ('商家', '商家', Merchant, 'name'),
    ('动力类型', '动力类型', PowerType, 'name'),
    ('机务段', '机务段', Depot, 'name'),
    ('车辆段', '机务段', Depot, 'name'),
    ('芯片接口', '芯片接口', ChipInterface, 'name'),
    ('芯片型号', '芯片型号', ChipModel, 'name'),
    ('机车系列', '机车系列', LocomotiveSeries, 'name'),
    ('车厢系列', '车厢系列', CarriageSeries, 'name'),
    ('动车组系列', '动车组系列', TrainsetSeries, 'name'),
  ]

  for sheet_name, display_name, model, field in system_name_checks:
    if sheet_name in all_data:
      for row in all_data[sheet_name]:
        name = row.get('名称') or row.get(field) or row.get(display_name)
        if name:
          existing = model.query.filter_by(name=name).first()
          if existing:
            conflicts.append({
              'type': display_name,
              'field': '名称',
              'value': name,
              'message': f"{display_name} '{name}' 已存在"
            })

  # 检查机车冲突：同一比例内机车号或编号唯一
  if '机车' in all_data:
    for row in all_data['机车']:
      scale = row.get('比例')
      locomotive_number = row.get('机车号')
      decoder_number = row.get('编号')

      if scale and locomotive_number:
        existing = Locomotive.query.filter_by(scale=scale, locomotive_number=locomotive_number).first()
        if existing:
          conflicts.append({
            'type': '机车模型',
            'field': '机车号',
            'value': f'{scale} 比例 - {locomotive_number}',
            'message': f"机车号 '{locomotive_number}' 在比例 '{scale}' 中已存在"
          })

      if scale and decoder_number:
        existing = Locomotive.query.filter_by(scale=scale, decoder_number=decoder_number).first()
        if existing:
          conflicts.append({
            'type': '机车模型',
            'field': '编号',
            'value': f'{scale} 比例 - {decoder_number}',
            'message': f"编号 '{decoder_number}' 在比例 '{scale}' 中已存在"
          })

  # 检查动车组冲突：同一比例内动车号唯一
  if '动车组' in all_data:
    for row in all_data['动车组']:
      scale = row.get('比例')
      trainset_number = row.get('动车号')

      if scale and trainset_number:
        existing = Trainset.query.filter_by(scale=scale, trainset_number=trainset_number).first()
        if existing:
          conflicts.append({
            'type': '动车组模型',
            'field': '动车号',
            'value': f'{scale} 比例 - {trainset_number}',
            'message': f"动车号 '{trainset_number}' 在比例 '{scale}' 中已存在"
          })

  return conflicts


def import_locomotive_data_with_mode(data, mode='skip'):
  """导入机车模型数据，支持 skip/overwrite 模式"""
  count = 0
  for row in data:
    brand_name = row.get('品牌')
    scale = row.get('比例')
    if not brand_name or not scale:
      logger.warning(f"跳过机车行：缺少必填字段 品牌={brand_name}, 比例={scale}")
      continue

    brand_id = find_id_by_name(Brand, brand_name)
    if not brand_id:
      logger.warning(f"跳过机车行：找不到品牌 '{brand_name}'")
      continue

    series_id = find_id_by_name(LocomotiveSeries, row.get('系列'))
    power_type_id = find_id_by_name(PowerType, row.get('动力'))
    model_id = find_id_by_name(LocomotiveModel, row.get('车型'),
      lambda q: q.filter_by(name=row.get('车型'), series_id=series_id, power_type_id=power_type_id).first())

    locomotive_number = row.get('机车号') or None
    decoder_number = row.get('编号') or None

    # 检查冲突
    existing = None
    if scale and locomotive_number:
      existing = Locomotive.query.filter_by(scale=scale, locomotive_number=locomotive_number).first()
    if not existing and scale and decoder_number:
      existing = Locomotive.query.filter_by(scale=scale, decoder_number=decoder_number).first()

    if existing:
      if mode == 'skip':
        logger.info(f"跳过机车：冲突数据 比例={scale}, 机车号={locomotive_number}, 编号={decoder_number}")
        continue
      elif mode == 'overwrite':
        # 更新现有记录
        existing.series_id = series_id
        existing.power_type_id = power_type_id
        existing.model_id = model_id
        existing.brand_id = brand_id
        existing.depot_id = find_id_by_name(Depot, row.get('机务段'))
        existing.plaque = row.get('挂牌') or None
        existing.color = row.get('颜色') or None
        existing.decoder_number = decoder_number
        existing.chip_interface_id = find_id_by_name(ChipInterface, row.get('芯片接口'))
        existing.chip_model_id = find_id_by_name(ChipModel, row.get('芯片型号'))
        existing.price = row.get('价格') or None
        existing.total_price = calculate_price(row.get('价格')) if row.get('价格') else 0
        existing.item_number = row.get('货号') or None
        existing.purchase_date = parse_purchase_date(row.get('购买日期'))
        existing.merchant_id = find_id_by_name(Merchant, row.get('购买商家'))
        count += 1
        continue

    locomotive = Locomotive(
      series_id=series_id,
      power_type_id=power_type_id,
      model_id=model_id,
      brand_id=brand_id,
      depot_id=find_id_by_name(Depot, row.get('机务段')),
      plaque=row.get('挂牌') or None,
      color=row.get('颜色') or None,
      scale=scale,
      locomotive_number=locomotive_number,
      decoder_number=decoder_number,
      chip_interface_id=find_id_by_name(ChipInterface, row.get('芯片接口')),
      chip_model_id=find_id_by_name(ChipModel, row.get('芯片型号')),
      price=row.get('价格') or None,
      total_price=calculate_price(row.get('价格')) if row.get('价格') else 0,
      item_number=row.get('货号') or None,
      purchase_date=parse_purchase_date(row.get('购买日期')),
      merchant_id=find_id_by_name(Merchant, row.get('购买商家'))
    )
    db.session.add(locomotive)
    count += 1

  db.session.commit()
  return count


def import_carriage_data_with_mode(data, mode='skip'):
  """导入车厢模型数据 - 按套装分组处理，车厢没有唯一性约束，直接导入"""
  count = 0
  current_set_data = None
  current_items = []

  def save_carriage_set():
    """保存当前套装及其车厢项"""
    nonlocal count
    if not current_set_data:
      return

    brand_id = current_set_data.get('brand_id')
    scale = current_set_data.get('scale')
    if not brand_id or not scale:
      logger.warning(f"跳过车厢套装：缺少必填字段 brand_id={brand_id}, scale={scale}")
      return

    carriage_set = CarriageSet(
      brand_id=brand_id,
      series_id=current_set_data.get('series_id'),
      depot_id=current_set_data.get('depot_id'),
      train_number=current_set_data.get('train_number'),
      plaque=current_set_data.get('plaque'),
      item_number=current_set_data.get('item_number'),
      scale=scale,
      total_price=current_set_data.get('total_price', 0),
      purchase_date=current_set_data.get('purchase_date'),
      merchant_id=current_set_data.get('merchant_id')
    )
    db.session.add(carriage_set)
    db.session.flush()

    for item_data in current_items:
      if item_data.get('model_id'):
        item = CarriageItem(
          set_id=carriage_set.id,
          model_id=item_data.get('model_id'),
          car_number=item_data.get('car_number'),
          color=item_data.get('color'),
          lighting=item_data.get('lighting')
        )
        db.session.add(item)

    count += 1

  for row in data:
    brand_name = row.get('品牌')
    scale = row.get('比例')
    is_new_set = brand_name and scale

    if is_new_set:
      save_carriage_set()
      brand_id = find_id_by_name(Brand, brand_name)
      current_set_data = {
        'brand_id': brand_id,
        'series_id': find_id_by_name(CarriageSeries, row.get('系列')),
        'depot_id': find_id_by_name(Depot, row.get('车辆段')),
        'train_number': row.get('车次') or None,
        'plaque': row.get('挂牌') or None,
        'item_number': row.get('货号') or None,
        'scale': scale,
        'total_price': safe_float(row.get('总价')),
        'purchase_date': parse_purchase_date(row.get('购买日期')),
        'merchant_id': find_id_by_name(Merchant, row.get('购买商家'))
      }
      current_items = []

    model_name = row.get('车型')
    if model_name and current_set_data:
      model_id = find_id_by_name(CarriageModel, model_name)
      current_items.append({
        'model_id': model_id,
        'car_number': row.get('车辆号') or None,
        'color': row.get('颜色') or None,
        'lighting': row.get('灯光') or None
      })

  save_carriage_set()
  db.session.commit()
  return count


def import_trainset_data_with_mode(data, mode='skip'):
  """导入动车组模型数据，支持 skip/overwrite 模式"""
  count = 0
  for row in data:
    brand_name = row.get('品牌')
    scale = row.get('比例')
    if not brand_name or not scale:
      logger.warning(f"跳过动车组行：缺少必填字段 品牌={brand_name}, 比例={scale}")
      continue

    brand_id = find_id_by_name(Brand, brand_name)
    if not brand_id:
      logger.warning(f"跳过动车组行：找不到品牌 '{brand_name}'")
      continue

    series_id = find_id_by_name(TrainsetSeries, row.get('系列'))
    power_type_id = find_id_by_name(PowerType, row.get('动力'))
    model_id = find_id_by_name(TrainsetModel, row.get('车型'),
      lambda q: q.filter_by(name=row.get('车型'), series_id=series_id, power_type_id=power_type_id).first())

    trainset_number = row.get('动车号') or None

    # 检查冲突
    existing = None
    if scale and trainset_number:
      existing = Trainset.query.filter_by(scale=scale, trainset_number=trainset_number).first()

    if existing:
      if mode == 'skip':
        logger.info(f"跳过动车组：冲突数据 比例={scale}, 动车号={trainset_number}")
        continue
      elif mode == 'overwrite':
        existing.series_id = series_id
        existing.power_type_id = power_type_id
        existing.model_id = model_id
        existing.brand_id = brand_id
        existing.depot_id = find_id_by_name(Depot, row.get('动车段'))
        existing.plaque = row.get('挂牌') or None
        existing.color = row.get('颜色') or None
        existing.formation = safe_int(row.get('编组'))
        existing.decoder_number = row.get('编号') or None
        existing.head_light = parse_boolean(row.get('头车灯')) or False
        existing.interior_light = row.get('室内灯') or None
        existing.chip_interface_id = find_id_by_name(ChipInterface, row.get('芯片接口'))
        existing.chip_model_id = find_id_by_name(ChipModel, row.get('芯片型号'))
        existing.price = row.get('价格') or None
        existing.total_price = calculate_price(row.get('价格')) if row.get('价格') else 0
        existing.item_number = row.get('货号') or None
        existing.purchase_date = parse_purchase_date(row.get('购买日期'))
        existing.merchant_id = find_id_by_name(Merchant, row.get('购买商家'))
        count += 1
        continue

    trainset = Trainset(
      series_id=series_id,
      power_type_id=power_type_id,
      model_id=model_id,
      brand_id=brand_id,
      depot_id=find_id_by_name(Depot, row.get('动车段')),
      plaque=row.get('挂牌') or None,
      color=row.get('颜色') or None,
      scale=scale,
      formation=safe_int(row.get('编组')),
      trainset_number=trainset_number,
      decoder_number=row.get('编号') or None,
      head_light=parse_boolean(row.get('头车灯')) or False,
      interior_light=row.get('室内灯') or None,
      chip_interface_id=find_id_by_name(ChipInterface, row.get('芯片接口')),
      chip_model_id=find_id_by_name(ChipModel, row.get('芯片型号')),
      price=row.get('价格') or None,
      total_price=calculate_price(row.get('价格')) if row.get('价格') else 0,
      item_number=row.get('货号') or None,
      purchase_date=parse_purchase_date(row.get('购买日期')),
      merchant_id=find_id_by_name(Merchant, row.get('购买商家'))
    )
    db.session.add(trainset)
    count += 1

  db.session.commit()
  return count


def import_locomotive_head_data_with_mode(data, mode='skip'):
  """导入先头车模型数据，先头车没有唯一性约束，直接导入"""
  count = 0
  for row in data:
    brand_name = row.get('品牌')
    scale = row.get('比例')
    if not brand_name or not scale:
      logger.warning(f"跳过先头车行：缺少必填字段 品牌={brand_name}, 比例={scale}")
      continue

    brand_id = find_id_by_name(Brand, brand_name)
    if not brand_id:
      logger.warning(f"跳过先头车行：找不到品牌 '{brand_name}'")
      continue

    model_id = find_id_by_name(TrainsetModel, row.get('车型'))

    locomotive_head = LocomotiveHead(
      model_id=model_id,
      brand_id=brand_id,
      special_color=row.get('涂装') or None,
      scale=scale,
      head_light=parse_boolean(row.get('头车灯')) or False,
      interior_light=row.get('室内灯') or None,
      price=row.get('价格') or None,
      total_price=calculate_price(row.get('价格')) if row.get('价格') else 0,
      item_number=row.get('货号') or None,
      purchase_date=parse_purchase_date(row.get('购买日期')),
      merchant_id=find_id_by_name(Merchant, row.get('购买商家'))
    )
    db.session.add(locomotive_head)
    count += 1

  db.session.commit()
  return count


# 系统信息导入函数（支持 skip/overwrite 模式）
def import_brand_data_with_mode(data, mode='skip'):
  """导入品牌数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('品牌')
    if not name:
      continue
    existing = Brand.query.filter_by(name=name).first()
    if existing:
      if mode == 'skip':
        continue
      elif mode == 'overwrite':
        existing.search_url = row.get('搜索地址') or row.get('search_url') or existing.search_url
        count += 1
        continue
    brand = Brand(
      name=name,
      search_url=row.get('搜索地址') or row.get('search_url') or None
    )
    db.session.add(brand)
    count += 1
  db.session.commit()
  return count


def import_depot_data_with_mode(data, mode='skip'):
  """导入机务段/车辆段数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('机务段') or row.get('车辆段')
    if not name:
      continue
    if Depot.query.filter_by(name=name).first():
      if mode == 'skip':
        continue
      # overwrite 模式下无需更新，因为只有 name 字段
      continue
    depot = Depot(name=name)
    db.session.add(depot)
    count += 1
  db.session.commit()
  return count


def import_merchant_data_with_mode(data, mode='skip'):
  """导入商家数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('商家')
    if not name:
      continue
    if Merchant.query.filter_by(name=name).first():
      if mode == 'skip':
        continue
      continue
    merchant = Merchant(name=name)
    db.session.add(merchant)
    count += 1
  db.session.commit()
  return count


def import_power_type_data_with_mode(data, mode='skip'):
  """导入动力类型数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('动力类型')
    if not name:
      continue
    if PowerType.query.filter_by(name=name).first():
      if mode == 'skip':
        continue
      continue
    power_type = PowerType(name=name)
    db.session.add(power_type)
    count += 1
  db.session.commit()
  return count


def import_chip_interface_data_with_mode(data, mode='skip'):
  """导入芯片接口数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('芯片接口')
    if not name:
      continue
    if ChipInterface.query.filter_by(name=name).first():
      if mode == 'skip':
        continue
      continue
    chip_interface = ChipInterface(name=name)
    db.session.add(chip_interface)
    count += 1
  db.session.commit()
  return count


def import_chip_model_data_with_mode(data, mode='skip'):
  """导入芯片型号数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('芯片型号')
    if not name:
      continue
    if ChipModel.query.filter_by(name=name).first():
      if mode == 'skip':
        continue
      continue
    chip_model = ChipModel(name=name)
    db.session.add(chip_model)
    count += 1
  db.session.commit()
  return count


def import_locomotive_series_data_with_mode(data, mode='skip'):
  """导入机车系列数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('系列')
    if not name:
      continue
    if LocomotiveSeries.query.filter_by(name=name).first():
      if mode == 'skip':
        continue
      continue
    series = LocomotiveSeries(name=name)
    db.session.add(series)
    count += 1
  db.session.commit()
  return count


def import_carriage_series_data_with_mode(data, mode='skip'):
  """导入车厢系列数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('系列')
    if not name:
      continue
    if CarriageSeries.query.filter_by(name=name).first():
      if mode == 'skip':
        continue
      continue
    series = CarriageSeries(name=name)
    db.session.add(series)
    count += 1
  db.session.commit()
  return count


def import_trainset_series_data_with_mode(data, mode='skip'):
  """导入动车组系列数据"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('系列')
    if not name:
      continue
    if TrainsetSeries.query.filter_by(name=name).first():
      if mode == 'skip':
        continue
      continue
    series = TrainsetSeries(name=name)
    db.session.add(series)
    count += 1
  db.session.commit()
  return count


def import_locomotive_model_data_with_mode(data, mode='skip'):
  """导入机车车型数据，车型没有唯一名称约束，直接导入"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('车型')
    if not name:
      continue
    model = LocomotiveModel(
      name=name,
      series_id=find_id_by_name(LocomotiveSeries, row.get('系列')),
      power_type_id=find_id_by_name(PowerType, row.get('动力类型'))
    )
    db.session.add(model)
    count += 1
  db.session.commit()
  return count


def import_carriage_model_data_with_mode(data, mode='skip'):
  """导入车厢车型数据，车型没有唯一名称约束，直接导入"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('车型')
    if not name:
      continue
    model = CarriageModel(
      name=name,
      series_id=find_id_by_name(CarriageSeries, row.get('系列')),
      type=row.get('类型') or '客车'
    )
    db.session.add(model)
    count += 1
  db.session.commit()
  return count


def import_trainset_model_data_with_mode(data, mode='skip'):
  """导入动车组车型数据，车型没有唯一名称约束，直接导入"""
  count = 0
  for row in data:
    name = row.get('名称') or row.get('车型')
    if not name:
      continue
    model = TrainsetModel(
      name=name,
      series_id=find_id_by_name(TrainsetSeries, row.get('系列')),
      power_type_id=find_id_by_name(PowerType, row.get('动力类型'))
    )
    db.session.add(model)
    count += 1
  db.session.commit()
  return count


@api_bp.route('/api/export/excel')
def export_to_excel():
  """导出数据到Excel - 支持 mode 参数: models(模型数据), system(系统信息), all(全部)"""
  try:
    mode = request.args.get('mode', 'models')

    # 检查是否有数据
    has_model_data = (
      Locomotive.query.count() > 0 or
      CarriageSet.query.count() > 0 or
      Trainset.query.count() > 0 or
      LocomotiveHead.query.count() > 0
    )
    has_system_data = (
      Brand.query.count() > 0 or
      Depot.query.count() > 0 or
      Merchant.query.count() > 0 or
      PowerType.query.count() > 0 or
      ChipInterface.query.count() > 0 or
      ChipModel.query.count() > 0
    )

    if mode == 'models' and not has_model_data:
      return jsonify({'success': False, 'error': '当前没有可导出的模型数据'}), 400
    if mode == 'system' and not has_system_data:
      return jsonify({'success': False, 'error': '当前没有可导出的系统信息'}), 400
    if mode == 'all' and not has_model_data and not has_system_data:
      return jsonify({'success': False, 'error': '当前没有可导出的数据'}), 400

    workbook = openpyxl.Workbook()

    if 'Sheet' in workbook.sheetnames:
      workbook.remove(workbook['Sheet'])

    # 导出模型数据（models 或 all 模式）
    if mode in ('models', 'all'):
      # 导出机车模型
      if Locomotive.query.count() > 0:
        sheet = workbook.create_sheet('机车')
        headers = ['系列', '动力', '车型', '品牌', '机务段', '挂牌', '颜色', '比例', '机车号', '编号',
               '芯片接口', '芯片型号', '价格', '总价', '货号', '购买日期', '购买商家']
        sheet.append(headers)

        for loco in Locomotive.query.all():
          sheet.append([
            loco.series.name if loco.series else '',
            loco.power_type.name if loco.power_type else '',
            loco.model.name if loco.model else '',
            loco.brand.name if loco.brand else '',
            loco.depot.name if loco.depot else '',
            loco.plaque or '',
            loco.color or '',
            loco.scale or '',
            loco.locomotive_number or '',
            loco.decoder_number or '',
            loco.chip_interface.name if loco.chip_interface else '',
            loco.chip_model.name if loco.chip_model else '',
            loco.price or '',
            loco.total_price or '',
            loco.item_number or '',
            loco.purchase_date.strftime('%Y-%m-%d') if loco.purchase_date else '',
            loco.merchant.name if loco.merchant else ''
          ])

      # 导出车厢模型（使用合并单元格标识套装）
      if CarriageSet.query.count() > 0:
        sheet = workbook.create_sheet('车厢')
        headers = ['品牌', '系列', '车辆段', '车次', '挂牌', '货号', '比例', '车型', '车辆号', '颜色', '灯光', '总价', '购买日期', '购买商家']
        sheet.append(headers)

        current_row = 2  # 从第2行开始（第1行是表头）
        for carriage_set in CarriageSet.query.all():
          items = carriage_set.items
          if not items:
            # 无车厢项的套装，单独一行
            sheet.append([
              carriage_set.brand.name if carriage_set.brand else '',
              carriage_set.series.name if carriage_set.series else '',
              carriage_set.depot.name if carriage_set.depot else '',
              carriage_set.train_number or '',
              carriage_set.plaque or '',
              carriage_set.item_number or '',
              carriage_set.scale or '',
              '', '', '', '',
              carriage_set.total_price or '',
              carriage_set.purchase_date.strftime('%Y-%m-%d') if carriage_set.purchase_date else '',
              carriage_set.merchant.name if carriage_set.merchant else ''
            ])
            current_row += 1
          else:
            # 有车厢项的套装，合并公共信息列
            start_row = current_row
            for item in items:
              sheet.append([
                carriage_set.brand.name if carriage_set.brand else '',
                carriage_set.series.name if carriage_set.series else '',
                carriage_set.depot.name if carriage_set.depot else '',
                carriage_set.train_number or '',
                carriage_set.plaque or '',
                carriage_set.item_number or '',
                carriage_set.scale or '',
                item.model.name if item.model else '',
                item.car_number or '',
                item.color or '',
                item.lighting or '',
                carriage_set.total_price or '',
                carriage_set.purchase_date.strftime('%Y-%m-%d') if carriage_set.purchase_date else '',
                carriage_set.merchant.name if carriage_set.merchant else ''
              ])
              current_row += 1

            # 合并公共信息列（前7列：A-G，即品牌到比例）
            # 合并总价、购买日期、购买商家列（L-N，即第12-14列）
            if len(items) > 1:
              end_row = current_row - 1
              # 合并前7列（品牌、系列、车辆段、车次、挂牌、货号、比例）
              for col in range(1, 8):  # A-G 列
                sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
              # 合并后3列（总价、购买日期、购买商家）
              for col in range(12, 15):  # L-N 列
                sheet.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)

      # 导出动车组模型
      if Trainset.query.count() > 0:
        sheet = workbook.create_sheet('动车组')
        headers = ['系列', '动力', '车型', '品牌', '动车段', '挂牌', '颜色', '比例', '编组', '动车号', '编号',
               '头车灯', '室内灯', '芯片接口', '芯片型号', '价格', '总价', '货号', '购买日期', '购买商家']
        sheet.append(headers)

        for ts in Trainset.query.all():
          sheet.append([
            ts.series.name if ts.series else '',
            ts.power_type.name if ts.power_type else '',
            ts.model.name if ts.model else '',
            ts.brand.name if ts.brand else '',
            ts.depot.name if ts.depot else '',
            ts.plaque or '',
            ts.color or '',
            ts.scale or '',
            ts.formation or '',
            ts.trainset_number or '',
            ts.decoder_number or '',
            '是' if ts.head_light else '否',
            ts.interior_light or '',
            ts.chip_interface.name if ts.chip_interface else '',
            ts.chip_model.name if ts.chip_model else '',
            ts.price or '',
            ts.total_price or '',
            ts.item_number or '',
            ts.purchase_date.strftime('%Y-%m-%d') if ts.purchase_date else '',
            ts.merchant.name if ts.merchant else ''
          ])

      # 导出先头车模型
      if LocomotiveHead.query.count() > 0:
        sheet = workbook.create_sheet('先头车')
        headers = ['车型', '品牌', '涂装', '比例', '头车灯', '室内灯', '价格', '总价', '货号', '购买日期', '购买商家']
        sheet.append(headers)

        for head in LocomotiveHead.query.all():
          sheet.append([
            head.model.name if head.model else '',
            head.brand.name if head.brand else '',
            head.special_color or '',
            head.scale or '',
            '是' if head.head_light else '否',
            head.interior_light or '',
            head.price or '',
            head.total_price or '',
            head.item_number or '',
            head.purchase_date.strftime('%Y-%m-%d') if head.purchase_date else '',
            head.merchant.name if head.merchant else ''
          ])

    # 导出系统信息（system 或 all 模式）
    if mode in ('system', 'all'):
      # 导出品牌
      if Brand.query.count() > 0:
        sheet = workbook.create_sheet('品牌')
        sheet.append(['名称', '搜索地址'])
        for brand in Brand.query.all():
          sheet.append([brand.name, brand.search_url or ''])

      # 导出机务段
      if Depot.query.count() > 0:
        sheet = workbook.create_sheet('机务段')
        sheet.append(['名称'])
        for depot in Depot.query.all():
          sheet.append([depot.name])

      # 导出商家
      if Merchant.query.count() > 0:
        sheet = workbook.create_sheet('商家')
        sheet.append(['名称'])
        for merchant in Merchant.query.all():
          sheet.append([merchant.name])

      # 导出动力类型
      if PowerType.query.count() > 0:
        sheet = workbook.create_sheet('动力类型')
        sheet.append(['名称'])
        for pt in PowerType.query.all():
          sheet.append([pt.name])

      # 导出芯片接口
      if ChipInterface.query.count() > 0:
        sheet = workbook.create_sheet('芯片接口')
        sheet.append(['名称'])
        for ci in ChipInterface.query.all():
          sheet.append([ci.name])

      # 导出芯片型号
      if ChipModel.query.count() > 0:
        sheet = workbook.create_sheet('芯片型号')
        sheet.append(['名称'])
        for cm in ChipModel.query.all():
          sheet.append([cm.name])

      # 导出机车系列
      if LocomotiveSeries.query.count() > 0:
        sheet = workbook.create_sheet('机车系列')
        sheet.append(['名称'])
        for series in LocomotiveSeries.query.all():
          sheet.append([series.name])

      # 导出车厢系列
      if CarriageSeries.query.count() > 0:
        sheet = workbook.create_sheet('车厢系列')
        sheet.append(['名称'])
        for series in CarriageSeries.query.all():
          sheet.append([series.name])

      # 导出动车组系列
      if TrainsetSeries.query.count() > 0:
        sheet = workbook.create_sheet('动车组系列')
        sheet.append(['名称'])
        for series in TrainsetSeries.query.all():
          sheet.append([series.name])

      # 导出机车车型
      if LocomotiveModel.query.count() > 0:
        sheet = workbook.create_sheet('机车车型')
        sheet.append(['名称', '系列', '动力类型'])
        for model in LocomotiveModel.query.all():
          sheet.append([
            model.name,
            model.series.name if model.series else '',
            model.power_type.name if model.power_type else ''
          ])

      # 导出车厢车型
      if CarriageModel.query.count() > 0:
        sheet = workbook.create_sheet('车厢车型')
        sheet.append(['名称', '系列', '类型'])
        for model in CarriageModel.query.all():
          sheet.append([
            model.name,
            model.series.name if model.series else '',
            model.type or ''
          ])

      # 导出动车组车型
      if TrainsetModel.query.count() > 0:
        sheet = workbook.create_sheet('动车组车型')
        sheet.append(['名称', '系列', '动力类型'])
        for model in TrainsetModel.query.all():
          sheet.append([
            model.name,
            model.series.name if model.series else '',
            model.power_type.name if model.power_type else ''
          ])

    # 加粗所有工作表的第一行（标题行）
    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    for sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
      for cell in sheet[1]:
        cell.font = bold_font

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # 根据模式设置文件名
    mode_names = {'models': 'Models', 'system': 'System', 'all': 'All'}
    filename = f'TMM_{mode_names.get(mode, "Export")}_{datetime.now().strftime("%Y%m%d_%H%M%S")}_{random.randint(1000, 9999)}.xlsx'

    logger.info(f"Excel export completed successfully, mode={mode}")
    return send_file(
      output,
      as_attachment=True,
      download_name=filename,
      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

  except Exception as e:
    logger.error(f"Excel export failed: {str(e)}", exc_info=True)
    return f"导出失败: {str(e)}", 500


# ==================== 自定义导入模板 CRUD API ====================

@api_bp.route('/api/import-templates', methods=['GET'])
def list_import_templates():
  """
  获取所有导入模板

  Returns:
    JSON: {
      'success': True,
      'templates': [{'id', 'name', 'config', 'created_at', 'updated_at'}, ...]
    }
  """
  try:
    templates = ImportTemplate.query.order_by(ImportTemplate.updated_at.desc()).all()
    return jsonify({
      'success': True,
      'templates': [{
        'id': t.id,
        'name': t.name,
        'config': t.config,
        'created_at': t.created_at.isoformat() if t.created_at else None,
        'updated_at': t.updated_at.isoformat() if t.updated_at else None
      } for t in templates]
    })
  except Exception as e:
    logger.error(f"Failed to list import templates: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


@api_bp.route('/api/import-templates', methods=['POST'])
def create_import_template():
  """
  创建导入模板

  Request JSON:
    {
      'name': str,     # 必填，模板名称
      'config': dict   # 必填，映射配置
    }

  Returns:
    JSON: {
      'success': True,
      'template': {'id', 'name', 'config', 'created_at', 'updated_at'}
    }
  """
  try:
    data = request.get_json()
    if not data:
      return jsonify({'success': False, 'error': '请求体不能为空'}), 400

    name = data.get('name')
    config = data.get('config')

    if not name:
      return jsonify({'success': False, 'error': '模板名称不能为空'}), 400
    if config is None:
      return jsonify({'success': False, 'error': '模板配置不能为空'}), 400
    if not isinstance(config, dict):
      return jsonify({'success': False, 'error': '模板配置必须是JSON对象'}), 400

    template = ImportTemplate(name=name, config=config)
    db.session.add(template)
    db.session.commit()

    logger.info(f"Created import template: {template.id} - {template.name}")
    return jsonify({
      'success': True,
      'template': {
        'id': template.id,
        'name': template.name,
        'config': template.config,
        'created_at': template.created_at.isoformat() if template.created_at else None,
        'updated_at': template.updated_at.isoformat() if template.updated_at else None
      }
    }), 201

  except Exception as e:
    db.session.rollback()
    logger.error(f"Failed to create import template: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


@api_bp.route('/api/import-templates/<int:template_id>', methods=['GET'])
def get_import_template(template_id):
  """
  获取单个导入模板

  Args:
    template_id: 模板ID

  Returns:
    JSON: {
      'success': True,
      'template': {'id', 'name', 'config', 'created_at', 'updated_at'}
    }
  """
  try:
    template = db.session.get(ImportTemplate, template_id)
    if not template:
      return jsonify({'success': False, 'error': '模板不存在'}), 404

    return jsonify({
      'success': True,
      'template': {
        'id': template.id,
        'name': template.name,
        'config': template.config,
        'created_at': template.created_at.isoformat() if template.created_at else None,
        'updated_at': template.updated_at.isoformat() if template.updated_at else None
      }
    })

  except Exception as e:
    logger.error(f"Failed to get import template: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


@api_bp.route('/api/import-templates/<int:template_id>', methods=['PUT'])
def update_import_template(template_id):
  """
  更新导入模板

  Args:
    template_id: 模板ID

  Request JSON:
    {
      'name': str,     # 可选，模板名称
      'config': dict   # 可选，映射配置
    }

  Returns:
    JSON: {
      'success': True,
      'template': {'id', 'name', 'config', 'created_at', 'updated_at'}
    }
  """
  try:
    template = db.session.get(ImportTemplate, template_id)
    if not template:
      return jsonify({'success': False, 'error': '模板不存在'}), 404

    data = request.get_json()
    if not data:
      return jsonify({'success': False, 'error': '请求体不能为空'}), 400

    if 'name' in data:
      template.name = data['name']
    if 'config' in data:
      config = data['config']
      if not isinstance(config, dict):
        return jsonify({'success': False, 'error': '模板配置必须是JSON对象'}), 400
      template.config = config

    db.session.commit()

    logger.info(f"Updated import template: {template.id} - {template.name}")
    return jsonify({
      'success': True,
      'template': {
        'id': template.id,
        'name': template.name,
        'config': template.config,
        'created_at': template.created_at.isoformat() if template.created_at else None,
        'updated_at': template.updated_at.isoformat() if template.updated_at else None
      }
    })

  except Exception as e:
    db.session.rollback()
    logger.error(f"Failed to update import template: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


@api_bp.route('/api/import-templates/<int:template_id>', methods=['DELETE'])
def delete_import_template(template_id):
  """
  删除导入模板

  Args:
    template_id: 模板ID

  Returns:
    JSON: {'success': True, 'message': '删除成功'}
  """
  try:
    template = db.session.get(ImportTemplate, template_id)
    if not template:
      return jsonify({'success': False, 'error': '模板不存在'}), 404

    template_name = template.name
    db.session.delete(template)
    db.session.commit()

    logger.info(f"Deleted import template: {template_id} - {template_name}")
    return jsonify({'success': True, 'message': '删除成功'})

  except Exception as e:
    db.session.rollback()
    logger.error(f"Failed to delete import template: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


@api_bp.route('/api/import-templates/<int:template_id>/copy', methods=['POST'])
def copy_import_template(template_id):
  """
  复制导入模板

  Args:
    template_id: 模板ID

  Request Body:
    {
      'name': str  # 新模板名称
    }

  Returns:
    JSON: {'success': True, 'template': {...}}
  """
  try:
    template = db.session.get(ImportTemplate, template_id)
    if not template:
      return jsonify({'success': False, 'error': '模板不存在'}), 404

    data = request.get_json()
    if not data or not data.get('name'):
      return jsonify({'success': False, 'error': '新模板名称不能为空'}), 400

    new_name = data['name'].strip()
    if not new_name:
      return jsonify({'success': False, 'error': '新模板名称不能为空'}), 400

    # 检查名称是否已存在
    existing = ImportTemplate.query.filter_by(name=new_name).first()
    if existing:
      return jsonify({'success': False, 'error': '模板名称已存在'}), 400

    # 创建新模板
    new_template = ImportTemplate(
      name=new_name,
      config=template.config.copy() if template.config else {}
    )
    db.session.add(new_template)
    db.session.commit()

    logger.info(f"Copied import template: {template_id} -> {new_template.id} ({new_name})")

    return jsonify({
      'success': True,
      'template': {
        'id': new_template.id,
        'name': new_template.name,
        'config': new_template.config,
        'created_at': new_template.created_at.isoformat() if new_template.created_at else None,
        'updated_at': new_template.updated_at.isoformat() if new_template.updated_at else None
      }
    })

  except Exception as e:
    db.session.rollback()
    logger.error(f"Failed to copy import template: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


@api_bp.route('/api/custom-import/tables', methods=['GET'])
def get_custom_import_tables():
  """
  获取自定义导入可用的系统表配置

  返回所有可映射的系统表和模型表信息，用于前端下拉菜单

  Returns:
    JSON: {
      'success': True,
      'tables': [
        {
          'name': str,           # 表名
          'display_name': str,   # 显示名称
          'category': str,       # 类别: 'system' 或 'model'
          'tooltip': str,        # 可选，提示文本
          'has_set_detection': bool  # 可选，是否支持套装检测
        },
        ...
      ]
    }
  """
  try:
    tables = get_table_display_info()
    return jsonify({
      'success': True,
      'tables': tables
    })

  except Exception as e:
    logger.error(f"Failed to get custom import tables: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


@api_bp.route('/api/custom-import/parse', methods=['POST'])
def parse_custom_import_file():
  """
  解析 Excel 文件，返回 Sheet 和列信息

  Request:
    multipart/form-data with 'file' field containing Excel file

  Returns:
    JSON: {
      'success': True,
      'filename': str,
      'sheets': [
        {
          'name': str,       # Sheet 名称
          'columns': list,   # 列名列表
          'row_count': int   # 数据行数
        },
        ...
      ]
    }
  """
  try:
    if 'file' not in request.files:
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
      return jsonify({'success': False, 'error': '文件格式错误，请上传Excel文件'}), 400

    workbook = openpyxl.load_workbook(file)
    sheets = []

    for sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
      columns = []
      # Read first row as column headers
      for cell in sheet[1]:
        if cell.value is not None:
          columns.append(str(cell.value))

      # Count data rows
      row_count = 0
      for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
          row_count += 1

      sheets.append({
        'name': sheet_name,
        'columns': columns,
        'row_count': row_count
      })

    logger.info(f"Parsed Excel file: {file.filename}, {len(sheets)} sheets")
    return jsonify({
      'success': True,
      'filename': file.filename,
      'sheets': sheets
    })

  except Exception as e:
    logger.error(f"Parse Excel failed: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


# 模型类映射
MODEL_CLASS_MAP = {
  'brand': Brand,
  'depot': Depot,
  'merchant': Merchant,
  'power_type': PowerType,
  'chip_interface': ChipInterface,
  'chip_model': ChipModel,
  'locomotive_series': LocomotiveSeries,
  'carriage_series': CarriageSeries,
  'trainset_series': TrainsetSeries,
  'locomotive': Locomotive,
  'carriage': CarriageSet,
  'trainset': Trainset,
  'locomotive_head': LocomotiveHead,
  'carriage_model': CarriageModel,
  'locomotive_model': LocomotiveModel,
  'trainset_model': TrainsetModel
}


@api_bp.route('/api/custom-import/preview', methods=['POST'])
def preview_custom_import():
  """
  预览自定义导入数据，检测冲突

  Request:
    multipart/form-data with:
      - 'file': Excel file
      - 'config': JSON string containing mapping configuration

  Config structure:
    {
      'sheet_mappings': [
        {'sheet_name': '品牌列表', 'table_name': 'brand'},
        ...
      ],
      'column_mappings': {
        'brand': {
          'columns': [
            {'source': '品牌名称', 'target': 'name', 'required': True},
            {'source': '官网', 'target': 'search_url', 'required': False}
          ],
          'conflict_mode': 'overwrite'
        },
        ...
      }
    }

  Returns:
    JSON: {
      'success': True,
      'previews': [
        {
          'table_name': 'brand',
          'display_name': '品牌',
          'row_count': 10,
          'conflicts': [
            {'type': '唯一名称冲突', 'field': 'name', 'value': 'xxx', 'message': '...'}
          ],
          'warnings': ['未映射字段: search_url'],
          'missing_required': ['scale']
        }
      ],
      'has_conflicts': False,
      'can_proceed': True
    }
  """
  try:
    # 验证文件
    if 'file' not in request.files:
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
      return jsonify({'success': False, 'error': '文件格式错误，请上传Excel文件'}), 400

    # 验证配置
    config_str = request.form.get('config')
    if not config_str:
      return jsonify({'success': False, 'error': '缺少映射配置'}), 400

    try:
      config = json.loads(config_str)
    except json.JSONDecodeError:
      return jsonify({'success': False, 'error': '配置格式错误，必须是有效的JSON'}), 400

    # 解析 Excel 文件
    workbook = openpyxl.load_workbook(file)

    # 构建工作表数据
    sheets_data = {}
    for sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
      headers = []
      data = []

      for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_idx == 1:
          headers = [str(cell) if cell is not None else '' for cell in row]
        else:
          if any(cell is not None for cell in row):
            row_dict = {}
            for i, cell in enumerate(row):
              if i < len(headers) and headers[i]:
                row_dict[headers[i]] = cell
            data.append(row_dict)

      sheets_data[sheet_name] = {'headers': headers, 'data': data}

    # 处理预览
    previews = []
    has_conflicts = False
    can_proceed = True

    sheet_mappings = config.get('sheet_mappings', [])
    column_mappings = config.get('column_mappings', {})

    for mapping in sheet_mappings:
      sheet_name = mapping.get('sheet_name')
      table_name = mapping.get('table_name')

      if not sheet_name or not table_name:
        continue

      # 获取表配置
      table_config = SYSTEM_TABLES.get(table_name)
      if not table_config:
        logger.warning(f"Unknown table name: {table_name}")
        continue

      # 获取列映射
      column_mapping = column_mappings.get(table_name, {})
      columns = column_mapping.get('columns', [])

      # 构建源列名到目标字段名的映射
      source_to_target = {}
      for col in columns:
        source = col.get('source')
        target = col.get('target')
        if source and target:
          source_to_target[source] = target

      # 检查缺失的必填字段
      missing_required = []
      mapped_targets = set(source_to_target.values())
      for field in table_config.get('fields', []):
        if field.get('required') and field.get('name') not in mapped_targets:
          missing_required.append(field.get('name'))

      if missing_required:
        can_proceed = False

      # 生成警告（未映射的可选字段）
      warnings = []
      for field in table_config.get('fields', []):
        field_name = field.get('name')
        if not field.get('required') and field_name not in mapped_targets:
          warnings.append(f"未映射字段: {field_name}")

      # 获取工作表数据
      sheet_data = sheets_data.get(sheet_name, {})
      rows = sheet_data.get('data', [])
      row_count = len(rows)

      # 检测冲突
      conflicts = []
      model_class = MODEL_CLASS_MAP.get(table_name)

      if model_class and row_count > 0:
        for row in rows:
          # 映射行数据到目标字段
          mapped_row = {}
          for source_col, target_field in source_to_target.items():
            value = row.get(source_col)
            mapped_row[target_field] = value

          # 检查唯一约束
          for field in table_config.get('fields', []):
            field_name = field.get('name')

            # 检查 unique 约束
            if field.get('unique'):
              value = mapped_row.get(field_name)
              if value:
                existing = model_class.query.filter_by(**{field_name: value}).first()
                if existing:
                  conflicts.append({
                    'type': '唯一名称冲突',
                    'field': field_name,
                    'value': str(value),
                    'message': f"{table_config.get('display_name')} '{value}' 已存在"
                  })
                  has_conflicts = True

            # 检查 unique_in_scale 约束
            if field.get('unique_in_scale'):
              value = mapped_row.get(field_name)
              scale = mapped_row.get('scale')
              if value and scale:
                existing = model_class.query.filter_by(
                  scale=scale,
                  **{field_name: value}
                ).first()
                if existing:
                  conflicts.append({
                    'type': '比例内唯一冲突',
                    'field': field_name,
                    'value': f"{scale} 比例 - {value}",
                    'message': f"{field.get('display', field_name)} '{value}' 在比例 '{scale}' 中已存在"
                  })
                  has_conflicts = True

      previews.append({
        'table_name': table_name,
        'display_name': table_config.get('display_name', table_name),
        'row_count': row_count,
        'conflicts': conflicts,
        'warnings': warnings,
        'missing_required': missing_required
      })

    logger.info(f"Preview completed: {len(previews)} tables, has_conflicts={has_conflicts}, can_proceed={can_proceed}")

    return jsonify({
      'success': True,
      'previews': previews,
      'has_conflicts': has_conflicts,
      'can_proceed': can_proceed
    })

  except Exception as e:
    logger.error(f"Preview custom import failed: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500


# 外键引用模型映射
FK_REFERENCE_MAP = {
  'brand': Brand,
  'depot': Depot,
  'merchant': Merchant,
  'power_type': PowerType,
  'chip_interface': ChipInterface,
  'chip_model': ChipModel,
  'locomotive_series': LocomotiveSeries,
  'carriage_series': CarriageSeries,
  'trainset_series': TrainsetSeries,
  'locomotive_model': LocomotiveModel,
  'carriage_model': CarriageModel,
  'trainset_model': TrainsetModel
}


def resolve_foreign_key(ref_table, value):
  """
  解析外键引用，将名称转换为 ID（大小写不敏感）

  Args:
    ref_table: 参考表名
    value: 名称值

  Returns:
    int or None: 找到的 ID，如果未找到则返回 None
  """
  if not value:
    return None

  model_class = FK_REFERENCE_MAP.get(ref_table)
  if not model_class:
    return None

  # 大小写不敏感匹配
  obj = model_class.query.filter(
    db.func.lower(model_class.name) == db.func.lower(value)
  ).first()
  return obj.id if obj else None


def check_unique_conflict(model_class, field_name, value, scale=None):
  """
  检查唯一约束冲突（大小写不敏感）

  Args:
    model_class: 模型类
    field_name: 字段名
    value: 值
    scale: 比例（用于 unique_in_scale 约束）

  Returns:
    object or None: 冲突的现有记录，无冲突则返回 None
  """
  if not value:
    return None

  # 获取字段属性
  field_attr = getattr(model_class, field_name)

  if scale:
    return model_class.query.filter(
      model_class.scale == scale,
      db.func.lower(field_attr) == db.func.lower(value)
    ).first()
  else:
    return model_class.query.filter(
      db.func.lower(field_attr) == db.func.lower(value)
    ).first()


def execute_system_table_import(table_name, table_config, rows, source_to_target, conflict_mode):
  """
  执行系统表数据导入

  Args:
    table_name: 表名
    table_config: 表配置
    rows: 数据行列表
    source_to_target: 源列到目标字段的映射
    conflict_mode: 冲突处理模式 ('skip' 或 'overwrite')

  Returns:
    int: 导入的记录数
  """
  model_class = MODEL_CLASS_MAP.get(table_name)
  if not model_class:
    return 0

  count = 0
  for row in rows:
    # 映射行数据到目标字段
    mapped_row = {}
    for source_col, target_field in source_to_target.items():
      value = row.get(source_col)
      mapped_row[target_field] = value

    # 获取名称字段值（用于唯一性检查）
    name = mapped_row.get('name')
    if not name:
      continue

    # 检查是否存在
    existing = model_class.query.filter_by(name=name).first()

    if existing:
      if conflict_mode == 'skip':
        continue
      elif conflict_mode == 'overwrite':
        # 更新现有记录
        for field, value in mapped_row.items():
          if field != 'name' and value is not None:
            setattr(existing, field, value)
        count += 1
        continue

    # 创建新记录
    instance = model_class(**mapped_row)
    db.session.add(instance)
    count += 1

  db.session.commit()
  return count


def execute_locomotive_import(rows, source_to_target, field_configs, conflict_mode):
  """
  执行机车模型导入

  Args:
    rows: 数据行列表
    source_to_target: 源列到目标字段的映射
    field_configs: 字段配置字典
    conflict_mode: 冲突处理模式

  Returns:
    int: 导入的记录数
  """
  count = 0
  for row in rows:
    # 映射行数据并解析外键
    mapped_data = {}
    for source_col, target_field in source_to_target.items():
      value = row.get(source_col)
      field_config = field_configs.get(target_field, {})
      ref_table = field_config.get('ref')

      if ref_table:
        # 解析外键
        resolved_id = resolve_foreign_key(ref_table, value)
        if field_config.get('required') and resolved_id is None and value:
          logger.warning(f"跳过机车行：找不到 {ref_table} '{value}'")
          mapped_data = None
          break
        mapped_data[target_field] = resolved_id
      else:
        mapped_data[target_field] = value

    if not mapped_data:
      continue

    # 检查必填字段
    brand_id = mapped_data.get('brand_id')
    scale = mapped_data.get('scale')
    if not brand_id or not scale:
      logger.warning(f"跳过机车行：缺少必填字段 brand_id={brand_id}, scale={scale}")
      continue

    # 检查冲突（locomotive_number 或 decoder_number 在比例内唯一）
    existing = None
    locomotive_number = mapped_data.get('locomotive_number')
    decoder_number = mapped_data.get('decoder_number')

    if scale and locomotive_number:
      existing = Locomotive.query.filter_by(scale=scale, locomotive_number=locomotive_number).first()
    if not existing and scale and decoder_number:
      existing = Locomotive.query.filter_by(scale=scale, decoder_number=decoder_number).first()

    if existing:
      if conflict_mode == 'skip':
        logger.info(f"跳过机车：冲突数据 比例={scale}, 机车号={locomotive_number}, 编号={decoder_number}")
        continue
      elif conflict_mode == 'overwrite':
        # 更新现有记录
        for field, value in mapped_data.items():
          if value is not None:
            if field == 'price':
              existing.price = value
              existing.total_price = calculate_price(value) if value else 0
            elif field == 'purchase_date':
              existing.purchase_date = parse_purchase_date(value)
            else:
              setattr(existing, field, value)
        count += 1
        continue

    # 创建新记录
    price = mapped_data.get('price')
    locomotive = Locomotive(
      series_id=mapped_data.get('series_id'),
      power_type_id=mapped_data.get('power_type_id'),
      model_id=mapped_data.get('model_id'),
      brand_id=brand_id,
      depot_id=mapped_data.get('depot_id'),
      plaque=mapped_data.get('plaque') or None,
      color=mapped_data.get('color') or None,
      scale=scale,
      locomotive_number=locomotive_number or None,
      decoder_number=mapped_data.get('decoder_number') or None,
      chip_interface_id=mapped_data.get('chip_interface_id'),
      chip_model_id=mapped_data.get('chip_model_id'),
      price=price or None,
      total_price=calculate_price(price) if price else 0,
      item_number=mapped_data.get('item_number') or None,
      purchase_date=parse_purchase_date(mapped_data.get('purchase_date')),
      merchant_id=mapped_data.get('merchant_id')
    )
    db.session.add(locomotive)
    count += 1

  db.session.commit()
  return count


def execute_trainset_import(rows, source_to_target, field_configs, conflict_mode):
  """
  执行动车组模型导入

  Args:
    rows: 数据行列表
    source_to_target: 源列到目标字段的映射
    field_configs: 字段配置字典
    conflict_mode: 冲突处理模式

  Returns:
    int: 导入的记录数
  """
  count = 0
  for row in rows:
    # 映射行数据并解析外键
    mapped_data = {}
    for source_col, target_field in source_to_target.items():
      value = row.get(source_col)
      field_config = field_configs.get(target_field, {})
      ref_table = field_config.get('ref')

      if ref_table:
        resolved_id = resolve_foreign_key(ref_table, value)
        if field_config.get('required') and resolved_id is None and value:
          logger.warning(f"跳过动车组行：找不到 {ref_table} '{value}'")
          mapped_data = None
          break
        mapped_data[target_field] = resolved_id
      else:
        mapped_data[target_field] = value

    if not mapped_data:
      continue

    # 检查必填字段
    brand_id = mapped_data.get('brand_id')
    scale = mapped_data.get('scale')
    if not brand_id or not scale:
      logger.warning(f"跳过动车组行：缺少必填字段 brand_id={brand_id}, scale={scale}")
      continue

    # 检查冲突（trainset_number 在比例内唯一）
    existing = None
    trainset_number = mapped_data.get('trainset_number')

    if scale and trainset_number:
      existing = Trainset.query.filter_by(scale=scale, trainset_number=trainset_number).first()

    if existing:
      if conflict_mode == 'skip':
        logger.info(f"跳过动车组：冲突数据 比例={scale}, 动车号={trainset_number}")
        continue
      elif conflict_mode == 'overwrite':
        # 更新现有记录
        for field, value in mapped_data.items():
          if value is not None:
            if field == 'price':
              existing.price = value
              existing.total_price = calculate_price(value) if value else 0
            elif field == 'purchase_date':
              existing.purchase_date = parse_purchase_date(value)
            elif field == 'head_light':
              existing.head_light = parse_boolean(value) or False
            elif field == 'formation':
              existing.formation = safe_int(value)
            else:
              setattr(existing, field, value)
        count += 1
        continue

    # 创建新记录
    price = mapped_data.get('price')
    trainset = Trainset(
      series_id=mapped_data.get('series_id'),
      power_type_id=mapped_data.get('power_type_id'),
      model_id=mapped_data.get('model_id'),
      brand_id=brand_id,
      depot_id=mapped_data.get('depot_id'),
      plaque=mapped_data.get('plaque') or None,
      color=mapped_data.get('color') or None,
      scale=scale,
      formation=safe_int(mapped_data.get('formation')),
      trainset_number=trainset_number or None,
      decoder_number=mapped_data.get('decoder_number') or None,
      head_light=parse_boolean(mapped_data.get('head_light')) or False,
      interior_light=mapped_data.get('interior_light') or None,
      chip_interface_id=mapped_data.get('chip_interface_id'),
      chip_model_id=mapped_data.get('chip_model_id'),
      price=price or None,
      total_price=calculate_price(price) if price else 0,
      item_number=mapped_data.get('item_number') or None,
      purchase_date=parse_purchase_date(mapped_data.get('purchase_date')),
      merchant_id=mapped_data.get('merchant_id')
    )
    db.session.add(trainset)
    count += 1

  db.session.commit()
  return count


def execute_locomotive_head_import(rows, source_to_target, field_configs, conflict_mode):
  """
  执行先头车模型导入（无唯一约束，直接导入）

  Args:
    rows: 数据行列表
    source_to_target: 源列到目标字段的映射
    field_configs: 字段配置字典
    conflict_mode: 冲突处理模式（先头车无唯一约束，此参数被忽略）

  Returns:
    int: 导入的记录数
  """
  count = 0
  for row in rows:
    # 映射行数据并解析外键
    mapped_data = {}
    for source_col, target_field in source_to_target.items():
      value = row.get(source_col)
      field_config = field_configs.get(target_field, {})
      ref_table = field_config.get('ref')

      if ref_table:
        resolved_id = resolve_foreign_key(ref_table, value)
        if field_config.get('required') and resolved_id is None and value:
          logger.warning(f"跳过先头车行：找不到 {ref_table} '{value}'")
          mapped_data = None
          break
        mapped_data[target_field] = resolved_id
      else:
        mapped_data[target_field] = value

    if not mapped_data:
      continue

    # 检查必填字段
    brand_id = mapped_data.get('brand_id')
    scale = mapped_data.get('scale')
    if not brand_id or not scale:
      logger.warning(f"跳过先头车行：缺少必填字段 brand_id={brand_id}, scale={scale}")
      continue

    # 创建新记录（先头车无唯一约束）
    price = mapped_data.get('price')
    locomotive_head = LocomotiveHead(
      model_id=mapped_data.get('model_id'),
      brand_id=brand_id,
      special_color=mapped_data.get('special_color') or None,
      scale=scale,
      head_light=parse_boolean(mapped_data.get('head_light')) or False,
      interior_light=mapped_data.get('interior_light') or None,
      price=price or None,
      total_price=calculate_price(price) if price else 0,
      item_number=mapped_data.get('item_number') or None,
      purchase_date=parse_purchase_date(mapped_data.get('purchase_date')),
      merchant_id=mapped_data.get('merchant_id')
    )
    db.session.add(locomotive_head)
    count += 1

  db.session.commit()
  return count


def get_cell_value_with_merge(sheet, row, col):
  """
  获取单元格值，支持合并单元格

  当单元格是合并单元格的一部分时，如果当前单元格值为 None，
  则尝试从合并范围的左上角单元格获取值

  Args:
    sheet: openpyxl worksheet 对象
    row: 行号（1-based）
    col: 列号（1-based）

  Returns:
    单元格值
  """
  cell = sheet.cell(row=row, column=col)
  if cell.value is not None:
    return cell.value

  # 检查是否在合并范围内
  for merged_range in sheet.merged_cells.ranges:
    if (merged_range.min_row <= row <= merged_range.max_row and
        merged_range.min_col <= col <= merged_range.max_col):
      # 从合并范围的左上角获取值
      return sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value

  return None


def detect_merged_cell_sets(sheet, headers, set_field_indices):
  """
  检测 Excel 工作表中的合并单元格，用于识别套装边界

  Args:
    sheet: openpyxl worksheet 对象
    headers: 列标题列表
    set_field_indices: 套装字段在 headers 中的索引列表

  Returns:
    list: 套装分组信息列表，每项包含 {'start_row': int, 'end_row': int}
          返回 None 表示无法检测合并单元格，应使用默认行为
  """
  if not sheet.merged_cells.ranges:
    return None

  # 构建 (列索引 -> 合并范围列表) 的映射
  # openpyxl 的行列索引从 1 开始
  column_merge_map = {}
  for merged_range in sheet.merged_cells.ranges:
    start_col = merged_range.min_col
    end_col = merged_range.max_col
    start_row = merged_range.min_row
    end_row = merged_range.max_row

    # 只处理单列的合并单元格
    if start_col == end_col:
      if start_col not in column_merge_map:
        column_merge_map[start_col] = []
      column_merge_map[start_col].append({
        'start_row': start_row,
        'end_row': end_row
      })

  # 检查套装字段列是否有合并单元格
  set_field_cols = [idx + 1 for idx in set_field_indices]  # 转换为 1-based 索引
  has_set_field_merges = any(col in column_merge_map for col in set_field_cols)

  if not has_set_field_merges:
    return None

  # 分析合并单元格的边界来识别套装
  # 关键假设：合并单元格的行边界决定套装边界
  # 使用第一个有合并单元格的套装字段列来识别套装

  primary_col = None
  for col in set_field_cols:
    if col in column_merge_map:
      primary_col = col
      break

  if not primary_col:
    return None

  # 收集所有边界行
  boundary_rows = set()
  boundary_rows.add(2)  # 数据从第 2 行开始

  for merge_info in column_merge_map[primary_col]:
    boundary_rows.add(merge_info['start_row'])
    boundary_rows.add(merge_info['end_row'] + 1)  # 下一个套装开始行

  # 获取总数据行数
  max_data_row = 0
  for row in sheet.iter_rows(min_row=2, values_only=True):
    if any(cell is not None for cell in row):
      max_data_row += 1
    else:
      break

  boundary_rows.add(max_data_row + 2)  # 结束边界

  # 排序并构建套装分组
  sorted_boundaries = sorted(boundary_rows)
  set_groups = []

  for i in range(len(sorted_boundaries) - 1):
    start_row = sorted_boundaries[i]
    end_row = sorted_boundaries[i + 1] - 1

    if start_row <= end_row and start_row >= 2:
      set_groups.append({
        'start_row': start_row,
        'end_row': end_row
      })

  return set_groups if set_groups else None


def validate_merged_cells_consistency(sheet, headers, set_field_indices, set_groups):
  """
  验证合并单元格的一致性

  Args:
    sheet: openpyxl worksheet 对象
    headers: 列标题列表
    set_field_indices: 套装字段索引列表
    set_groups: 套装分组列表

  Returns:
    list: 警告消息列表
  """
  warnings = []

  if not sheet.merged_cells.ranges or not set_groups:
    return warnings

  # 检查每个套装分组内套装字段的一致性
  for group in set_groups:
    start_row = group['start_row']
    end_row = group['end_row']

    for field_idx in set_field_indices:
      col = field_idx + 1  # 1-based
      values = []

      for row_idx in range(start_row, end_row + 1):
        cell = sheet.cell(row=row_idx, column=col)
        value = cell.value
        if value is not None:
          values.append(str(value).strip())

      # 检查是否所有值都相同
      if values and len(set(values)) > 1:
        field_name = headers[field_idx] if field_idx < len(headers) else f"列{col}"
        warnings.append(
          f"行 {start_row}-{end_row}，{field_name} 列值不一致: {', '.join(set(values))}"
        )

  return warnings


def execute_carriage_import(rows, source_to_target, field_configs, conflict_mode,
                            sheet=None, headers=None, set_detection_mode='merged'):
  """
  执行车厢套装导入（套装无唯一约束，按套装分组处理）

  Args:
    rows: 数据行列表
    source_to_target: 源列到目标字段的映射
    field_configs: 字段配置字典
    conflict_mode: 冲突处理模式（车厢无唯一约束，此参数被忽略）
    sheet: openpyxl worksheet 对象（用于合并单元格检测）
    headers: 列标题列表（用于合并单元格检测）
    set_detection_mode: 套装检测模式
      - 'merged': 按合并单元格识别套装（默认）
      - 'row': 每行作为一个独立套装

  Returns:
    dict: {'count': int, 'warnings': list}
  """
  from models import CarriageItem

  count = 0
  warnings = []
  current_set_data = None
  current_items = []

  def save_carriage_set():
    """保存当前套装及其车厢项"""
    nonlocal count
    if not current_set_data:
      return

    brand_id = current_set_data.get('brand_id')
    scale = current_set_data.get('scale')
    if not brand_id or not scale:
      logger.warning(f"跳过车厢套装：缺少必填字段 brand_id={brand_id}, scale={scale}")
      return

    carriage_set = CarriageSet(
      brand_id=brand_id,
      series_id=current_set_data.get('series_id'),
      depot_id=current_set_data.get('depot_id'),
      train_number=current_set_data.get('train_number'),
      plaque=current_set_data.get('plaque'),
      item_number=current_set_data.get('item_number'),
      scale=scale,
      total_price=current_set_data.get('total_price', 0),
      purchase_date=current_set_data.get('purchase_date'),
      merchant_id=current_set_data.get('merchant_id')
    )
    db.session.add(carriage_set)
    db.session.flush()

    for item_data in current_items:
      if item_data.get('model_id'):
        item = CarriageItem(
          set_id=carriage_set.id,
          model_id=item_data.get('model_id'),
          car_number=item_data.get('car_number'),
          color=item_data.get('color'),
          lighting=item_data.get('lighting')
        )
        db.session.add(item)

    count += 1

  # 区分套装字段和车厢项字段
  set_fields = set()
  item_fields = set()
  for field_name, config in field_configs.items():
    if config.get('is_set_field'):
      set_fields.add(field_name)
    elif config.get('is_item_field'):
      item_fields.add(field_name)

  # 获取套装字段在 headers 中的索引
  set_field_indices = []
  if headers and set_detection_mode == 'merged':
    for source_col, target_field in source_to_target.items():
      if target_field in set_fields:
        try:
          idx = headers.index(source_col)
          set_field_indices.append(idx)
        except ValueError:
          pass

  # 尝试检测合并单元格
  set_groups = None
  if sheet and headers and set_detection_mode == 'merged' and set_field_indices:
    set_groups = detect_merged_cell_sets(sheet, headers, set_field_indices)
    if set_groups:
      # 验证合并单元格一致性
      consistency_warnings = validate_merged_cells_consistency(
        sheet, headers, set_field_indices, set_groups
      )
      warnings.extend(consistency_warnings)

  # 如果检测到合并单元格，按分组处理
  if set_groups:
    logger.info(f"检测到 {len(set_groups)} 个合并单元格套装分组")

    for group in set_groups:
      start_row = group['start_row'] - 2  # 转换为 0-based 行索引（跳过标题行）
      end_row = group['end_row'] - 2

      if start_row < 0 or start_row >= len(rows):
        continue

      # 收集该套装的所有行
      group_rows = rows[start_row:end_row + 1]
      if not group_rows:
        continue

      # 从第一行获取套装字段值
      first_row = group_rows[0]
      mapped_first = {}
      for source_col, target_field in source_to_target.items():
        value = first_row.get(source_col)
        field_config = field_configs.get(target_field, {})
        ref_table = field_config.get('ref')

        if ref_table:
          resolved_id = resolve_foreign_key(ref_table, value)
          mapped_first[target_field] = resolved_id
        else:
          mapped_first[target_field] = value

      brand_id = mapped_first.get('brand_id')
      scale = mapped_first.get('scale')

      if not brand_id or not scale:
        logger.warning(f"跳过套装分组（行 {start_row + 2}-{end_row + 2}）：缺少必填字段")
        continue

      # 创建套装数据
      current_set_data = {
        'brand_id': brand_id,
        'series_id': mapped_first.get('series_id'),
        'depot_id': mapped_first.get('depot_id'),
        'train_number': mapped_first.get('train_number') or None,
        'plaque': mapped_first.get('plaque') or None,
        'item_number': mapped_first.get('item_number') or None,
        'scale': scale,
        'total_price': safe_float(mapped_first.get('total_price')),
        'purchase_date': parse_purchase_date(mapped_first.get('purchase_date')),
        'merchant_id': mapped_first.get('merchant_id')
      }
      current_items = []

      # 添加所有车厢项
      for row in group_rows:
        mapped_data = {}
        for source_col, target_field in source_to_target.items():
          value = row.get(source_col)
          field_config = field_configs.get(target_field, {})
          ref_table = field_config.get('ref')

          if ref_table:
            resolved_id = resolve_foreign_key(ref_table, value)
            mapped_data[target_field] = resolved_id
          else:
            mapped_data[target_field] = value

        model_id = mapped_data.get('model_id')
        if model_id:
          current_items.append({
            'model_id': model_id,
            'car_number': mapped_data.get('car_number') or None,
            'color': mapped_data.get('color') or None,
            'lighting': mapped_data.get('lighting') or None
          })

      save_carriage_set()

  else:
    # 默认行为：按品牌+比例识别新套装
    # 如果 set_detection_mode == 'row'，则每行都是一个独立套装
    for row_idx, row in enumerate(rows):
      # 映射行数据并解析外键
      mapped_data = {}
      excel_row = row_idx + 2  # Excel 行号（1-based，跳过标题行）

      for source_col, target_field in source_to_target.items():
        value = row.get(source_col)
        field_config = field_configs.get(target_field, {})

        # 对于套装字段，如果值为 None 且有 worksheet，尝试从合并单元格获取值
        if value is None and sheet and headers and target_field in set_fields:
          try:
            col_idx = headers.index(source_col) + 1  # 转换为 1-based 列索引
            value = get_cell_value_with_merge(sheet, excel_row, col_idx)
          except (ValueError, IndexError):
            pass

        ref_table = field_config.get('ref')

        if ref_table:
          resolved_id = resolve_foreign_key(ref_table, value)
          mapped_data[target_field] = resolved_id
        else:
          mapped_data[target_field] = value

      # 判断是否是新套装
      brand_id = mapped_data.get('brand_id')
      scale = mapped_data.get('scale')

      if set_detection_mode == 'row':
        # 每行都是独立套装：只要有 brand+scale 就是新套装
        is_new_set = brand_id and scale
      else:
        # 默认：brand_id 和 scale 同时存在时认为是新套装
        is_new_set = brand_id and scale

      if is_new_set:
        # 保存之前的套装
        save_carriage_set()
        # 开始新套装
        current_set_data = {
          'brand_id': brand_id,
          'series_id': mapped_data.get('series_id'),
          'depot_id': mapped_data.get('depot_id'),
          'train_number': mapped_data.get('train_number') or None,
          'plaque': mapped_data.get('plaque') or None,
          'item_number': mapped_data.get('item_number') or None,
          'scale': scale,
          'total_price': safe_float(mapped_data.get('total_price')),
          'purchase_date': parse_purchase_date(mapped_data.get('purchase_date')),
          'merchant_id': mapped_data.get('merchant_id')
        }
        current_items = []

      # 添加车厢项
      model_id = mapped_data.get('model_id')
      if model_id and current_set_data:
        current_items.append({
          'model_id': model_id,
          'car_number': mapped_data.get('car_number') or None,
          'color': mapped_data.get('color') or None,
          'lighting': mapped_data.get('lighting') or None
        })

      # 在 'row' 模式下，每行结束时保存套装
      if set_detection_mode == 'row' and current_set_data:
        save_carriage_set()
        current_set_data = None
        current_items = []

    # 保存最后一个套装（非 row 模式）
    if set_detection_mode != 'row':
      save_carriage_set()

  db.session.commit()
  return {'count': count, 'warnings': warnings}


def execute_model_series_import(table_name, rows, source_to_target, conflict_mode):
  """
  执行机车/车厢/动车组系列导入

  Args:
    table_name: 表名 (locomotive_series, carriage_series, trainset_series)
    rows: 数据行列表
    source_to_target: 源列到目标字段的映射
    conflict_mode: 冲突处理模式

  Returns:
    int: 导入的记录数
  """
  model_class = MODEL_CLASS_MAP.get(table_name)
  if not model_class:
    return 0

  count = 0
  for row in rows:
    name = None
    for source_col, target_field in source_to_target.items():
      if target_field == 'name':
        name = row.get(source_col)
        break

    if not name:
      continue

    existing = model_class.query.filter_by(name=name).first()
    if existing:
      if conflict_mode == 'skip':
        continue
      # overwrite 模式下无需更新，因为只有 name 字段
      continue

    series = model_class(name=name)
    db.session.add(series)
    count += 1

  db.session.commit()
  return count


def execute_model_model_import(table_name, rows, source_to_target, field_configs, conflict_mode):
  """
  执行机车/车厢/动车组车型导入（车型没有唯一名称约束，直接导入）

  Args:
    table_name: 表名 (locomotive_model, carriage_model, trainset_model)
    rows: 数据行列表
    source_to_target: 源列到目标字段的映射
    field_configs: 字段配置字典
    conflict_mode: 冲突处理模式（车型无唯一约束，此参数被忽略）

  Returns:
    int: 导入的记录数
  """
  model_class = MODEL_CLASS_MAP.get(table_name)
  if not model_class:
    return 0

  count = 0
  for row in rows:
    mapped_data = {}
    for source_col, target_field in source_to_target.items():
      value = row.get(source_col)
      field_config = field_configs.get(target_field, {})
      ref_table = field_config.get('ref')

      if ref_table:
        resolved_id = resolve_foreign_key(ref_table, value)
        mapped_data[target_field] = resolved_id
      else:
        mapped_data[target_field] = value

    name = mapped_data.get('name')
    if not name:
      continue

    model_instance = model_class(**mapped_data)
    db.session.add(model_instance)
    count += 1

  db.session.commit()
  return count


@api_bp.route('/api/custom-import/execute', methods=['POST'])
def execute_custom_import():
  """
  执行自定义导入

  Request:
    multipart/form-data with:
      - 'file': Excel file
      - 'config': JSON string containing mapping configuration

  Config structure:
    {
      'sheet_mappings': [
        {'sheet_name': '品牌列表', 'table_name': 'brand'},
        ...
      ],
      'column_mappings': {
        'brand': {
          'columns': [
            {'source': '品牌名称', 'target': 'name', 'required': True},
            {'source': '官网', 'target': 'search_url', 'required': False}
          ],
          'conflict_mode': 'overwrite'
        },
        ...
      }
    }

  Returns:
    JSON: {
      'success': True,
      'summary': {'brand': 10, 'locomotive': 5, ...},
      'errors': [],
      'message': '导入完成'
    }
  """
  try:
    # 验证文件
    if 'file' not in request.files:
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
      return jsonify({'success': False, 'error': '文件格式错误，请上传Excel文件'}), 400

    # 验证配置
    config_str = request.form.get('config')
    if not config_str:
      return jsonify({'success': False, 'error': '缺少映射配置'}), 400

    try:
      config = json.loads(config_str)
    except json.JSONDecodeError:
      return jsonify({'success': False, 'error': '配置格式错误，必须是有效的JSON'}), 400

    # 解析 Excel 文件
    workbook = openpyxl.load_workbook(file)

    # 构建工作表数据
    sheets_data = {}
    for sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
      headers = []
      data = []

      for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_idx == 1:
          headers = [str(cell) if cell is not None else '' for cell in row]
        else:
          if any(cell is not None for cell in row):
            row_dict = {}
            for i, cell in enumerate(row):
              if i < len(headers) and headers[i]:
                row_dict[headers[i]] = cell
            data.append(row_dict)

      sheets_data[sheet_name] = {'headers': headers, 'data': data, 'sheet': sheet}

    # 执行导入
    summary = {}
    errors = []
    all_warnings = []

    sheet_mappings = config.get('sheet_mappings', [])
    column_mappings = config.get('column_mappings', {})

    for mapping in sheet_mappings:
      sheet_name = mapping.get('sheet_name')
      table_name = mapping.get('table_name')

      if not sheet_name or not table_name:
        continue

      # 获取表配置
      table_config = SYSTEM_TABLES.get(table_name)
      if not table_config:
        logger.warning(f"Unknown table name: {table_name}")
        continue

      # 获取列映射
      column_mapping = column_mappings.get(table_name, {})
      columns = column_mapping.get('columns', [])
      conflict_mode = column_mapping.get('conflict_mode', 'skip')

      # 构建源列名到目标字段名的映射
      source_to_target = {}
      for col in columns:
        source = col.get('source')
        target = col.get('target')
        if source and target:
          source_to_target[source] = target

      # 构建字段配置字典
      field_configs = {f['name']: f for f in table_config.get('fields', [])}

      # 获取工作表数据
      sheet_data = sheets_data.get(sheet_name, {})
      rows = sheet_data.get('data', [])

      if not rows:
        summary[table_name] = 0
        continue

      try:
        count = 0

        # 系统信息表
        if table_name in ['brand', 'depot', 'merchant', 'power_type', 'chip_interface', 'chip_model']:
          count = execute_system_table_import(table_name, table_config, rows, source_to_target, conflict_mode)

        # 系列表
        elif table_name in ['locomotive_series', 'carriage_series', 'trainset_series']:
          count = execute_model_series_import(table_name, rows, source_to_target, conflict_mode)

        # 车型表
        elif table_name in ['locomotive_model', 'carriage_model', 'trainset_model']:
          count = execute_model_model_import(table_name, rows, source_to_target, field_configs, conflict_mode)

        # 模型数据表
        elif table_name == 'locomotive':
          count = execute_locomotive_import(rows, source_to_target, field_configs, conflict_mode)
        elif table_name == 'trainset':
          count = execute_trainset_import(rows, source_to_target, field_configs, conflict_mode)
        elif table_name == 'locomotive_head':
          count = execute_locomotive_head_import(rows, source_to_target, field_configs, conflict_mode)
        elif table_name == 'carriage':
          # 获取套装检测模式（默认使用合并单元格检测）
          set_detection_mode = column_mapping.get('set_detection_mode', 'merged')
          result = execute_carriage_import(
            rows, source_to_target, field_configs, conflict_mode,
            sheet=sheet_data.get('sheet'),
            headers=sheet_data.get('headers', []),
            set_detection_mode=set_detection_mode
          )
          count = result['count']
          if result.get('warnings'):
            all_warnings.extend([f"车厢: {w}" for w in result['warnings']])

        summary[table_name] = count

      except Exception as e:
        db.session.rollback()
        errors.append(f"{table_config.get('display_name', table_name)}: {str(e)}")
        logger.error(f"Error importing table {table_name}: {str(e)}", exc_info=True)

    if errors:
      return jsonify({
        'success': False,
        'error': '部分导入失败: ' + '; '.join(errors),
        'summary': summary
      }), 400

    logger.info(f"Custom import completed: {summary}")
    response = {
      'success': True,
      'summary': summary,
      'errors': [],
      'message': '导入完成'
    }
    if all_warnings:
      response['warnings'] = all_warnings
    return jsonify(response)

  except Exception as e:
    db.session.rollback()
    logger.error(f"Execute custom import failed: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': '服务器内部错误'}), 500
