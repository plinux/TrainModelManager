from flask import Flask, render_template, request, jsonify, redirect, url_for
from config import Config
from models import db, Locomotive, CarriageSet, Trainset, LocomotiveHead
from models import LocomotiveModel, LocomotiveSeries, PowerType, Brand, Depot, ChipInterface, ChipModel, Merchant
from models import CarriageModel, CarriageSeries, TrainsetModel, TrainsetSeries
from models import CarriageItem
from datetime import date, datetime
import re
import ast
import operator
import logging
import subprocess
import openpyxl
from werkzeug.utils import secure_filename

# 配置日志
logging.basicConfig(
  level=logging.INFO,
  format='%(asctime)s %(name)s %(levelname)s %(message)s',
  handlers=[
    logging.FileHandler('app.log'),
    logging.StreamHandler()
  ]
)
logger = logging.getLogger(__name__)

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)

@app.route('/')
def index():
  """汇总统计页面"""
  return render_template('index.html')

@app.route('/options')
def options():
  """选项维护页面"""
  power_types = PowerType.query.all()
  brands = Brand.query.all()
  merchants = Merchant.query.all()
  depots = Depot.query.all()
  chip_interfaces = ChipInterface.query.all()
  chip_models = ChipModel.query.all()
  locomotive_series = LocomotiveSeries.query.all()
  locomotive_models = LocomotiveModel.query.all()
  carriage_series = CarriageSeries.query.all()
  carriage_models = CarriageModel.query.all()
  trainset_series = TrainsetSeries.query.all()
  trainset_models = TrainsetModel.query.all()
  return render_template('options.html',
    power_types=power_types,
    brands=brands,
    merchants=merchants,
    depots=depots,
    chip_interfaces=chip_interfaces,
    chip_models=chip_models,
    locomotive_series=locomotive_series,
    locomotive_models=locomotive_models,
    carriage_series=carriage_series,
    carriage_models=carriage_models,
    trainset_series=trainset_series,
    trainset_models=trainset_models
  )

@app.route('/api/statistics')
def statistics():
  """获取汇总统计数据"""
  locomotives = Locomotive.query.all()
  carriage_sets = CarriageSet.query.all()
  trainsets = Trainset.query.all()
  locomotive_heads = LocomotiveHead.query.all()

  # 按比例分组
  locomotives_by_scale = {l.scale: len([x for x in locomotives if x.scale == l.scale]) for l in locomotives}
  carriage_sets_by_scale = {c.scale: len([x for x in carriage_sets if x.scale == c.scale]) for c in carriage_sets}
  trainsets_by_scale = {t.scale: len([x for x in trainsets if x.scale == t.scale]) for t in trainsets}
  locomotive_heads_by_scale = {l.scale: len([x for x in locomotive_heads if x.scale == l.scale]) for l in locomotive_heads}

  # 按品牌分组
  locomotives_by_brand = {}
  for l in locomotives:
    brand = l.brand.name if l.brand else '未知'
    locomotives_by_brand[brand] = locomotives_by_brand.get(brand, 0) + 1

  carriage_sets_by_brand = {}
  for c in carriage_sets:
    brand = c.brand.name if c.brand else '未知'
    carriage_sets_by_brand[brand] = carriage_sets_by_brand.get(brand, 0) + 1

  trainsets_by_brand = {}
  for t in trainsets:
    brand = t.brand.name if t.brand else '未知'
    trainsets_by_brand[brand] = trainsets_by_brand.get(brand, 0) + 1

  locomotive_heads_by_brand = {}
  for h in locomotive_heads:
    brand = h.brand.name if h.brand else '未知'
    locomotive_heads_by_brand[brand] = locomotive_heads_by_brand.get(brand, 0) + 1

  return jsonify({
    'locomotive': {
      'count': len(locomotives),
      'total': sum(l.total_price or 0 for l in locomotives),
      'by_scale': locomotives_by_scale,
      'by_brand': locomotives_by_brand
    },
    'carriage': {
      'count': len(carriage_sets),
      'total': sum(c.total_price or 0 for c in carriage_sets),
      'by_scale': carriage_sets_by_scale,
      'by_brand': carriage_sets_by_brand
    },
    'trainset': {
      'count': len(trainsets),
      'total': sum(t.total_price or 0 for t in trainsets),
      'by_scale': trainsets_by_scale,
      'by_brand': trainsets_by_brand
    },
    'locomotive_head': {
      'count': len(locomotive_heads),
      'total': sum(l.total_price or 0 for l in locomotive_heads),
      'by_scale': locomotive_heads_by_scale,
      'by_brand': locomotive_heads_by_brand
    }
  })

@app.route('/api/locomotive/add', methods=['POST'])
def api_add_locomotive():
  """AJAX 添加机车模型"""
  try:
    data = request.get_json()
    scale = data.get('scale')
    locomotive_number = data.get('locomotive_number')
    decoder_number = data.get('decoder_number')

    errors = []

    # 格式验证
    if locomotive_number and not validate_locomotive_number(locomotive_number):
      errors.append({'field': 'locomotive_number', 'message': '机车号格式错误：应为4-12位数字，允许前导0'})
    if decoder_number and not validate_decoder_number(decoder_number):
      errors.append({'field': 'decoder_number', 'message': '编号格式错误：应为1-4位数字，无前导0'})

    # 唯一性验证
    if locomotive_number and check_duplicate('locomotive', 'locomotive_number', locomotive_number, scale):
      errors.append({'field': 'locomotive_number', 'message': f'机车号 {locomotive_number} 在 {scale} 比例下已存在'})
    if decoder_number and check_duplicate('locomotive', 'decoder_number', decoder_number, scale):
      errors.append({'field': 'decoder_number', 'message': f'编号 {decoder_number} 在 {scale} 比例下已存在'})

    if errors:
      return jsonify({'success': False, 'errors': errors})

    purchase_date = data.get('purchase_date')
    purchase_date = purchase_date if purchase_date and purchase_date.strip() else date.today()

    # 调试价格数据
    price_value = data.get('price')
    logger.info(f"Received price field: '{price_value}', type: {type(price_value)}")

    locomotive = Locomotive(
      model_id=int(data.get('model_id')),
      series_id=data.get('series_id'),
      power_type_id=data.get('power_type_id'),
      brand_id=int(data.get('brand_id')),
      depot_id=data.get('depot_id'),
      plaque=data.get('plaque'),
      color=data.get('color'),
      scale=scale,
      locomotive_number=locomotive_number,
      decoder_number=decoder_number,
      chip_interface_id=data.get('chip_interface_id'),
      chip_model_id=data.get('chip_model_id'),
      price=data.get('price'),
      total_price=calculate_price(data.get('price')),
      item_number=data.get('item_number'),
      purchase_date=purchase_date,
      merchant_id=data.get('merchant_id')
    )
    db.session.add(locomotive)
    db.session.commit()
    logger.info(f"Locomotive added: ID={locomotive.id}")

    return jsonify({'success': True, 'message': '机车模型添加成功'})
  except Exception as e:
    logger.error(f"Error adding locomotive: {e}")
    return jsonify({'success': False, 'error': str(e)})

@app.route('/api/carriage/add', methods=['POST'])
def api_add_carriage():
  """AJAX 添加车厢模型"""
  try:
    data = request.get_json()
    errors = []

    # 验证车辆号格式
    for i in range(10):
      model_key = f'model_{i}'
      car_number_key = f'car_number_{i}'

      if data.get(model_key):
        car_number = data.get(car_number_key)
        if car_number and not validate_car_number(car_number):
          errors.append({'field': car_number_key, 'message': f'车辆号 {car_number} 格式错误：应为3-10位数字，无前导0'})

    if errors:
      return jsonify({'success': False, 'errors': errors})

    purchase_date = data.get('purchase_date')
    purchase_date = purchase_date if purchase_date and purchase_date.strip() else date.today()

    carriage_set = CarriageSet(
      brand_id=int(data.get('brand_id')),
      series_id=data.get('series_id'),
      scale=data.get('scale'),
      plaque=data.get('plaque'),
      total_price=calculate_price(data.get('total_price')),
      item_number=data.get('item_number'),
      purchase_date=purchase_date,
      merchant_id=data.get('merchant_id')
    )
    db.session.add(carriage_set)
    db.session.flush()

    # 添加车厢项
    for i in range(10):
      model_key = f'model_{i}'
      if data.get(model_key):
        carriage_item = CarriageItem(
          set_id=carriage_set.id,
          model_id=int(data.get(model_key)),
          car_number=data.get(f'car_number_{i}'),
          color=data.get(f'color_{i}'),
          lighting=data.get(f'lighting_{i}')
        )
        db.session.add(carriage_item)

    db.session.commit()
    logger.info(f"Carriage set added: ID={carriage_set.id}")

    return jsonify({'success': True, 'message': '车厢套装添加成功'})
  except Exception as e:
    logger.error(f"Error adding carriage: {e}")
    return jsonify({'success': False, 'error': str(e)})

@app.route('/api/trainset/add', methods=['POST'])
def api_add_trainset():
  """AJAX 添加动车组模型"""
  try:
    data = request.get_json()
    scale = data.get('scale')
    trainset_number = data.get('trainset_number')
    decoder_number = data.get('decoder_number')

    errors = []

    # 格式验证
    if trainset_number and not validate_trainset_number(trainset_number):
      errors.append({'field': 'trainset_number', 'message': '动车号格式错误：应为3-12位数字，允许前导0'})
    if decoder_number and not validate_decoder_number(decoder_number):
      errors.append({'field': 'decoder_number', 'message': '编号格式错误：应为1-4位数字，无前导0'})

    # 唯一性验证
    if trainset_number and check_duplicate('trainset', 'trainset_number', trainset_number, scale):
      errors.append({'field': 'trainset_number', 'message': f'动车号 {trainset_number} 在 {scale} 比例下已存在'})
    if decoder_number and check_duplicate('trainset', 'decoder_number', decoder_number, scale):
      errors.append({'field': 'decoder_number', 'message': f'编号 {decoder_number} 在 {scale} 比例下已存在'})

    if errors:
      return jsonify({'success': False, 'errors': errors})

    purchase_date = data.get('purchase_date')
    purchase_date = purchase_date if purchase_date and purchase_date.strip() else date.today()

    trainset = Trainset(
      model_id=int(data.get('model_id')),
      series_id=data.get('series_id'),
      power_type_id=data.get('power_type_id'),
      brand_id=int(data.get('brand_id')),
      depot_id=data.get('depot_id'),
      plaque=data.get('plaque'),
      color=data.get('color'),
      scale=scale,
      trainset_number=trainset_number,
      decoder_number=decoder_number,
      head_light=True if data.get('head_light') == 'true' else False,
      interior_light=data.get('interior_light'),
      chip_interface_id=data.get('chip_interface_id'),
      chip_model_id=data.get('chip_model_id'),
      price=data.get('price'),
      total_price=calculate_price(data.get('price')),
      item_number=data.get('item_number'),
      purchase_date=purchase_date,
      merchant_id=data.get('merchant_id')
    )
    db.session.add(trainset)
    db.session.commit()
    logger.info(f"Trainset added: ID={trainset.id}")

    return jsonify({'success': True, 'message': '动车组模型添加成功'})
  except Exception as e:
    logger.error(f"Error adding trainset: {e}")
    return jsonify({'success': False, 'error': str(e)})

@app.route('/api/locomotive-head/add', methods=['POST'])
def api_add_locomotive_head():
  """AJAX 添加先头车模型"""
  try:
    data = request.get_json()

    purchase_date = data.get('purchase_date')
    purchase_date = purchase_date if purchase_date and purchase_date.strip() else date.today()

    locomotive_head = LocomotiveHead(
      model_id=int(data.get('model_id')),
      brand_id=int(data.get('brand_id')),
      depot_id=data.get('depot_id'),
      special_color=data.get('special_color'),
      scale=data.get('scale'),
      head_light=True if data.get('head_light') == 'true' else False,
      interior_light=data.get('interior_light'),
      price=data.get('price'),
      total_price=calculate_price(data.get('price')),
      item_number=data.get('item_number'),
      purchase_date=purchase_date,
      merchant_id=data.get('merchant_id')
    )
    db.session.add(locomotive_head)
    db.session.commit()
    logger.info(f"Locomotive head added: ID={locomotive_head.id}")

    return jsonify({'success': True, 'message': '先头车模型添加成功'})
  except Exception as e:
    logger.error(f"Error adding locomotive head: {e}")
    return jsonify({'success': False, 'error': str(e)})

def validate_locomotive_number(number):
  """验证机车号格式：4-12位数字，允许前导0"""
  return bool(re.match(r'^\d{4,12}$', number))

def validate_decoder_number(number):
  """验证编号格式：1-4位数字，无前导0"""
  return bool(re.match(r'^[1-9]\d{0,3}$', number))

def validate_trainset_number(number):
  """验证动车号格式：3-12位数字，允许前导0"""
  return bool(re.match(r'^\d{3,12}$', number))

def validate_car_number(number):
  """验证车辆号格式：3-10位数字，无前导0"""
  return bool(re.match(r'^[1-9]\d{2,9}$', number))

class SafeEval(ast.NodeVisitor):
  """安全的表达式求值器，只允许数字和基本运算"""

  def __init__(self):
    self._operators = {
      ast.Add: operator.add,
      ast.Sub: operator.sub,
      ast.Mult: operator.mul,
      ast.Div: operator.truediv,
      ast.USub: operator.neg
    }

  def visit(self, node):
    if not self._allowed_nodes(node):
      raise ValueError(f"不安全的表达式节点: {type(node).__name__}")
    return super().visit(node)

  def _allowed_nodes(self, node):
    return isinstance(node, (
      ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant
    ))

  def generic_visit(self, node):
    if isinstance(node, ast.BinOp):
      left = self.visit(node.left)
      right = self.visit(node.right)
      return self._operators[type(node.op)](left, right)
    elif isinstance(node, ast.UnaryOp):
      operand = self.visit(node.operand)
      return self._operators[type(node.op)](operand)
    elif isinstance(node, ast.Constant):
      if isinstance(node.value, (int, float)):
        return node.value
      raise ValueError(f"不支持的常量类型: {type(node.value).__name__}")
    return super().generic_visit(node)

def calculate_price(price_expr):
  """安全计算价格表达式"""
  logger.info(f"calculate_price called with: '{price_expr}', type: {type(price_expr)}")
  if not price_expr or not str(price_expr).strip():
    logger.info("Price expression is empty, returning 0")
    return 0
  price_str = str(price_expr)
  # 只允许数字、+、-、*、/、()、小数点
  if not re.match(r'^[\d+\-*/().\s]+$', price_str):
    logger.info(f"Price expression '{price_str}' does not match pattern")
    return 0
  try:
    # 使用 AST 解析表达式
    expr = ast.parse(price_str, mode='eval')
    evaluator = SafeEval()
    result = evaluator.visit(expr.body)
    logger.info(f"Calculated price: {result}, type: {type(result)}")
    # 确保结果是数字
    if isinstance(result, (int, float)):
      return result
    return 0
  except (ValueError, SyntaxError, TypeError, ZeroDivisionError) as e:
    logger.info(f"Error calculating price: {e}")
    return 0

def check_duplicate(table, field, value, scale=None):
  """检查同一比例内的唯一性"""
  if table == 'locomotive':
    if field == 'locomotive_number':
      return Locomotive.query.filter_by(locomotive_number=value, scale=scale).first()
    elif field == 'decoder_number':
      return Locomotive.query.filter_by(decoder_number=value, scale=scale).first()
  elif table == 'trainset':
    if field == 'trainset_number':
      return Trainset.query.filter_by(trainset_number=value, scale=scale).first()
    elif field == 'decoder_number':
      return Trainset.query.filter_by(decoder_number=value, scale=scale).first()
  return None

@app.route('/locomotive', methods=['GET', 'POST'])
def locomotive():
  """机车模型列表和添加"""
  locomotives = Locomotive.query.all()
  locomotive_models = LocomotiveModel.query.all()
  locomotive_series = LocomotiveSeries.query.all()
  power_types = PowerType.query.all()
  brands = Brand.query.all()
  depots = Depot.query.all()
  chip_interfaces = ChipInterface.query.all()
  chip_models = ChipModel.query.all()
  merchants = Merchant.query.all()

  errors = []

  if request.method == 'POST':
    scale = request.form.get('scale')
    locomotive_number = request.form.get('locomotive_number')
    decoder_number = request.form.get('decoder_number')

    # 格式验证
    if locomotive_number and not validate_locomotive_number(locomotive_number):
      errors.append("机车号格式错误：应为4-12位数字，允许前导0")
    if decoder_number and not validate_decoder_number(decoder_number):
      errors.append("编号格式错误：应为1-4位数字，无前导0")

    # 唯一性验证
    if locomotive_number and check_duplicate('locomotive', 'locomotive_number', locomotive_number, scale):
      errors.append(f"机车号 {locomotive_number} 在 {scale} 比例下已存在")
    if decoder_number and check_duplicate('locomotive', 'decoder_number', decoder_number, scale):
      errors.append(f"编号 {decoder_number} 在 {scale} 比例下已存在")

    if not errors:
      purchase_date = request.form.get('purchase_date')
      purchase_date = purchase_date if purchase_date.strip() else date.today()

      locomotive = Locomotive(
        model_id=int(request.form.get('model_id')),
        series_id=request.form.get('series_id'),
        power_type_id=request.form.get('power_type_id'),
        brand_id=int(request.form.get('brand_id')),
        depot_id=request.form.get('depot_id'),
        plaque=request.form.get('plaque'),
        color=request.form.get('color'),
        scale=scale,
        locomotive_number=locomotive_number,
        decoder_number=decoder_number,
        chip_interface_id=request.form.get('chip_interface_id'),
        chip_model_id=request.form.get('chip_model_id'),
        price=request.form.get('price'),
        total_price=calculate_price(request.form.get('price')),
        item_number=request.form.get('item_number'),
        purchase_date=purchase_date,
        merchant_id=request.form.get('merchant_id')
      )
      db.session.add(locomotive)
      db.session.commit()
      return redirect(url_for('locomotive'))

  return render_template('locomotive.html',
    locomotives=locomotives,
    locomotive_models=locomotive_models,
    locomotive_series=locomotive_series,
    power_types=power_types,
    brands=brands,
    depots=depots,
    chip_interfaces=chip_interfaces,
    chip_models=chip_models,
    merchants=merchants,
    errors=errors
  )

@app.route('/locomotive/delete/<int:id>', methods=['POST'])
def delete_locomotive(id):
  """删除机车模型"""
  locomotive = Locomotive.query.get_or_404(id)
  logger.info(f"Deleting locomotive: {locomotive}")
  db.session.delete(locomotive)
  db.session.commit()
  logger.info(f"Locomotive deleted: ID={id}")
  return redirect(url_for('locomotive'))

@app.route('/locomotive/edit/<int:id>', methods=['GET', 'POST'])
def edit_locomotive(id):
  """编辑机车模型"""
  locomotive = Locomotive.query.get_or_404(id)
  locomotive_models = LocomotiveModel.query.all()
  locomotive_series = LocomotiveSeries.query.all()
  power_types = PowerType.query.all()
  brands = Brand.query.all()
  depots = Depot.query.all()
  chip_interfaces = ChipInterface.query.all()
  chip_models = ChipModel.query.all()
  merchants = Merchant.query.all()

  errors = []

  if request.method == 'POST':
    scale = request.form.get('scale')
    locomotive_number = request.form.get('locomotive_number')
    decoder_number = request.form.get('decoder_number')

    # 格式验证
    if locomotive_number and not validate_locomotive_number(locomotive_number):
      errors.append("机车号格式错误：应为4-12位数字，允许前导0")
    if decoder_number and not validate_decoder_number(decoder_number):
      errors.append("编号格式错误：应为1-4位数字，无前导0")

    # 唯一性验证（排除当前记录）
    if locomotive_number:
      duplicate = Locomotive.query.filter(
        Locomotive.locomotive_number == locomotive_number,
        Locomotive.scale == scale,
        Locomotive.id != id
      ).first()
      if duplicate:
        errors.append(f"机车号 {locomotive_number} 在 {scale} 比例下已存在")
    if decoder_number:
      duplicate = Locomotive.query.filter(
        Locomotive.decoder_number == decoder_number,
        Locomotive.scale == scale,
        Locomotive.id != id
      ).first()
      if duplicate:
        errors.append(f"编号 {decoder_number} 在 {scale} 比例下已存在")

    if not errors:
      purchase_date = request.form.get('purchase_date')
      if purchase_date and purchase_date.strip():
        purchase_date = datetime.strptime(purchase_date, '%Y-%m-%d').date()
      else:
        purchase_date = date.today()

      locomotive.model_id = int(request.form.get('model_id'))
      locomotive.series_id = request.form.get('series_id')
      locomotive.power_type_id = request.form.get('power_type_id')
      locomotive.brand_id = int(request.form.get('brand_id'))
      locomotive.depot_id = request.form.get('depot_id')
      locomotive.plaque = request.form.get('plaque')
      locomotive.color = request.form.get('color')
      locomotive.scale = scale
      locomotive.locomotive_number = locomotive_number
      locomotive.decoder_number = decoder_number
      locomotive.chip_interface_id = request.form.get('chip_interface_id')
      locomotive.chip_model_id = request.form.get('chip_model_id')
      locomotive.price = request.form.get('price')
      locomotive.total_price = calculate_price(request.form.get('price'))
      locomotive.item_number = request.form.get('item_number')
      locomotive.purchase_date = purchase_date
      locomotive.merchant_id = request.form.get('merchant_id')

      db.session.commit()
      logger.info(f"Locomotive updated: ID={id}")
      return redirect(url_for('locomotive'))

  return render_template('locomotive_edit.html',
    locomotive=locomotive,
    locomotive_models=locomotive_models,
    locomotive_series=locomotive_series,
    power_types=power_types,
    brands=brands,
    depots=depots,
    chip_interfaces=chip_interfaces,
    chip_models=chip_models,
    merchants=merchants,
    errors=errors
  )

@app.route('/carriage', methods=['GET', 'POST'])
def carriage():
  """车厢模型列表和添加"""
  carriage_sets = CarriageSet.query.all()
  carriage_models = CarriageModel.query.all()
  carriage_series = CarriageSeries.query.all()
  brands = Brand.query.all()
  depots = Depot.query.all()
  merchants = Merchant.query.all()

  errors = []

  if request.method == 'POST':
    # 先验证车辆号格式
    for i in range(10):
      model_key = f'model_{i}'
      car_number_key = f'car_number_{i}'

      if model_key in request.form and request.form.get(model_key):
        car_number = request.form.get(car_number_key)
        # 格式验证
        if car_number and not validate_car_number(car_number):
          errors.append(f"车辆号 {car_number} 格式错误：应为3-10位数字，无前导0")

    if not errors:
      purchase_date = request.form.get('purchase_date')
      purchase_date = purchase_date if purchase_date.strip() else date.today()

      # 验证通过后添加数据
      carriage_set = CarriageSet(
        brand_id=int(request.form.get('brand_id')),
        series_id=request.form.get('series_id'),
        depot_id=request.form.get('depot_id'),
        train_number=request.form.get('train_number'),
        plaque=request.form.get('plaque'),
        item_number=request.form.get('item_number'),
        scale=request.form.get('scale'),
        total_price=float(request.form.get('total_price') or 0),
        purchase_date=purchase_date,
        merchant_id=request.form.get('merchant_id')
      )
      db.session.add(carriage_set)
      db.session.commit()
      db.session.refresh(carriage_set)

      # 添加车厢项
      for i in range(10):
        model_key = f'model_{i}'
        car_number_key = f'car_number_{i}'
        color_key = f'color_{i}'
        lighting_key = f'lighting_{i}'

        if model_key in request.form and request.form.get(model_key):
          carriage_item = CarriageItem(
            set_id=carriage_set.id,
            model_id=int(request.form.get(model_key)),
            car_number=request.form.get(car_number_key),
            color=request.form.get(color_key),
            lighting=request.form.get(lighting_key)
          )
          db.session.add(carriage_item)

      db.session.commit()
      return redirect(url_for('carriage'))

  return render_template('carriage.html',
    carriage_sets=carriage_sets,
    carriage_models=carriage_models,
    carriage_series=carriage_series,
    brands=brands,
    depots=depots,
    merchants=merchants,
    errors=errors
  )

@app.route('/carriage/delete/<int:id>', methods=['POST'])
def delete_carriage(id):
  """删除车厢套装"""
  carriage_set = CarriageSet.query.get_or_404(id)
  logger.info(f"Deleting carriage set: {carriage_set}")
  db.session.delete(carriage_set)
  db.session.commit()
  logger.info(f"Carriage set deleted: ID={id}")
  return redirect(url_for('carriage'))

@app.route('/carriage/edit/<int:id>', methods=['GET', 'POST'])
def edit_carriage(id):
  """编辑车厢套装"""
  carriage_set = CarriageSet.query.get_or_404(id)
  carriage_models = CarriageModel.query.all()
  carriage_series = CarriageSeries.query.all()
  brands = Brand.query.all()
  depots = Depot.query.all()
  merchants = Merchant.query.all()

  errors = []

  if request.method == 'POST':
    # 先验证车辆号格式
    for i in range(10):
      model_key = f'model_{i}'
      car_number_key = f'car_number_{i}'

      if model_key in request.form and request.form.get(model_key):
        car_number = request.form.get(car_number_key)
        # 格式验证
        if car_number and not validate_car_number(car_number):
          errors.append(f"车辆号 {car_number} 格式错误：应为3-10位数字，无前导0")

    if not errors:
      purchase_date = request.form.get('purchase_date')
      if purchase_date and purchase_date.strip():
        purchase_date = datetime.strptime(purchase_date, '%Y-%m-%d').date()
      else:
        purchase_date = date.today()

      # 更新车厢套装
      carriage_set.brand_id = int(request.form.get('brand_id'))
      carriage_set.series_id = request.form.get('series_id')
      carriage_set.depot_id = request.form.get('depot_id')
      carriage_set.train_number = request.form.get('train_number')
      carriage_set.plaque = request.form.get('plaque')
      carriage_set.item_number = request.form.get('item_number')
      carriage_set.scale = request.form.get('scale')
      carriage_set.total_price = float(request.form.get('total_price') or 0)
      carriage_set.purchase_date = purchase_date
      carriage_set.merchant_id = request.form.get('merchant_id')

      # 删除旧的车厢项
      CarriageItem.query.filter_by(set_id=id).delete()

      # 添加新的车厢项
      for i in range(10):
        model_key = f'model_{i}'
        car_number_key = f'car_number_{i}'
        color_key = f'color_{i}'
        lighting_key = f'lighting_{i}'

        if model_key in request.form and request.form.get(model_key):
          carriage_item = CarriageItem(
            set_id=carriage_set.id,
            model_id=int(request.form.get(model_key)),
            car_number=request.form.get(car_number_key),
            color=request.form.get(color_key),
            lighting=request.form.get(lighting_key)
          )
          db.session.add(carriage_item)

      db.session.commit()
      logger.info(f"Carriage set updated: ID={id}")
      return redirect(url_for('carriage'))

  return render_template('carriage_edit.html',
    carriage_set=carriage_set,
    carriage_models=carriage_models,
    carriage_series=carriage_series,
    brands=brands,
    depots=depots,
    merchants=merchants,
    errors=errors
  )

@app.route('/trainset', methods=['GET', 'POST'])
def trainset():
  """动车组模型列表和添加"""
  trainsets = Trainset.query.all()
  trainset_models = TrainsetModel.query.all()
  trainset_series = TrainsetSeries.query.all()
  power_types = PowerType.query.all()
  brands = Brand.query.all()
  depots = Depot.query.all()
  chip_interfaces = ChipInterface.query.all()
  chip_models = ChipModel.query.all()
  merchants = Merchant.query.all()

  errors = []

  if request.method == 'POST':
    scale = request.form.get('scale')
    trainset_number = request.form.get('trainset_number')
    decoder_number = request.form.get('decoder_number')

    # 格式验证
    if trainset_number and not validate_trainset_number(trainset_number):
      errors.append("动车号格式错误：应为3-12位数字，允许前导0")
    if decoder_number and not validate_decoder_number(decoder_number):
      errors.append("编号格式错误：应为1-4位数字，无前导0")

    # 唯一性验证
    if trainset_number and check_duplicate('trainset', 'trainset_number', trainset_number, scale):
      errors.append(f"动车号 {trainset_number} 在 {scale} 比例下已存在")
    if decoder_number and check_duplicate('trainset', 'decoder_number', decoder_number, scale):
      errors.append(f"编号 {decoder_number} 在 {scale} 比例下已存在")

    if not errors:
      purchase_date = request.form.get('purchase_date')
      purchase_date = purchase_date if purchase_date.strip() else date.today()

      trainset = Trainset(
        model_id=int(request.form.get('model_id')),
        series_id=request.form.get('series_id'),
        power_type_id=request.form.get('power_type_id'),
        brand_id=int(request.form.get('brand_id')),
        depot_id=request.form.get('depot_id'),
        plaque=request.form.get('plaque'),
        color=request.form.get('color'),
        scale=scale,
        formation=int(request.form.get('formation')) if request.form.get('formation') else None,
        trainset_number=trainset_number,
        decoder_number=decoder_number,
        head_light=True if request.form.get('head_light') == 'true' else False,
        interior_light=request.form.get('interior_light'),
        chip_interface_id=request.form.get('chip_interface_id'),
        chip_model_id=request.form.get('chip_model_id'),
        price=request.form.get('price'),
        total_price=calculate_price(request.form.get('price')),
        item_number=request.form.get('item_number'),
        purchase_date=purchase_date,
        merchant_id=request.form.get('merchant_id')
      )
      db.session.add(trainset)
      db.session.commit()
      return redirect(url_for('trainset'))

  return render_template('trainset.html',
    trainsets=trainsets,
    trainset_models=trainset_models,
    trainset_series=trainset_series,
    power_types=power_types,
    brands=brands,
    depots=depots,
    chip_interfaces=chip_interfaces,
    chip_models=chip_models,
    merchants=merchants,
    errors=errors
  )

@app.route('/trainset/delete/<int:id>', methods=['POST'])
def delete_trainset(id):
  """删除动车组模型"""
  trainset = Trainset.query.get_or_404(id)
  logger.info(f"Deleting trainset: {trainset}")
  db.session.delete(trainset)
  db.session.commit()
  logger.info(f"Trainset deleted: ID={id}")
  return redirect(url_for('trainset'))

@app.route('/trainset/edit/<int:id>', methods=['GET', 'POST'])
def edit_trainset(id):
  """编辑动车组模型"""
  trainset = Trainset.query.get_or_404(id)
  trainset_models = TrainsetModel.query.all()
  trainset_series = TrainsetSeries.query.all()
  power_types = PowerType.query.all()
  brands = Brand.query.all()
  depots = Depot.query.all()
  chip_interfaces = ChipInterface.query.all()
  chip_models = ChipModel.query.all()
  merchants = Merchant.query.all()

  errors = []

  if request.method == 'POST':
    scale = request.form.get('scale')
    trainset_number = request.form.get('trainset_number')
    decoder_number = request.form.get('decoder_number')

    # 格式验证
    if trainset_number and not validate_trainset_number(trainset_number):
      errors.append("动车号格式错误：应为3-12位数字，允许前导0")
    if decoder_number and not validate_decoder_number(decoder_number):
      errors.append("编号格式错误：应为1-4位数字，无前导0")

    # 唯一性验证（排除当前记录）
    if trainset_number:
      duplicate = Trainset.query.filter(
        Trainset.trainset_number == trainset_number,
        Trainset.scale == scale,
        Trainset.id != id
      ).first()
      if duplicate:
        errors.append(f"动车号 {trainset_number} 在 {scale} 比例下已存在")
    if decoder_number:
      duplicate = Trainset.query.filter(
        Trainset.decoder_number == decoder_number,
        Trainset.scale == scale,
        Trainset.id != id
      ).first()
      if duplicate:
        errors.append(f"编号 {decoder_number} 在 {scale} 比例下已存在")

    if not errors:
      purchase_date = request.form.get('purchase_date')
      if purchase_date and purchase_date.strip():
        purchase_date = datetime.strptime(purchase_date, '%Y-%m-%d').date()
      else:
        purchase_date = date.today()

      trainset.model_id = int(request.form.get('model_id'))
      trainset.series_id = request.form.get('series_id')
      trainset.power_type_id = request.form.get('power_type_id')
      trainset.brand_id = int(request.form.get('brand_id'))
      trainset.depot_id = request.form.get('depot_id')
      trainset.plaque = request.form.get('plaque')
      trainset.color = request.form.get('color')
      trainset.scale = scale
      trainset.formation = int(request.form.get('formation')) if request.form.get('formation') else None
      trainset.trainset_number = trainset_number
      trainset.decoder_number = decoder_number
      trainset.head_light = True if request.form.get('head_light') == 'true' else False
      trainset.interior_light = request.form.get('interior_light')
      trainset.chip_interface_id = request.form.get('chip_interface_id')
      trainset.chip_model_id = request.form.get('chip_model_id')
      trainset.price = request.form.get('price')
      trainset.total_price = calculate_price(request.form.get('price'))
      trainset.item_number = request.form.get('item_number')
      trainset.purchase_date = purchase_date
      trainset.merchant_id = request.form.get('merchant_id')

      db.session.commit()
      logger.info(f"Trainset updated: ID={id}")
      return redirect(url_for('trainset'))

  return render_template('trainset_edit.html',
    trainset=trainset,
    trainset_models=trainset_models,
    trainset_series=trainset_series,
    power_types=power_types,
    brands=brands,
    depots=depots,
    chip_interfaces=chip_interfaces,
    chip_models=chip_models,
    merchants=merchants,
    errors=errors
  )

@app.route('/locomotive-head', methods=['GET', 'POST'])
def locomotive_head():
  """先头车模型列表和添加"""
  locomotive_heads = LocomotiveHead.query.all()
  trainset_models = TrainsetModel.query.all()
  brands = Brand.query.all()
  depots = Depot.query.all()
  merchants = Merchant.query.all()

  if request.method == 'POST':
    purchase_date = request.form.get('purchase_date')
    purchase_date = purchase_date if purchase_date.strip() else date.today()

    locomotive_head = LocomotiveHead(
      model_id=int(request.form.get('model_id')),
      brand_id=int(request.form.get('brand_id')),
      depot_id=request.form.get('depot_id'),
      special_color=request.form.get('special_color'),
      scale=request.form.get('scale'),
      head_light=True if request.form.get('head_light') == 'true' else False,
      interior_light=request.form.get('interior_light'),
      price=request.form.get('price'),
      total_price=calculate_price(request.form.get('price')),
      item_number=request.form.get('item_number'),
      purchase_date=purchase_date,
      merchant_id=request.form.get('merchant_id')
    )
    db.session.add(locomotive_head)
    db.session.commit()
    return redirect(url_for('locomotive_head'))

  return render_template('locomotive_head.html',
    locomotive_heads=locomotive_heads,
    trainset_models=trainset_models,
    brands=brands,
    depots=depots,
    merchants=merchants
  )

@app.route('/locomotive-head/delete/<int:id>', methods=['POST'])
def delete_locomotive_head(id):
  """删除先头车模型"""
  locomotive_head = LocomotiveHead.query.get_or_404(id)
  logger.info(f"Deleting locomotive head: {locomotive_head}")
  db.session.delete(locomotive_head)
  db.session.commit()
  logger.info(f"Locomotive head deleted: ID={id}")
  return redirect(url_for('locomotive_head'))

@app.route('/locomotive-head/edit/<int:id>', methods=['GET', 'POST'])
def edit_locomotive_head(id):
  """编辑先头车模型"""
  locomotive_head = LocomotiveHead.query.get_or_404(id)
  trainset_models = TrainsetModel.query.all()
  brands = Brand.query.all()
  depots = Depot.query.all()
  merchants = Merchant.query.all()

  if request.method == 'POST':
    purchase_date = request.form.get('purchase_date')
    if purchase_date and purchase_date.strip():
      purchase_date = datetime.strptime(purchase_date, '%Y-%m-%d').date()
    else:
      purchase_date = date.today()

    locomotive_head.model_id = int(request.form.get('model_id'))
    locomotive_head.brand_id = int(request.form.get('brand_id'))
    locomotive_head.depot_id = request.form.get('depot_id')
    locomotive_head.special_color = request.form.get('special_color')
    locomotive_head.scale = request.form.get('scale')
    locomotive_head.head_light = True if request.form.get('head_light') == 'true' else False
    locomotive_head.interior_light = request.form.get('interior_light')
    locomotive_head.price = request.form.get('price')
    locomotive_head.total_price = calculate_price(request.form.get('price'))
    locomotive_head.item_number = request.form.get('item_number')
    locomotive_head.purchase_date = purchase_date
    locomotive_head.merchant_id = request.form.get('merchant_id')

    db.session.commit()
    logger.info(f"Locomotive head updated: ID={id}")
    return redirect(url_for('locomotive_head'))

  return render_template('locomotive_head_edit.html',
    locomotive_head=locomotive_head,
    trainset_models=trainset_models,
    brands=brands,
    depots=depots,
    merchants=merchants
  )

# 选项维护路由
@app.route('/options/power_type', methods=['POST'])
def add_power_type():
  power_type = PowerType(name=request.form.get('name'))
  logger.info(f"Adding power type: {power_type.name}")
  db.session.add(power_type)
  db.session.commit()
  logger.info(f"Power type added: ID={power_type.id}")
  return redirect(url_for('options'))

@app.route('/options/power_type/delete/<int:id>', methods=['POST'])
def delete_power_type(id):
  power_type = PowerType.query.get_or_404(id)
  db.session.delete(power_type)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/power_type/edit/<int:id>', methods=['GET', 'POST'])
def edit_power_type(id):
  """编辑动力类型"""
  power_type = PowerType.query.get_or_404(id)

  if request.method == 'POST':
    power_type.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=power_type,
    title='编辑动力类型',
    action_url=url_for('edit_power_type', id=id)
  )

@app.route('/options/brand', methods=['POST'])
def add_brand():
  brand = Brand(name=request.form.get('name'))
  db.session.add(brand)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/brand/delete/<int:id>', methods=['POST'])
def delete_brand(id):
  brand = Brand.query.get_or_404(id)
  db.session.delete(brand)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/brand/edit/<int:id>', methods=['GET', 'POST'])
def edit_brand(id):
  """编辑品牌"""
  brand = Brand.query.get_or_404(id)

  if request.method == 'POST':
    brand.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=brand,
    title='编辑品牌',
    action_url=url_for('edit_brand', id=id)
  )

@app.route('/options/merchant', methods=['POST'])
def add_merchant():
  merchant = Merchant(name=request.form.get('name'))
  db.session.add(merchant)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/merchant/delete/<int:id>', methods=['POST'])
def delete_merchant(id):
  merchant = Merchant.query.get_or_404(id)
  db.session.delete(merchant)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/merchant/edit/<int:id>', methods=['GET', 'POST'])
def edit_merchant(id):
  """编辑商家"""
  merchant = Merchant.query.get_or_404(id)

  if request.method == 'POST':
    merchant.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=merchant,
    title='编辑商家',
    action_url=url_for('edit_merchant', id=id)
  )

@app.route('/options/depot', methods=['POST'])
def add_depot():
  depot = Depot(name=request.form.get('name'))
  db.session.add(depot)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/depot/delete/<int:id>', methods=['POST'])
def delete_depot(id):
  depot = Depot.query.get_or_404(id)
  db.session.delete(depot)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/depot/edit/<int:id>', methods=['GET', 'POST'])
def edit_depot(id):
  """编辑车辆段"""
  depot = Depot.query.get_or_404(id)

  if request.method == 'POST':
    depot.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=depot,
    title='编辑车辆段',
    action_url=url_for('edit_depot', id=id)
  )

@app.route('/options/chip_interface', methods=['POST'])
def add_chip_interface():
  chip_interface = ChipInterface(name=request.form.get('name'))
  db.session.add(chip_interface)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/chip_interface/delete/<int:id>', methods=['POST'])
def delete_chip_interface(id):
  chip_interface = ChipInterface.query.get_or_404(id)
  db.session.delete(chip_interface)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/chip_interface/edit/<int:id>', methods=['GET', 'POST'])
def edit_chip_interface(id):
  """编辑芯片接口"""
  chip_interface = ChipInterface.query.get_or_404(id)

  if request.method == 'POST':
    chip_interface.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=chip_interface,
    title='编辑芯片接口',
    action_url=url_for('edit_chip_interface', id=id)
  )

@app.route('/options/chip_model', methods=['POST'])
def add_chip_model():
  chip_model = ChipModel(name=request.form.get('name'))
  db.session.add(chip_model)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/chip_model/delete/<int:id>', methods=['POST'])
def delete_chip_model(id):
  chip_model = ChipModel.query.get_or_404(id)
  db.session.delete(chip_model)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/chip_model/edit/<int:id>', methods=['GET', 'POST'])
def edit_chip_model(id):
  """编辑芯片型号"""
  chip_model = ChipModel.query.get_or_404(id)

  if request.method == 'POST':
    chip_model.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=chip_model,
    title='编辑芯片型号',
    action_url=url_for('edit_chip_model', id=id)
  )

@app.route('/options/locomotive_series', methods=['POST'])
def add_locomotive_series():
  series = LocomotiveSeries(name=request.form.get('name'))
  db.session.add(series)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/locomotive_series/delete/<int:id>', methods=['POST'])
def delete_locomotive_series(id):
  series = LocomotiveSeries.query.get_or_404(id)
  if series.locomotives or series.models:
    return f"该机车系列正在被使用，无法删除！<script>setTimeout(()=>location.href='/options', 2000);</script>"
  db.session.delete(series)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/locomotive_series/edit/<int:id>', methods=['GET', 'POST'])
def edit_locomotive_series(id):
  """编辑机车系列"""
  series = LocomotiveSeries.query.get_or_404(id)

  if request.method == 'POST':
    series.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=series,
    title='编辑机车系列',
    action_url=url_for('edit_locomotive_series', id=id)
  )

@app.route('/options/locomotive_model', methods=['POST'])
def add_locomotive_model():
  model = LocomotiveModel(
    name=request.form.get('name'),
    series_id=int(request.form.get('series_id')),
    power_type_id=int(request.form.get('power_type_id'))
  )
  db.session.add(model)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/locomotive_model/delete/<int:id>', methods=['POST'])
def delete_locomotive_model(id):
  model = LocomotiveModel.query.get_or_404(id)
  if model.locomotives:
    return f"该机车型号正在被使用，无法删除！<script>setTimeout(()=>location.href='/options', 2000);</script>"
  db.session.delete(model)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/locomotive_model/edit/<int:id>', methods=['GET', 'POST'])
def edit_locomotive_model(id):
  """编辑机车型号"""
  model = LocomotiveModel.query.get_or_404(id)
  locomotive_series = LocomotiveSeries.query.all()
  power_types = PowerType.query.all()

  if request.method == 'POST':
    model.name = request.form.get('name')
    model.series_id = int(request.form.get('series_id'))
    model.power_type_id = int(request.form.get('power_type_id'))
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_locomotive_model.html',
    model=model,
    locomotive_series=locomotive_series,
    power_types=power_types
  )

@app.route('/options/carriage_series', methods=['POST'])
def add_carriage_series():
  series = CarriageSeries(name=request.form.get('name'))
  db.session.add(series)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/carriage_series/delete/<int:id>', methods=['POST'])
def delete_carriage_series(id):
  series = CarriageSeries.query.get_or_404(id)
  if series.carriages or series.models:
    return f"该车厢系列正在被使用，无法删除！<script>setTimeout(()=>location.href='/options', 2000);</script>"
  db.session.delete(series)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/carriage_series/edit/<int:id>', methods=['GET', 'POST'])
def edit_carriage_series(id):
  """编辑车厢系列"""
  series = CarriageSeries.query.get_or_404(id)

  if request.method == 'POST':
    series.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=series,
    title='编辑车厢系列',
    action_url=url_for('edit_carriage_series', id=id)
  )

@app.route('/options/carriage_model', methods=['POST'])
def add_carriage_model():
  model = CarriageModel(
    name=request.form.get('name'),
    series_id=int(request.form.get('series_id')),
    type=request.form.get('type')
  )
  db.session.add(model)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/carriage_model/delete/<int:id>', methods=['POST'])
def delete_carriage_model(id):
  model = CarriageModel.query.get_or_404(id)
  if model.items:
    return f"该车厢型号正在被使用，无法删除！<script>setTimeout(()=>location.href='/options', 2000);</script>"
  db.session.delete(model)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/carriage_model/edit/<int:id>', methods=['GET', 'POST'])
def edit_carriage_model(id):
  """编辑车厢型号"""
  model = CarriageModel.query.get_or_404(id)
  carriage_series = CarriageSeries.query.all()

  if request.method == 'POST':
    model.name = request.form.get('name')
    model.series_id = int(request.form.get('series_id'))
    model.type = request.form.get('type')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_carriage_model.html',
    model=model,
    carriage_series=carriage_series
  )

@app.route('/options/trainset_series', methods=['POST'])
def add_trainset_series():
  series = TrainsetSeries(name=request.form.get('name'))
  db.session.add(series)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/trainset_series/delete/<int:id>', methods=['POST'])
def delete_trainset_series(id):
  series = TrainsetSeries.query.get_or_404(id)
  if series.trainsets or series.models:
    return f"该动车组系列正在被使用，无法删除！<script>setTimeout(()=>location.href='/options', 2000);</script>"
  db.session.delete(series)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/trainset_series/edit/<int:id>', methods=['GET', 'POST'])
def edit_trainset_series(id):
  """编辑动车组系列"""
  series = TrainsetSeries.query.get_or_404(id)

  if request.method == 'POST':
    series.name = request.form.get('name')
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_simple.html',
    item=series,
    title='编辑动车组系列',
    action_url=url_for('edit_trainset_series', id=id)
  )

@app.route('/options/trainset_model', methods=['POST'])
def add_trainset_model():
  model = TrainsetModel(
    name=request.form.get('name'),
    series_id=int(request.form.get('series_id')),
    power_type_id=int(request.form.get('power_type_id'))
  )
  db.session.add(model)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/trainset_model/delete/<int:id>', methods=['POST'])
def delete_trainset_model(id):
  model = TrainsetModel.query.get_or_404(id)
  if model.trainsets or model.locomotive_heads:
    return f"该动车组车型正在被使用，无法删除！<script>setTimeout(()=>location.href='/options', 2000);</script>"
  db.session.delete(model)
  db.session.commit()
  return redirect(url_for('options'))

@app.route('/options/trainset_model/edit/<int:id>', methods=['GET', 'POST'])
def edit_trainset_model(id):
  """编辑动车组车型"""
  model = TrainsetModel.query.get_or_404(id)
  trainset_series = TrainsetSeries.query.all()
  power_types = PowerType.query.all()

  if request.method == 'POST':
    model.name = request.form.get('name')
    model.series_id = int(request.form.get('series_id'))
    model.power_type_id = int(request.form.get('power_type_id'))
    db.session.commit()
    return redirect(url_for('options'))

  return render_template('option_edit_trainset_model.html',
    model=model,
    trainset_series=trainset_series,
    power_types=power_types
  )

# Auto-fill API
@app.route('/api/auto-fill/locomotive/<int:model_id>')
def auto_fill_locomotive(model_id):
  """机车车型自动填充"""
  model = LocomotiveModel.query.get_or_404(model_id)
  return jsonify({
    'series_id': model.series_id,
    'power_type_id': model.power_type_id
  })

@app.route('/api/auto-fill/carriage/<int:model_id>')
def auto_fill_carriage(model_id):
  """车厢车型自动填充"""
  model = CarriageModel.query.get_or_404(model_id)
  return jsonify({
    'series_id': model.series_id,
    'type': model.type
  })

@app.route('/api/auto-fill/trainset/<int:model_id>')
def auto_fill_trainset(model_id):
  """动车组车型自动填充"""
  model = TrainsetModel.query.get_or_404(model_id)
  return jsonify({
    'series_id': model.series_id,
    'power_type_id': model.power_type_id
  })

# 选项维护 API 路由（用于行内编辑）
@app.route('/api/options/<string:type>/edit', methods=['POST'])
def edit_option_api(type):
  """选项编辑 API"""
  id = request.form.get('id')

  if type == 'power_type':
    item = PowerType.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'brand':
    item = Brand.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'merchant':
    item = Merchant.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'depot':
    item = Depot.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'chip_interface':
    item = ChipInterface.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'chip_model':
    item = ChipModel.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'locomotive_series':
    item = LocomotiveSeries.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'carriage_series':
    item = CarriageSeries.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'trainset_series':
    item = TrainsetSeries.query.get_or_404(id)
    item.name = request.form.get('name')
  elif type == 'locomotive_model':
    item = LocomotiveModel.query.get_or_404(id)
    item.name = request.form.get('name')
    series_id = request.form.get('series_id')
    power_type_id = request.form.get('power_type_id')
    if series_id:
      item.series_id = int(series_id)
    if power_type_id:
      item.power_type_id = int(power_type_id)
  elif type == 'carriage_model':
    item = CarriageModel.query.get_or_404(id)
    item.name = request.form.get('name')
    series_id = request.form.get('series_id')
    type_value = request.form.get('type')
    if series_id:
      item.series_id = int(series_id)
    if type_value:
      item.type = type_value
  elif type == 'trainset_model':
    item = TrainsetModel.query.get_or_404(id)
    item.name = request.form.get('name')
    series_id = request.form.get('series_id')
    power_type_id = request.form.get('power_type_id')
    if series_id:
      item.series_id = int(series_id)
    if power_type_id:
      item.power_type_id = int(power_type_id)
  else:
    return jsonify({'success': False, 'error': '未知类型'}), 400

  try:
    db.session.commit()
    return jsonify({'success': True})
  except Exception as e:
    db.session.rollback()
    logger.error(f"Error editing {type}: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': str(e)}), 500

# 重新初始化
@app.route('/options/reinit', methods=['POST'])
def reinit_database():
  """重新初始化数据库：删除所有数据并重新加载预置数据"""
  # 删除所有数据（按外键依赖顺序）
  CarriageItem.query.delete()
  Locomotive.query.delete()
  CarriageSet.query.delete()
  Trainset.query.delete()
  LocomotiveHead.query.delete()
  LocomotiveModel.query.delete()
  TrainsetModel.query.delete()
  CarriageModel.query.delete()
  LocomotiveSeries.query.delete()
  CarriageSeries.query.delete()
  TrainsetSeries.query.delete()
  ChipModel.query.delete()
  ChipInterface.query.delete()
  Depot.query.delete()
  Merchant.query.delete()
  Brand.query.delete()
  PowerType.query.delete()
  db.session.commit()

  # 运行初始化脚本
  subprocess.run(['python', 'init_db.py'], check=True)

  logger.info("Database reinitialized successfully")
  return redirect(url_for('options'))

@app.route('/api/import/excel', methods=['POST'])
def import_from_excel():
  """从 Excel 文件导入数据"""
  try:
    if 'file' not in request.files:
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    file = request.files['file']
    if file.filename == '':
      return jsonify({'success': False, 'error': '未选择文件'}), 400

    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
      return jsonify({'success': False, 'error': '文件格式错误，请上传Excel文件'}), 400

    # 读取 Excel 文件
    workbook = openpyxl.load_workbook(file)
    summary = {}
    errors = []

    # 处理每个工作表
    for sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
      data = []

      # 读取表头（第一行）
      headers = None
      for row_idx, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if row_idx == 1:
          headers = [cell for cell in row if cell is not None]
        else:
          if any(cell is not None for cell in row):
            data.append(dict(zip(headers, row)))

      if not data:
        continue

      # 根据工作表名导入到对应表
      try:
        if sheet_name == '机车':
          count = import_locomotive_data(data)
          summary['机车模型'] = count
        elif sheet_name == '车厢':
          count = import_carriage_data(data)
          summary['车厢模型'] = count
        elif sheet_name == '动车组':
          count = import_trainset_data(data)
          summary['动车组模型'] = count
        elif sheet_name == '先头车':
          count = import_locomotive_head_data(data)
          summary['先头车模型'] = count
        else:
          logger.warning(f"Unknown sheet name: {sheet_name}")
      except Exception as e:
        errors.append(f"{sheet_name}: {str(e)}")
        logger.error(f"Error importing sheet {sheet_name}: {str(e)}", exc_info=True)

    if errors:
      return jsonify({'success': False, 'error': '部分导入失败: ' + '; '.join(errors)}), 400

    logger.info(f"Excel import completed: {summary}")
    return jsonify({'success': True, 'summary': summary})

  except Exception as e:
    db.session.rollback()
    logger.error(f"Excel import failed: {str(e)}", exc_info=True)
    return jsonify({'success': False, 'error': str(e)}), 500


def import_locomotive_data(data):
  """导入机车模型数据"""
  count = 0
  for row in data:
    # 查找关联对象的ID
    series_id = find_id_by_name(LocomotiveSeries, row.get('系列'))
    power_type_id = find_id_by_name(PowerType, row.get('动力类型'))
    model_id = find_id_by_name(LocomotiveModel, row.get('车型'), lambda q: q.filter_by(name=row.get('车型'), series_id=series_id, power_type_id=power_type_id).first())
    brand_id = find_id_by_name(Brand, row.get('品牌'))
    depot_id = find_id_by_name(Depot, row.get('机务段'))
    chip_interface_id = find_id_by_name(ChipInterface, row.get('芯片接口'))
    chip_model_id = find_id_by_name(ChipModel, row.get('芯片型号'))
    merchant_id = find_id_by_name(Merchant, row.get('购买商家'))

    # 处理日期
    purchase_date = parse_date(row.get('购买日期'))

    # 计算总价
    total_price = calculate_total_price(row.get('价格'))

    locomotive = Locomotive(
      series_id=series_id,
      power_type_id=power_type_id,
      model_id=model_id,
      brand_id=brand_id,
      depot_id=depot_id,
      plaque=row.get('挂牌'),
      color=row.get('颜色'),
      scale=row.get('比例'),
      locomotive_number=row.get('机车号'),
      decoder_number=row.get('编号'),
      chip_interface_id=chip_interface_id,
      chip_model_id=chip_model_id,
      price=row.get('价格'),
      total_price=total_price,
      item_number=row.get('货号'),
      purchase_date=purchase_date,
      merchant_id=merchant_id
    )
    db.session.add(locomotive)
    count += 1

  db.session.commit()
  return count


def import_carriage_data(data):
  """导入车厢模型数据"""
  count = 0
  for row in data:
    brand_id = find_id_by_name(Brand, row.get('品牌'))
    series_id = find_id_by_name(CarriageSeries, row.get('系列'))
    depot_id = find_id_by_name(Depot, row.get('车辆段'))
    merchant_id = find_id_by_name(Merchant, row.get('购买商家'))

    purchase_date = parse_date(row.get('购买日期'))
    total_price = float(row.get('总价', 0)) if row.get('总价') else 0

    # 创建车厢套装
    carriage_set = CarriageSet(
      brand_id=brand_id,
      series_id=series_id,
      depot_id=depot_id,
      train_number=row.get('车次'),
      plaque=row.get('挂牌'),
      item_number=row.get('货号'),
      scale=row.get('比例'),
      total_price=total_price,
      purchase_date=purchase_date,
      merchant_id=merchant_id
    )
    db.session.add(carriage_set)
    db.session.flush()  # 获取ID

    # 处理车厢项（如果Excel中有详细数据）
    # 简化版：这里假设每行是一个车厢套装
    # 如需要支持多车厢，需要Excel格式调整

    count += 1

  db.session.commit()
  return count


def import_trainset_data(data):
  """导入动车组模型数据"""
  count = 0
  for row in data:
    series_id = find_id_by_name(TrainsetSeries, row.get('系列'))
    power_type_id = find_id_by_name(PowerType, row.get('动力类型'))
    model_id = find_id_by_name(TrainsetModel, row.get('车型'), lambda q: q.filter_by(name=row.get('车型'), series_id=series_id, power_type_id=power_type_id).first())
    brand_id = find_id_by_name(Brand, row.get('品牌'))
    depot_id = find_id_by_name(Depot, row.get('动车段'))
    chip_interface_id = find_id_by_name(ChipInterface, row.get('芯片接口'))
    chip_model_id = find_id_by_name(ChipModel, row.get('芯片型号'))
    merchant_id = find_id_by_name(Merchant, row.get('购买商家'))

    purchase_date = parse_date(row.get('购买日期'))
    total_price = calculate_total_price(row.get('价格'))

    trainset = Trainset(
      series_id=series_id,
      power_type_id=power_type_id,
      model_id=model_id,
      brand_id=brand_id,
      depot_id=depot_id,
      plaque=row.get('挂牌'),
      color=row.get('颜色'),
      scale=row.get('比例'),
      formation=int(row.get('编组')) if row.get('编组') else None,
      trainset_number=row.get('动车号'),
      decoder_number=row.get('编号'),
      head_light=parse_boolean(row.get('头车灯')),
      interior_light=row.get('室内灯'),
      chip_interface_id=chip_interface_id,
      chip_model_id=chip_model_id,
      price=row.get('价格'),
      total_price=total_price,
      item_number=row.get('货号'),
      purchase_date=purchase_date,
      merchant_id=merchant_id
    )
    db.session.add(trainset)
    count += 1

  db.session.commit()
  return count


def import_locomotive_head_data(data):
  """导入先头车模型数据"""
  count = 0
  for row in data:
    model_id = find_id_by_name(TrainsetModel, row.get('车型'))
    brand_id = find_id_by_name(Brand, row.get('品牌'))
    depot_id = find_id_by_name(Depot, row.get('动车段'))
    merchant_id = find_id_by_name(Merchant, row.get('购买商家'))

    purchase_date = parse_date(row.get('购买日期'))
    total_price = calculate_total_price(row.get('价格'))

    locomotive_head = LocomotiveHead(
      model_id=model_id,
      brand_id=brand_id,
      depot_id=depot_id,
      special_color=row.get('特涂'),
      scale=row.get('比例'),
      head_light=parse_boolean(row.get('头车灯')),
      interior_light=row.get('室内灯'),
      price=row.get('价格'),
      total_price=total_price,
      item_number=row.get('货号'),
      purchase_date=purchase_date,
      merchant_id=merchant_id
    )
    db.session.add(locomotive_head)
    count += 1

  db.session.commit()
  return count


def find_id_by_name(model, name, custom_query=None):
  """根据名称查找模型的ID"""
  if not name:
    return None

  if custom_query:
    return custom_query(model.query).id if custom_query(model.query) else None

  obj = model.query.filter_by(name=name).first()
  return obj.id if obj else None


def parse_date(date_str):
  """解析日期字符串"""
  if not date_str:
    return date.today()

  try:
    if isinstance(date_str, datetime):
      return date_str.date()
    return datetime.strptime(str(date_str), '%Y-%m-%d').date()
  except:
    return date.today()


def parse_boolean(value):
  """解析布尔值"""
  if not value:
    return None
  if isinstance(value, bool):
    return value
  return str(value).lower() in ('true', '1', '是', '有', 'yes')


def calculate_total_price(price_str):
  """计算总价（支持表达式）"""
  if not price_str:
    return 0

  try:
    # 使用现有的 SafeEval 逻辑
    tree = ast.parse(str(price_str), mode='eval')

    class SafeEval(ast.NodeVisitor):
      allowed_nodes = (ast.Expression, ast.BinOp, ast.UnaryOp, ast.Constant)

      def visit(self, node):
        if not isinstance(node, self.allowed_nodes):
          raise ValueError(f"Invalid node: {type(node).__name__}")
        return super().visit(node)

      def generic_visit(self, node):
        if isinstance(node, ast.BinOp):
          return self.visit(node.left) + self.visit(node.right)
        elif isinstance(node, ast.UnaryOp):
          return -self.visit(node.operand)
        elif isinstance(node, ast.Constant):
          return node.value
        return super().generic_visit(node)

    return SafeEval().visit(tree)
  except:
    return float(price_str) if str(price_str).replace('.', '').isdigit() else 0


@app.route('/api/export/excel')
def export_to_excel():
  """导出数据到Excel"""
  try:
    from flask import send_file
    from io import BytesIO
    from datetime import datetime

    # 检查是否有任何数据
    has_data = (
      Locomotive.query.count() > 0 or
      CarriageSet.query.count() > 0 or
      Trainset.query.count() > 0 or
      LocomotiveHead.query.count() > 0
    )

    if not has_data:
      return jsonify({'success': False, 'error': '当前没有可导出的数据，请先添加模型后再导出'}), 400

    # 创建工作簿
    workbook = openpyxl.Workbook()

    # 移除默认的工作表
    if 'Sheet' in workbook.sheetnames:
      workbook.remove(workbook['Sheet'])

    # 导出机车模型
    if Locomotive.query.count() > 0:
      sheet = workbook.create_sheet('机车')
      headers = ['系列', '动力类型', '车型', '品牌', '机务段', '挂牌', '颜色', '比例', '机车号', '编号',
                 '芯片接口', '芯片型号', '价格', '总价', '货号', '购买日期', '购买商家']
      sheet.append(headers)

      for locomotive in Locomotive.query.all():
        row = [
          locomotive.series.name if locomotive.series else '',
          locomotive.power_type.name if locomotive.power_type else '',
          locomotive.model.name if locomotive.model else '',
          locomotive.brand.name if locomotive.brand else '',
          locomotive.depot.name if locomotive.depot else '',
          locomotive.plaque or '',
          locomotive.color or '',
          locomotive.scale or '',
          locomotive.locomotive_number or '',
          locomotive.decoder_number or '',
          locomotive.chip_interface.name if locomotive.chip_interface else '',
          locomotive.chip_model.name if locomotive.chip_model else '',
          locomotive.price or '',
          locomotive.total_price or '',
          locomotive.item_number or '',
          locomotive.purchase_date.strftime('%Y-%m-%d') if locomotive.purchase_date else '',
          locomotive.merchant.name if locomotive.merchant else ''
        ]
        sheet.append(row)

    # 导出车厢模型（按每节车厢一行）
    if CarriageSet.query.count() > 0:
      sheet = workbook.create_sheet('车厢')
      headers = ['品牌', '系列', '车辆段', '车次', '挂牌', '货号', '比例', '车型', '车辆号', '颜色', '灯光', '总价', '购买日期', '购买商家']
      sheet.append(headers)

      row_offset = 2  # 从第2行开始（第1行是表头）
      for carriage_set in CarriageSet.query.all():
        items = carriage_set.items
        if not items:
          # 没有车厢项，添加一行套装信息
          row = [
            carriage_set.brand.name if carriage_set.brand else '',
            carriage_set.series.name if carriage_set.series else '',
            carriage_set.depot.name if carriage_set.depot else '',
            carriage_set.train_number or '',
            carriage_set.plaque or '',
            carriage_set.item_number or '',
            carriage_set.scale or '',
            '', '', '', '',  # 车型、车辆号、颜色、灯光为空
            carriage_set.total_price or '',
            carriage_set.purchase_date.strftime('%Y-%m-%d') if carriage_set.purchase_date else '',
            carriage_set.merchant.name if carriage_set.merchant else ''
          ]
          sheet.append(row)
        else:
          # 有车厢项，为每节车厢添加一行，合并套装字段
          set_row_start = row_offset
          for item in items:
            row = [
              carriage_set.brand.name if carriage_set.brand else '',
              carriage_set.series.name if carriage_set.series else '',
              carriage_set.depot.name if carriage_set.depot else '',
              carriage_set.train_number or '',
              carriage_set.plaque or '',
              carriage_set.item_number or '',
              carriage_set.scale or '',
              item.model.name if item.model else '',
              item.car_number or '',
              item.color or '',
              item.lighting or '',
              carriage_set.total_price or '',
              carriage_set.purchase_date.strftime('%Y-%m-%d') if carriage_set.purchase_date else '',
              carriage_set.merchant.name if carriage_set.merchant else ''
            ]
            sheet.append(row)
            row_offset += 1

          # 合并套装公共字段单元格（品牌、系列、车辆段、车次、挂牌、货号、比例、总价、购买日期、购买商家）
          set_row_end = row_offset - 1
          merge_columns = [1, 2, 3, 4, 5, 6, 7, 12, 13, 14]  # 1-based column indices
          for col in merge_columns:
            if set_row_start != set_row_end:
              sheet.merge_cells(start_row=set_row_start, start_column=col,
                           end_row=set_row_end, end_column=col)

    # 导出动车组模型
    if Trainset.query.count() > 0:
      sheet = workbook.create_sheet('动车组')
      headers = ['系列', '动力类型', '车型', '品牌', '动车段', '挂牌', '颜色', '比例', '编组', '动车号', '编号',
                 '头车灯', '室内灯', '芯片接口', '芯片型号', '价格', '总价', '货号', '购买日期', '购买商家']
      sheet.append(headers)

      for trainset in Trainset.query.all():
        row = [
          trainset.series.name if trainset.series else '',
          trainset.power_type.name if trainset.power_type else '',
          trainset.model.name if trainset.model else '',
          trainset.brand.name if trainset.brand else '',
          trainset.depot.name if trainset.depot else '',
          trainset.plaque or '',
          trainset.color or '',
          trainset.scale or '',
          trainset.formation or '',
          trainset.trainset_number or '',
          trainset.decoder_number or '',
          '是' if trainset.head_light else '否',
          trainset.interior_light or '',
          trainset.chip_interface.name if trainset.chip_interface else '',
          trainset.chip_model.name if trainset.chip_model else '',
          trainset.price or '',
          trainset.total_price or '',
          trainset.item_number or '',
          trainset.purchase_date.strftime('%Y-%m-%d') if trainset.purchase_date else '',
          trainset.merchant.name if trainset.merchant else ''
        ]
        sheet.append(row)

    # 导出先头车模型
    if LocomotiveHead.query.count() > 0:
      sheet = workbook.create_sheet('先头车')
      headers = ['车型', '品牌', '动车段', '特涂', '比例', '头车灯', '室内灯', '价格', '总价', '货号', '购买日期', '购买商家']
      sheet.append(headers)

      for head in LocomotiveHead.query.all():
        row = [
          head.model.name if head.model else '',
          head.brand.name if head.brand else '',
          head.depot.name if head.depot else '',
          head.special_color or '',
          head.scale or '',
          '是' if head.head_light else '否',
          head.interior_light or '',
          head.price or '',
          head.total_price or '',
          head.item_number or '',
          head.purchase_date.strftime('%Y-%m-%d') if head.purchase_date else '',
          head.merchant.name if head.merchant else ''
        ]
        sheet.append(row)

    # 生成文件
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # 生成文件名
    filename = f'火车模型数据导出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    logger.info("Excel export completed successfully")
    return send_file(
      output,
      as_attachment=True,
      download_name=filename,
      mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

  except Exception as e:
    logger.error(f"Excel export failed: {str(e)}", exc_info=True)
    return f"导出失败: {str(e)}", 500

# 错误处理器
@app.errorhandler(404)
def not_found(error):
  """404 错误处理"""
  return render_template('404.html'), 404

@app.errorhandler(500)
def internal_error(error):
  """500 错误处理"""
  return render_template('500.html'), 500

@app.errorhandler(Exception)
def handle_exception(error):
  """全局异常处理"""
  db.session.rollback()
  logger.error(f"Unhandled exception: {str(error)}", exc_info=True)
  return f"服务器错误: {str(error)}<script>setTimeout(()=>location.href='/', 3000);</script>", 500

if __name__ == '__main__':
  app.run(debug=True)
